# -*- coding: utf-8 -*-
from __future__ import annotations
import base64
import copy
import csv
import hashlib
import hmac
import io
import json
import os
import re
import time
import threading
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import Any, Dict, Iterable, List, Tuple, Optional
from urllib.parse import urlparse
import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from flask import Flask, Response, jsonify, render_template, request, send_file
from werkzeug.exceptions import HTTPException
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TEMPLATES_DIR = os.path.join(BASE_DIR, "templates")
SAMPLES_DIR = os.path.join(BASE_DIR, "samples")
OPENAPI_BASE_URL = "https://api.searchad.naver.com"
app = Flask(__name__, template_folder=TEMPLATES_DIR)
HTTP_SESSION = requests.Session()
HTTP_ADAPTER = requests.adapters.HTTPAdapter(pool_connections=32, pool_maxsize=64, max_retries=0)
HTTP_SESSION.mount("https://", HTTP_ADAPTER)
HTTP_SESSION.mount("http://", HTTP_ADAPTER)
_CACHE_LOCK = threading.RLock()
_CACHE_TTL_SECONDS = 10.0
_CAMPAIGN_CACHE: Dict[str, Tuple[float, List[Dict[str, Any]]]] = {}
_ADGROUP_CACHE: Dict[str, Tuple[float, List[Dict[str, Any]]]] = {}
_CHANNEL_CACHE: Dict[str, Tuple[float, List[Dict[str, Any]]]] = {}
FAST_IO_WORKERS = 8
DELETE_IO_WORKERS = 12
BID_IO_WORKERS = 8
LOG_DIR = os.path.join(BASE_DIR, "logs")
os.makedirs(LOG_DIR, exist_ok=True)
ACTION_LOG_PATH = os.path.join(LOG_DIR, "action_history.jsonl")
_ACTION_LOG_LOCK = threading.RLock()
_ACTION_LOG_MAX_LINES = 2000
def _stable_cache_key(api_key: str, secret_key: str, cid: str, scope: str) -> str:
    raw = f"{scope}::{str(api_key or '').strip()}::{str(secret_key or '').strip()}::{str(cid or '').strip()}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()
def _cache_get(store: Dict[str, Tuple[float, Any]], key: str, ttl: float = _CACHE_TTL_SECONDS):
    now = time.time()
    with _CACHE_LOCK:
        item = store.get(key)
        if not item:
            return None
        ts, value = item
        if now - ts > ttl:
            store.pop(key, None)
            return None
        return copy.deepcopy(value)
def _cache_set(store: Dict[str, Tuple[float, Any]], key: str, value: Any):
    with _CACHE_LOCK:
        store[key] = (time.time(), copy.deepcopy(value))
def _cache_invalidate(api_key: str, secret_key: str, cid: str):
    camp_key = _stable_cache_key(api_key, secret_key, cid, "campaigns")
    ch_key = _stable_cache_key(api_key, secret_key, cid, "channels")
    with _CACHE_LOCK:
        _CAMPAIGN_CACHE.pop(camp_key, None)
        _CHANNEL_CACHE.pop(ch_key, None)
        _ADGROUP_CACHE.clear()
LOG_ACTION_LABELS = {
    "/create_campaign": "캠페인 생성",
    "/create_adgroup_simple": "광고그룹 생성",
    "/create_keywords_simple": "키워드 등록",
    "/bulk_upload_text_ads": "텍스트 소재 일괄 업로드",
    "/create_text_ad_simple": "텍스트 소재 생성",
    "/create_ad_advanced": "고급 소재 생성",
    "/create_extension_simple": "확장소재 등록",
    "/bulk_upload_headlines": "추가제목 일괄 업로드",
    "/create_shopping_ad_simple": "쇼핑 소재 생성",
    "/create_extension_raw": "확장소재 원본 등록",
    "/create_restricted_keywords_simple": "제외키워드 등록",
    "/copy_entities_to_adgroups": "엔티티 복사",
    "/copy_campaigns": "캠페인 복사",
    "/copy_adgroups_to_target": "광고그룹 복사",
    "/rename_adgroups_bulk": "광고그룹명 일괄 변경",
    "/update_media": "매체 설정 변경",
    "/update_adgroup_options": "광고그룹 옵션 변경",
    "/update_budget": "예산 변경",
    "/update_schedule": "시간대 설정 변경",
    "/update_schedule_campaign_bulk": "캠페인 시간대 일괄 변경",
    "/update_non_search_keyword_exclusion": "비검색영역 노출제외 변경",
    "/update_keyword_bids": "키워드 입찰가 변경",
    "/update_bid_mode_by_scope": "입찰가 방식 변경",
    "/search_powerlink_keywords": "검색어 기준 키워드 조회",
    "/export_powerlink_keywords_excel": "검색어 기준 키워드 엑셀 다운로드",
    "/query_account_keywords": "계정 키워드 조회",
    "/export_account_keywords_excel": "계정 키워드 엑셀 다운로드",
    "/query_account_ads": "계정 소재 조회",
    "/query_account_extensions": "계정 확장소재 조회",
    "/export_account_ads_excel": "계정 소재 엑셀 다운로드",
    "/export_account_extensions_excel": "계정 확장소재 엑셀 다운로드",
    "/update_keyword_bids_by_search": "검색어 기준 키워드 입찰가 변경",
    "/update_keyword_bid_weights_by_search": "검색어 기준 키워드 입찰가중치 변경",
    "/set_searched_powerlink_keyword_state": "조회 키워드 ON/OFF 변경",
    "/adjust_keyword_bids_by_threshold": "상하한 기준 입찰가 조정",
    "/update_keyword_bids_avg_position": "평균순위 기준 입찰가 적용",
    "/bulk_delete_by_parent": "부모 기준 일괄 삭제",
    "/bulk_register": "일괄 등록",
    "/bulk_delete": "일괄 삭제",
    "/set_campaign_state": "캠페인 상태 변경",
    "/delete_selected": "선택 삭제",
    "/clear_action_logs": "로그 비우기",
}
ACTION_LOG_EXCLUDED_PATHS = {
    "/get_campaigns", "/get_adgroups", "/get_biz_channels", "/get_keywords", "/get_ads",
    "/get_ad_extensions", "/get_restricted_keywords", "/find_powerlink_duplicate_keywords", "/find_account_powerlink_duplicate_keywords", "/get_powerlink_keyword_stats", "/search_powerlink_keywords", "/export_powerlink_keywords_excel", "/query_account_ads", "/query_account_extensions", "/query_account_keywords", "/export_account_ads_excel", "/export_account_extensions_excel", "/export_account_keywords_excel", "/get_action_logs", "/health", "/favicon.ico", "/",
    "/sample_headers", "/delete_sample_headers", "/clear_action_logs",
}
def _safe_json_body() -> Dict[str, Any]:
    data = request.get_json(silent=True)
    if isinstance(data, dict):
        return data
    if request.form:
        out: Dict[str, Any] = {}
        for k in request.form.keys():
            vals = request.form.getlist(k)
            out[k] = vals if len(vals) > 1 else (vals[0] if vals else "")
        return out
    return {}
def _line_count(text_value: Any) -> int:
    return len([x for x in str(text_value or "").replace(",", "\n").splitlines() if x.strip()])
def _action_summary(path: str, payload: Dict[str, Any]) -> str:
    if not payload:
        return ""
    if path == "/create_campaign":
        return f"이름={str(payload.get('name') or '').strip()}"
    if path == "/create_adgroup_simple":
        return f"이름={str(payload.get('name') or '').strip()} | 캠페인={str(payload.get('campaign_id') or '').strip()}"
    if path == "/create_keywords_simple":
        if isinstance(payload.get('keyword_batches'), list):
            batches = payload.get('keyword_batches') or []
            slot_cnt = sum(1 for x in batches if str((x or {}).get('adgroup_id') or '').strip() and str((x or {}).get('keywords_text') or '').strip())
            kw_cnt = sum(_line_count((x or {}).get('keywords_text')) for x in batches)
            return f"슬롯={slot_cnt} | 키워드={kw_cnt}"
        return f"키워드={_line_count(payload.get('keywords_text'))}"
    if path in {"/copy_campaigns", "/copy_adgroups_to_target", "/copy_entities_to_adgroups"}:
        src = payload.get('source_ids') or payload.get('entity_ids') or []
        if not isinstance(src, list):
            src = [src] if src else []
        custom_cnt = _line_count(payload.get('custom_names_text'))
        pieces = [f"대상={len(src)}"]
        if str(payload.get('copy_count') or '').strip():
            pieces.append(f"복사수={payload.get('copy_count')}")
        if custom_cnt:
            pieces.append(f"지정명={custom_cnt}")
        tgt = str(payload.get('target_campaign_id') or '').strip()
        if tgt:
            pieces.append(f"타겟캠페인={tgt}")
        return " | ".join(pieces)
    if path == "/rename_adgroups_bulk":
        ids = payload.get('entity_ids') or []
        if not isinstance(ids, list):
            ids = [ids] if ids else []
        return f"그룹={len(ids)} | 새이름={_line_count(payload.get('target_names_text'))}"
    if path in {"/update_budget", "/set_campaign_state", "/update_media", "/update_adgroup_options", "/update_schedule", "/update_schedule_campaign_bulk", "/update_non_search_keyword_exclusion", "/update_keyword_bids", "/update_bid_mode_by_scope", "/update_keyword_bids_by_search", "/update_keyword_bid_weights_by_search", "/adjust_keyword_bids_by_threshold", "/update_keyword_bids_avg_position"}:
        ids = payload.get('entity_ids') or payload.get('campaign_ids') or payload.get('adgroup_ids') or []
        if not isinstance(ids, list):
            ids = [ids] if ids else []
        label = str(payload.get('entity_type') or payload.get('media_type') or '').strip()
        extra = []
        if label:
            extra.append(label)
        if ids:
            extra.append(f"건수={len(ids)}")
        if str(payload.get('search_text') or '').strip():
            match_label = "완전일치" if _boolish(payload.get('exact_match'), False) else "부분일치"
            extra.append(f"검색={str(payload.get('search_text') or '').strip()} ({match_label})")
        if str(payload.get('bid_amt') or '').strip():
            extra.append(f"입찰가={payload.get('bid_amt')}")
        if str(payload.get('bid_weight') or payload.get('bidWeight') or '').strip():
            extra.append(f"입찰가중치={payload.get('bid_weight') or payload.get('bidWeight')}%")
        return " | ".join(extra)
    if path in {"/bulk_delete", "/delete_selected"}:
        rows = payload.get('rows') or payload.get('ids') or []
        if not isinstance(rows, list):
            rows = [rows] if rows else []
        return f"유형={str(payload.get('entity_type') or '').strip()} | 건수={len(rows)}"
    if path == "/bulk_delete_by_parent":
        ids = payload.get('parent_ids') or []
        if not isinstance(ids, list):
            ids = [ids] if ids else []
        return f"범위={str(payload.get('parent_type') or '').strip()} | 대상={str(payload.get('target_entity') or '').strip()} | 부모={len(ids)}"
    if path == "/bulk_register":
        rows = payload.get('rows') or []
        if not isinstance(rows, list):
            rows = [rows] if rows else []
        return f"유형={str(payload.get('entity_type') or '').strip()} | 건수={len(rows)}"
    if path == "/create_restricted_keywords_simple":
        return f"제외키워드={_line_count(payload.get('keywords_text'))}"
    if path in {"/bulk_upload_text_ads", "/bulk_upload_headlines"}:
        file_names = []
        try:
            file_names = [f.filename for f in request.files.values() if getattr(f, 'filename', '')]
        except Exception:
            file_names = []
        return f"파일={', '.join(file_names[:2])}" if file_names else "파일 업로드"
    count_keys = ('ids', 'entity_ids', 'source_ids', 'campaign_ids', 'adgroup_ids', 'parent_ids', 'rows')
    for key in count_keys:
        val = payload.get(key)
        if isinstance(val, list) and val:
            return f"건수={len(val)}"
    return ""
def _extract_response_message(response: Response) -> str:
    try:
        if response.is_json:
            data = response.get_json(silent=True) or {}
            if isinstance(data, dict):
                return str(data.get('message') or data.get('error') or data.get('details') or '').strip()
    except Exception:
        pass
    return ""
def _prune_action_log_file():
    try:
        if not os.path.exists(ACTION_LOG_PATH):
            return
        with open(ACTION_LOG_PATH, 'r', encoding='utf-8') as fp:
            lines = fp.readlines()
        if len(lines) <= _ACTION_LOG_MAX_LINES:
            return
        keep = lines[-_ACTION_LOG_MAX_LINES:]
        with open(ACTION_LOG_PATH, 'w', encoding='utf-8') as fp:
            fp.writelines(keep)
    except Exception:
        pass
def _append_action_log(entry: Dict[str, Any]):
    line = json.dumps(entry, ensure_ascii=False)
    with _ACTION_LOG_LOCK:
        with open(ACTION_LOG_PATH, 'a', encoding='utf-8') as fp:
            fp.write(line + "\n")
        _prune_action_log_file()
def _read_action_logs(limit: int = 150) -> List[Dict[str, Any]]:
    if limit <= 0:
        return []
    with _ACTION_LOG_LOCK:
        if not os.path.exists(ACTION_LOG_PATH):
            return []
        with open(ACTION_LOG_PATH, 'r', encoding='utf-8') as fp:
            lines = fp.readlines()[-limit:]
    out: List[Dict[str, Any]] = []
    for line in reversed(lines):
        try:
            item = json.loads(line)
            if isinstance(item, dict):
                out.append(item)
        except Exception:
            continue
    return out
def _write_action_log_from_request(response: Response) -> None:
    path = request.path or ''
    if request.method != 'POST' or path in ACTION_LOG_EXCLUDED_PATHS:
        return
    try:
        payload = _safe_json_body()
        status = 'success' if 200 <= int(response.status_code) < 400 else 'error'
        customer_id = str(payload.get('customer_id') or payload.get('customerId') or '').strip()
        entry = {
            'ts': time.strftime('%Y-%m-%d %H:%M:%S'),
            'path': path,
            'action': LOG_ACTION_LABELS.get(path, path.strip('/') or '요청'),
            'status': status,
            'http_status': int(response.status_code),
            'customer_id': customer_id,
            'summary': _action_summary(path, payload),
            'message': _extract_response_message(response),
        }
        _append_action_log(entry)
    except Exception:
        pass
DAY_NUM_TO_CODE = {1: "MON", 2: "TUE", 3: "WED", 4: "THU", 5: "FRI", 6: "SAT", 7: "SUN"}
SHOPPING_AD_TYPES = {"SHOPPING_PRODUCT_AD", "CATALOG_PRODUCT_AD", "CATALOG_AD", "SHOPPING_BRAND_AD"}
SHOPPING_TARGETABLE_ADGROUP_TYPES = {"SHOPPING", "CATALOG", "SHOPPING_BRAND"}
SHOPPING_ITEM_BID_AD_TYPES = {"SHOPPING_PRODUCT_AD", "CATALOG_PRODUCT_AD", "CATALOG_AD"}
SHOPPING_ADGROUP_TYPES_WITH_AD_LEVEL_BID = {"SHOPPING", "CATALOG"}
CAMPAIGN_TYPE_LABELS = {
    "WEB_SITE": "파워링크",
    "SHOPPING": "쇼핑검색",
}
CAMPAIGN_TYPE_COLORS = {
    "WEB_SITE": "blue",
    "SHOPPING": "green",
}
ADGROUP_TYPE_LABELS = {
    "WEB_SITE": "파워링크",
    "SHOPPING": "쇼핑검색",
    "SHOPPING_BRAND": "쇼핑브랜드",
    "CATALOG": "카탈로그",
}
AD_TYPE_LABELS = {
    "TEXT_45": "기본소재(제목+설명)",
    "RSA_AD": "반응형 검색소재",
    "SHOPPING_PRODUCT_AD": "상품소재",
    "CATALOG_PRODUCT_AD": "카탈로그 상품소재",
    "CATALOG_AD": "카탈로그 소재",
    "SHOPPING_BRAND_AD": "쇼핑브랜드 소재",
}
AD_EXTENSION_TYPE_LABELS = {
    "HEADLINE": "추가 제목",
    "SUB_LINKS": "서브링크",
    "DESCRIPTION": "추가 설명문구",
    "DESCRIPTION_EXTRA": "설명 확장문구",
    "SHOPPING_PROMO_TEXT": "쇼핑 추가홍보문구",
    "SHOPPING_EXTRA": "쇼핑상품부가정보",
    "PHONE": "전화번호",
    "LOCATION": "위치정보",
    "PROMOTION": "프로모션",
    "PRICE_LINKS": "가격링크",
    "POWER_LINK_IMAGE": "파워링크 이미지",
    "WEBSITE_INFO": "웹사이트 정보",
    "IMAGE_SUB_LINKS": "이미지 서브링크",
}
COMMON_EXTENSION_TYPES = [
    "HEADLINE", "SUB_LINKS", "DESCRIPTION", "DESCRIPTION_EXTRA", "PHONE",
    "LOCATION", "PROMOTION", "PRICE_LINKS", "POWER_LINK_IMAGE", "WEBSITE_INFO", "IMAGE_SUB_LINKS"
]
SHOPPING_EXTRA_AD_TYPES = {"SHOPPING_PRODUCT_AD"}
SYSTEM_FIELDS = {
    "nccAdId", "adExtensionId", "nccKeywordId", "regTm", "editTm", "status", "statusReason",
    "inspectStatus", "delFlag", "managedKeyword", "referenceData", "referenceKeyData", "nccQi",
}
BOOL_FIELDS = {
    "useDailyBudget", "useGroupBidAmt", "userLock", "enable", "keywordPlusFlag", "useStoreUrl",
    "paused", "mobilePreferred", "descriptionPin", "headlinePin",
}
INT_FIELDS = {
    "customerId", "dailyBudget", "bidAmt", "contentsNetworkBidAmt", "priority", "displayOrder",
    "price", "discountPrice", "extraCost", "bidWeight",
}
FLOAT_FIELDS = {"pcCtr", "mobileCtr"}
ENTITY_SAMPLE_HEADERS: Dict[str, List[str]] = {
    "campaign": ["캠페인명", "캠페인유형", "일예산사용", "일예산"],
    "adgroup": ["캠페인ID", "광고그룹명", "광고그룹유형", "일예산사용", "일예산", "입찰가", "비즈채널ID"],
    "keyword": ["광고그룹ID", "키워드", "그룹입찰가사용", "입찰가", "사용자잠금"],
    "ad": ["광고그룹ID", "소재유형", "제목", "설명", "PC랜딩URL", "모바일랜딩URL", "사용자잠금"],
    "ad_extension": ["소유ID", "확장소재유형", "원본JSON"],
    "restricted_keyword": ["광고그룹ID", "제외키워드"],
}
DELETE_SAMPLE_HEADERS: Dict[str, List[str]] = {
    "campaign": ["캠페인ID"],
    "adgroup": ["광고그룹ID"],
    "keyword": ["키워드ID", "광고그룹ID", "키워드"],
    "ad": ["소재ID"],
    "ad_extension": ["확장소재ID"],
    "restricted_keyword": ["광고그룹ID", "제외키워드"],
}
KO_HEADER_ALIASES: Dict[str, str] = {
    "고객아이디": "customerId", "고객id": "customerId",
    "캠페인id": "nccCampaignId", "캠페인아이디": "nccCampaignId",
    "광고그룹id": "nccAdgroupId", "광고그룹아이디": "nccAdgroupId",
    "키워드id": "nccKeywordId", "키워드아이디": "nccKeywordId",
    "비즈채널id": "nccBusinessChannelId", "비즈채널아이디": "nccBusinessChannelId",
    "비즈채널": "nccBusinessChannelId",
    "소재id": "nccAdId", "소재아이디": "nccAdId",
    "확장소재id": "adExtensionId", "확장소재아이디": "adExtensionId",
    "소유id": "ownerId", "소유아이디": "ownerId",
    "캠페인명": "name", "캠페인유형": "campaignTp", "일예산사용": "useDailyBudget", "일예산": "dailyBudget",
    "광고그룹명": "name", "광고그룹유형": "adgroupType", "입찰가": "bidAmt",
    "키워드": "keyword", "제외키워드": "keyword", "그룹입찰가사용": "useGroupBidAmt", "사용자잠금": "userLock",
    "소재유형": "type", "제목": "headline", "설명": "description", "pc랜딩url": "pcFinalUrl",
    "모바일랜딩url": "mobileFinalUrl", "원본json": "rawJson",
    "확장소재유형": "type",
}
@app.errorhandler(Exception)
def handle_exception(e):
    if isinstance(e, HTTPException):
        return jsonify({"error": e.description or str(e)}), int(getattr(e, "code", 500) or 500)
    return jsonify({"error": f"서버 내부 오류: {str(e)}"}), 500
@app.after_request
def after_request_write_action_log(response):
    _write_action_log_from_request(response)
    return response
@app.route("/favicon.ico")
def favicon():
    return Response(status=204)
def _display_url_from_final(final_url: str) -> str:
    s = str(final_url or "").strip()
    if not s:
        return s
    try:
        parsed = urlparse(s)
        if parsed.scheme and parsed.netloc:
            return f"{parsed.scheme}://{parsed.netloc}"
    except Exception:
        pass
    return s
def _label_negative_type(value: Any) -> str:
    s = str(value or "").strip().upper()
    if s in {"2", "PHRASE", "구문"}:
        return "구문"
    if s in {"1", "EXACT", "일치"}:
        return "일치"
    if s in {"KEYWORD_PLUS_RESTRICT"}:
        return "일치"
    return s or "-"
def _normalize_negative_type(value: Any) -> int:
    s = str(value or "").strip().upper()
    if s in {"2", "PHRASE", "구문"}:
        return 2
    return 1
def _sig(ts: str, method: str, uri: str, secret_key: str) -> str:
    msg = f"{ts}.{method.upper()}.{uri}"
    dig = hmac.new(str(secret_key).strip().encode("utf-8"), msg.encode("utf-8"), hashlib.sha256).digest()
    return base64.b64encode(dig).decode()
def _open_headers(api_key: str, secret_key: str, customer_id: str, method: str, uri: str) -> dict:
    ts = str(int(time.time() * 1000))
    return {
        "X-Timestamp": ts,
        "X-API-KEY": str(api_key).strip(),
        "X-Customer": str(customer_id).strip(),
        "X-Signature": _sig(ts, method, uri, secret_key),
        "Content-Type": "application/json; charset=UTF-8",
    }
def _make_fake_response(status_code: int, text: str):
    class FakeResponse:
        def __init__(self, status_code: int, text: str):
            self.status_code = status_code
            self.text = text
            self.content = text.encode("utf-8", errors="ignore")
        def json(self):
            try:
                return json.loads(self.text)
            except Exception:
                return {"error": self.text}
    return FakeResponse(status_code, text)
def _do_req(method, api_key, secret_key, cid, uri, params=None, json_body=None, max_retries=3):
    url = OPENAPI_BASE_URL + uri
    last_err = None
    for i in range(max_retries):
        headers = _open_headers(api_key, secret_key, cid, method, uri)
        try:
            r = HTTP_SESSION.request(method, url, headers=headers, params=params, json=json_body, timeout=(5, 20))
            if r.status_code in [200, 201, 204]:
                return r
            if r.status_code == 429:
                time.sleep(1.25)
                continue
            if r.status_code == 404 and "1018" in r.text:
                time.sleep(1.0)
                continue
            return r
        except requests.exceptions.RequestException as e:
            last_err = str(e)
            time.sleep(1.25)
    return _make_fake_response(500, f"네트워크 통신 실패: {last_err or '알 수 없는 오류'}")
def _campaign_label(value: Any) -> str:
    return CAMPAIGN_TYPE_LABELS.get(str(value or "").strip(), str(value or "-"))
def _adgroup_label(value: Any) -> str:
    return ADGROUP_TYPE_LABELS.get(str(value or "").strip(), str(value or "-"))
def _ad_label(value: Any) -> str:
    return AD_TYPE_LABELS.get(str(value or "").strip(), str(value or "-"))
def _extension_label(value: Any) -> str:
    return AD_EXTENSION_TYPE_LABELS.get(str(value or "").strip(), str(value or "-"))
def _snake_to_camel(key: str) -> str:
    parts = str(key).strip().split("_")
    if len(parts) == 1:
        return parts[0]
    return parts[0] + "".join(p[:1].upper() + p[1:] for p in parts[1:])
def _boolify(v: Any) -> bool:
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    return s in {"1", "true", "y", "yes", "t", "on"}
def _normalize_value(key: str, value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, float) and pd.isna(value):
        return None
    if isinstance(value, str):
        value = value.strip()
        if value == "":
            return None
        lowered = value.lower()
        if lowered in {"예", "네", "사용", "y", "yes", "true", "on"}:
            value = True
        elif lowered in {"아니오", "미사용", "n", "no", "false", "off"}:
            value = False
        elif value.startswith("{") or value.startswith("["):
            try:
                return json.loads(value)
            except Exception:
                pass
    if key in BOOL_FIELDS:
        return _boolify(value)
    if key in INT_FIELDS:
        try:
            return int(float(str(value).replace(",", "").strip()))
        except Exception:
            return value
    if key in FLOAT_FIELDS:
        try:
            return float(str(value).replace(",", "").strip())
        except Exception:
            return value
    return value
def _strip_empty(data: Any) -> Any:
    if isinstance(data, dict):
        cleaned = {}
        for k, v in data.items():
            vv = _strip_empty(v)
            if vv is None or vv == "":
                continue
            cleaned[k] = vv
        return cleaned
    if isinstance(data, list):
        out = []
        for item in data:
            vv = _strip_empty(item)
            if vv is None:
                continue
            out.append(vv)
        return out
    return data
def _special_alias_map(entity_type: str) -> Dict[str, str]:
    base = {
        "customer_id": "customerId",
        "customerid": "customerId",
        "campaign_tp": "campaignTp",
        "daily_budget": "dailyBudget",
        "use_daily_budget": "useDailyBudget",
        "bid_amt": "bidAmt",
        "use_group_bid_amt": "useGroupBidAmt",
        "user_lock": "userLock",
        "owner_id": "ownerId",
        "reference_key": "referenceKey",
        "contents_network_bid_amt": "contentsNetworkBidAmt",
        "raw_json": "rawJson",
        "pc_final_url": "pcFinalUrl",
        "mobile_final_url": "mobileFinalUrl",
    }
    if entity_type == "adgroup":
        base.update({"campaign_id": "nccCampaignId"})
    elif entity_type in {"keyword", "ad", "restricted_keyword"}:
        base.update({"adgroup_id": "nccAdgroupId"})
    elif entity_type == "ad_extension":
        base.update({"adgroup_id": "ownerId"})
    return base
def _read_table_text(text: str) -> List[Dict[str, Any]]:
    if not text or not text.strip():
        return []
    sample = text[:2048]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        sep = dialect.delimiter
    except Exception:
        sep = "\t" if "\t" in sample else ","
    df = pd.read_csv(io.StringIO(text), sep=sep, dtype=str, keep_default_na=False)
    renamed = []
    for c in df.columns:
        key = str(c).strip()
        key_norm = re.sub(r"\s+", "", key).lower()
        renamed.append(KO_HEADER_ALIASES.get(key_norm, key))
    df.columns = renamed
    return df.to_dict(orient="records")
def _result_item(row_no: int, ok: bool, name: str, detail: str = "") -> Dict[str, Any]:
    return {"row_no": row_no, "ok": ok, "name": name, "detail": detail}
def _read_uploaded_table(upload) -> List[Dict[str, Any]]:
    if upload is None or not getattr(upload, "filename", ""):
        return []
    filename = str(upload.filename or "").lower()
    raw = upload.read()
    try:
        upload.stream.seek(0)
    except Exception:
        pass
    if not raw:
        return []
    if filename.endswith('.csv') or filename.endswith('.txt') or filename.endswith('.tsv'):
        text = raw.decode('utf-8-sig', errors='ignore')
        return _read_table_text(text)
    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        try:
            df = pd.read_excel(io.BytesIO(raw), dtype=str, keep_default_na=False)
        except Exception as e:
            raise ValueError(f'엑셀 파일을 읽지 못했습니다: {e}')
        renamed = []
        for c in df.columns:
            key = str(c).strip()
            key_norm = re.sub(r"\s+", "", key).lower()
            renamed.append(KO_HEADER_ALIASES.get(key_norm, key))
        df.columns = renamed
        return df.to_dict(orient='records')
    raise ValueError('지원 파일 형식은 csv, txt, tsv, xls, xlsx 입니다.')
def _looks_like_biz_channel_id(value: Any) -> bool:
    return bool(re.match(r"^bsn-", str(value or "").strip(), re.I))
def _normalize_campaign_tp(value: Any) -> str:
    s = str(value or "").strip().upper()
    if s in {"파워링크", "POWERLINK", "WEB", "WEB_SITE"}:
        return "WEB_SITE"
    if s in {"쇼핑검색", "쇼검", "SHOPPING"}:
        return "SHOPPING"
    return s or "WEB_SITE"
def _is_shopping_campaign_type(value: Any) -> bool:
    s = str(value or "").strip().upper()
    if not s:
        return False
    return s in {"SHOPPING", "SHOPPING_BRAND", "CATALOG", "PRODUCT", "SHOPPING_PRODUCT"} or ("SHOPPING" in s) or ("CATALOG" in s) or ("PRODUCT" in s and s != "WEB_SITE")
def _default_adgroup_type_for_campaign(value: Any) -> str:
    s = str(value or "").strip().upper()
    if s == "SHOPPING_BRAND":
        return "SHOPPING_BRAND"
    if s == "CATALOG":
        return "CATALOG"
    if s == "SHOPPING":
        return "SHOPPING"
    if _is_shopping_campaign_type(s):
        return "SHOPPING"
    return "WEB_SITE"
def _normalize_adgroup_tp(value: Any, campaign_tp: str = "WEB_SITE") -> str:
    s = str(value or "").strip().upper()
    campaign_tp = str(campaign_tp or "").strip().upper()
    default_tp = _default_adgroup_type_for_campaign(campaign_tp)
    if s in {"", "기본", "파워링크", "WEB_SITE"}:
        return default_tp
    if s in {"쇼핑검색", "쇼검", "SHOPPING", "SHOPING", "PRODUCT", "SHOPPING_PRODUCT"}:
        return default_tp if _is_shopping_campaign_type(campaign_tp) else "WEB_SITE"
    if s in {"카탈로그", "CATALOG"}:
        return "CATALOG"
    if s in {"쇼핑브랜드", "SHOPPING_BRAND"}:
        return "SHOPPING_BRAND"
    return s
def _normalize_ad_type(value: Any) -> str:
    s = str(value or "").strip().upper()
    reverse = {v.upper(): k for k, v in AD_TYPE_LABELS.items()}
    if s in reverse:
        return reverse[s]
    if s in {"기본소재", "텍스트", "파워링크기본소재", "TEXT_45"}:
        return "TEXT_45"
    return s
def _normalize_extension_type(value: Any) -> str:
    s = str(value or "").strip().upper()
    reverse = {v.upper(): k for k, v in AD_EXTENSION_TYPE_LABELS.items()}
    if s in reverse:
        s = reverse[s]
    # API/화면/기존 데이터에서 섞여 들어오는 별칭을 최대한 같은 타입으로 정규화
    alias_map = {
        "SUB_LINK": "SUB_LINKS",
        "SUBLINK": "SUB_LINKS",
        "SUBLINKS": "SUB_LINKS",
        "HEAD_LINE": "HEADLINE",
        "HEADLINES": "HEADLINE",
        "HEAD LINE": "HEADLINE",
        "PROMO": "PROMOTION",
        "SHOPPING_PROMO_TEXT": "PROMOTION",
        "SHOPPING_EXTRA_INFO": "SHOPPING_EXTRA",
        "EXTRA_DESCRIPTION": "DESCRIPTION_EXTRA",
        "DESCRIPTIONEXTRA": "DESCRIPTION_EXTRA",
        "DESCRIPTION EXTRA": "DESCRIPTION_EXTRA",
    }
    if s in alias_map:
        return alias_map[s]
    if s == "쇼핑상품부가정보".upper():
        return "SHOPPING_EXTRA"
    return s
def _build_text45_ad_object(payload: Dict[str, Any], existing_ad: Any = None) -> Dict[str, Any]:
    ad_obj = copy.deepcopy(existing_ad) if isinstance(existing_ad, dict) else {}
    if isinstance(existing_ad, str):
        try:
            parsed = json.loads(existing_ad)
            if isinstance(parsed, dict):
                ad_obj = parsed
        except Exception:
            ad_obj = {}
    pc_obj = ad_obj.get("pc") if isinstance(ad_obj.get("pc"), dict) else {}
    mobile_obj = ad_obj.get("mobile") if isinstance(ad_obj.get("mobile"), dict) else {}
    fallback_final = str(payload.pop("finalUrl", "") or "").strip()
    headline = str(
        payload.pop("headline", payload.pop("title", ""))
        or ad_obj.get("headline")
        or ad_obj.get("title")
        or ""
    ).strip()
    description = str(
        payload.pop("description", "")
        or ad_obj.get("description")
        or ad_obj.get("desc")
        or ""
    ).strip()
    pc_final = str(payload.pop("pcFinalUrl", "") or pc_obj.get("final") or fallback_final).strip()
    mobile_final = str(
        payload.pop("mobileFinalUrl", payload.pop("mobileUrl", ""))
        or mobile_obj.get("final")
        or pc_final
        or fallback_final
    ).strip()
    merged = copy.deepcopy(ad_obj) if isinstance(ad_obj, dict) else {}
    if headline:
        merged["headline"] = headline
    if description:
        merged["description"] = description
    if pc_final:
        pc_payload = copy.deepcopy(pc_obj) if isinstance(pc_obj, dict) else {}
        pc_payload["final"] = pc_final
        merged["pc"] = pc_payload
    if mobile_final:
        mobile_payload = copy.deepcopy(mobile_obj) if isinstance(mobile_obj, dict) else {}
        mobile_payload["final"] = mobile_final
        merged["mobile"] = mobile_payload
    return _strip_empty(merged)
def _prepare_payload_row(row: Dict[str, Any], entity_type: str, cid: str) -> Dict[str, Any]:
    alias_map = _special_alias_map(entity_type)
    payload: Dict[str, Any] = {}
    for raw_key, raw_value in row.items():
        if raw_key is None:
            continue
        key0 = str(raw_key).strip()
        if not key0:
            continue
        key0_norm = re.sub(r"\s+", "", key0).lower()
        key = KO_HEADER_ALIASES.get(key0_norm, alias_map.get(key0, key0))
        if "_" in key and key not in {"nccCampaignId", "nccAdgroupId", "ownerId", "referenceKey", "rawJson"}:
            key = _snake_to_camel(key)
        value = _normalize_value(key, raw_value)
        if value is None:
            continue
        payload[key] = value
    payload["customerId"] = int(cid)
    if entity_type == "campaign":
        payload["campaignTp"] = _normalize_campaign_tp(payload.get("campaignTp"))
        payload.pop("nccCampaignId", None)
    if entity_type == "adgroup":
        campaign_tp = _normalize_campaign_tp(payload.get("campaignTp") or payload.get("parentCampaignTp") or "")
        payload["adgroupType"] = _normalize_adgroup_tp(payload.get("adgroupType"), campaign_tp)
        biz_channel_id = str(payload.get("nccBusinessChannelId") or "").strip()
        if biz_channel_id and _looks_like_biz_channel_id(biz_channel_id):
            payload.setdefault("pcChannelId", biz_channel_id)
            payload.setdefault("mobileChannelId", biz_channel_id)
        payload.pop("nccBusinessChannelId", None)
    if entity_type == "ad":
        ad_type = _normalize_ad_type(payload.get("type"))
        payload["type"] = ad_type
        raw_json = payload.pop("rawJson", None)
        if raw_json and isinstance(raw_json, str):
            try:
                raw_json = json.loads(raw_json)
            except Exception:
                raw_json = None
        if ad_type == "TEXT_45":
            existing_ad = payload.pop("ad", None)
            if isinstance(raw_json, dict):
                payload["ad"] = raw_json
            else:
                payload["ad"] = _build_text45_ad_object(payload, existing_ad)
        elif raw_json is not None:
            payload["ad"] = raw_json
        elif isinstance(payload.get("ad"), str):
            try:
                payload["ad"] = json.loads(payload["ad"])
            except Exception:
                pass
    if entity_type == "ad_extension":
        payload.pop("nccAdgroupId", None)
        payload["type"] = _normalize_extension_type(payload.get("type"))
        raw_json = payload.pop("rawJson", None)
        if raw_json and isinstance(raw_json, str):
            try:
                raw_payload = json.loads(raw_json)
                if isinstance(raw_payload, dict):
                    raw_payload.setdefault("ownerId", payload.get("ownerId"))
                    raw_payload.setdefault("type", payload.get("type"))
                    raw_payload.setdefault("customerId", int(cid))
                    payload = raw_payload
            except Exception:
                payload["rawJson"] = raw_json
    if entity_type == "restricted_keyword":
        rk_type = payload.get("type") or payload.get("matchType")
        payload = {
            "nccAdgroupId": str(payload.get("nccAdgroupId") or "").strip(),
            "customerId": int(cid),
            "keyword": str(payload.get("keyword") or payload.get("restrictedKeyword") or "").strip(),
        }
        if rk_type:
            payload["type"] = rk_type
    if entity_type != "ad":
        payload.pop("referenceData", None)
    for field in list(payload.keys()):
        if field in SYSTEM_FIELDS:
            payload.pop(field, None)
    return _strip_empty(payload)
def _normalize_campaign_item(item: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": item.get("nccCampaignId") or item.get("id") or "",
        "name": item.get("name") or "",
        "campaignTp": item.get("campaignTp") or "",
        "label": _campaign_label(item.get("campaignTp")),
        "badgeColor": CAMPAIGN_TYPE_COLORS.get(item.get("campaignTp"), "gray"),
        "raw": item,
    }
def _normalize_adgroup_item(item: Dict[str, Any]) -> Dict[str, Any]:
    return {
        "id": item.get("nccAdgroupId") or "",
        "name": item.get("name") or "",
        "adgroupType": item.get("adgroupType") or "",
        "label": _adgroup_label(item.get("adgroupType")),
        "pcChannelId": item.get("pcChannelId") or "",
        "mobileChannelId": item.get("mobileChannelId") or "",
        "nccCampaignId": item.get("nccCampaignId") or "",
        "raw": item,
    }
def _normalize_channel_item(item: Dict[str, Any]) -> Dict[str, Any]:
    cid = item.get("nccBusinessChannelId") or item.get("bizChannelId") or item.get("nccChannelId") or item.get("channelId") or ""
    name = item.get("name") or item.get("channelContents") or item.get("siteUrl") or cid
    return {
        "id": cid,
        "name": name,
        "channelTp": item.get("channelTp") or item.get("bizChannelType") or item.get("channelType") or "",
        "siteUrl": item.get("siteUrl") or "",
        "raw": item,
    }
def _fetch_campaigns(api_key: str, secret_key: str, cid: str):
    res = _do_req("GET", api_key, secret_key, cid, "/ncc/campaigns")
    if res.status_code != 200:
        return res, []
    rows = [_normalize_campaign_item(x) for x in (res.json() or [])]
    rows.sort(key=lambda x: (x["label"], x["name"]))
    return res, rows
def _fetch_campaign_detail(api_key: str, secret_key: str, cid: str, campaign_id: str):
    campaign_id = str(campaign_id or "").strip()
    if not campaign_id:
        return None
    try:
        res = _do_req("GET", api_key, secret_key, cid, f"/ncc/campaigns/{campaign_id}")
        if res.status_code == 200 and isinstance(res.json(), dict):
            return res.json()
    except Exception:
        pass
    try:
        _, rows = _fetch_campaigns(api_key, secret_key, cid)
        for row in (rows or []):
            if str(row.get("id") or "").strip() == campaign_id:
                raw = row.get("raw")
                if isinstance(raw, dict):
                    return raw
                return {"nccCampaignId": row.get("id"), "campaignTp": row.get("campaignTp"), "name": row.get("name")}
    except Exception:
        pass
    return None
def _pc_mobile_label(pc: Any, mobile: Any) -> str:
    if pc is True and mobile is False:
        return "PC"
    if pc is False and mobile is True:
        return "MOBILE"
    if pc is True and mobile is True:
        return "ALL"
    return "UNKNOWN"
def _extract_pc_mobile_flags(detail_obj: Dict[str, Any] | None, pm_target_obj: Dict[str, Any] | None = None) -> Tuple[Optional[bool], Optional[bool]]:
    detail_obj = detail_obj or {}
    pc = detail_obj.get("pcDevice")
    mobile = detail_obj.get("mobileDevice")
    if isinstance(pc, bool) and isinstance(mobile, bool):
        return pc, mobile
    target = (pm_target_obj or {}).get("target") if isinstance(pm_target_obj, dict) else None
    if isinstance(target, dict):
        pc_t = target.get("pc")
        mobile_t = target.get("mobile")
        if isinstance(pc_t, bool) and isinstance(mobile_t, bool):
            return pc_t, mobile_t
    return None, None
def _enrich_adgroup_media_row(api_key: str, secret_key: str, cid: str, row: Dict[str, Any]) -> Dict[str, Any]:
    row = copy.deepcopy(row)
    raw = row.get("raw") if isinstance(row.get("raw"), dict) else {}
    adgroup_id = str(row.get("id") or raw.get("nccAdgroupId") or "").strip()
    if not adgroup_id:
        return row
    detail_obj: Dict[str, Any] | None = None
    pm_target_obj: Dict[str, Any] | None = None
    res_detail, detail_obj = _fetch_adgroup_detail(api_key, secret_key, cid, adgroup_id)
    if res_detail.status_code == 200 and isinstance(detail_obj, dict):
        merged_raw = copy.deepcopy(raw)
        merged_raw.update(detail_obj)
        raw = merged_raw
    if not isinstance(raw.get("pcDevice"), bool) or not isinstance(raw.get("mobileDevice"), bool):
        try:
            res_target, pm_target_obj = _fetch_target_object(api_key, secret_key, cid, adgroup_id, "PC_MOBILE_TARGET")
            if res_target.status_code != 200:
                pm_target_obj = None
        except Exception:
            pm_target_obj = None
    pc, mobile = _extract_pc_mobile_flags(detail_obj if isinstance(detail_obj, dict) else raw, pm_target_obj)
    if isinstance(pc, bool):
        raw["pcDevice"] = pc
    if isinstance(mobile, bool):
        raw["mobileDevice"] = mobile
    target_summary = raw.get("targetSummary") if isinstance(raw.get("targetSummary"), dict) else {}
    media_label = _pc_mobile_label(raw.get("pcDevice"), raw.get("mobileDevice"))
    if media_label != "UNKNOWN":
        target_summary["pcMobile"] = media_label
        raw["targetSummary"] = target_summary
        row["mediaType"] = media_label
    row["pcDevice"] = raw.get("pcDevice")
    row["mobileDevice"] = raw.get("mobileDevice")
    row["raw"] = raw
    return row
def _fetch_adgroups(api_key: str, secret_key: str, cid: str, campaign_id: str, enrich_media: bool = True):
    res = _do_req("GET", api_key, secret_key, cid, "/ncc/adgroups", params={"nccCampaignId": campaign_id})
    if res.status_code != 200:
        return res, []
    rows = [_normalize_adgroup_item(x) for x in (res.json() or [])]
    if enrich_media and rows:
        max_workers = min(8, max(1, len(rows)))
        enriched_rows: List[Dict[str, Any]] = []
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            future_map = {ex.submit(_enrich_adgroup_media_row, api_key, secret_key, cid, row): idx for idx, row in enumerate(rows)}
            ordered: Dict[int, Dict[str, Any]] = {}
            for fut in as_completed(future_map):
                idx = future_map[fut]
                try:
                    ordered[idx] = fut.result()
                except Exception:
                    ordered[idx] = rows[idx]
            enriched_rows = [ordered[i] for i in range(len(rows))]
        rows = enriched_rows
    rows.sort(key=lambda x: x["name"])
    return res, rows
def _fetch_adgroup_detail(api_key: str, secret_key: str, cid: str, adgroup_id: str):
    res = _do_req("GET", api_key, secret_key, cid, f"/ncc/adgroups/{adgroup_id}")
    if res.status_code == 200 and isinstance(res.json(), dict):
        return res, res.json()
    return res, None
def _fetch_first_biz_channel_id(api_key: str, secret_key: str, cid: str) -> str:
    res = _do_req("GET", api_key, secret_key, cid, "/ncc/channels")
    if res.status_code != 200:
        return ""
    for item in (res.json() or []):
        norm = _normalize_channel_item(item if isinstance(item, dict) else {})
        cid_val = str(norm.get("id") or "").strip()
        if cid_val:
            return cid_val
    return ""
def _is_shopping_adgroup(adgroup_obj: Dict[str, Any] | None) -> bool:
    adg_type = str((adgroup_obj or {}).get("adgroupType") or "").upper()
    return adg_type in {"SHOPPING", "CATALOG", "SHOPPING_BRAND"}
def _fetch_keywords(api_key: str, secret_key: str, cid: str, adgroup_id: str):
    res = _do_req("GET", api_key, secret_key, cid, "/ncc/keywords", params={"nccAdgroupId": adgroup_id})
    if res.status_code != 200:
        return res, []
    return res, res.json() or []
def _normalize_keyword_search_text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()
def _parse_keyword_search_groups(search_text: Any) -> List[str]:
    raw = str(search_text or "")
    parts = re.split(r"[\r\n,;|]+", raw)
    groups: List[str] = []
    seen: set[str] = set()
    for part in parts:
        norm = _normalize_keyword_search_text(part)
        if not norm or norm in seen:
            continue
        seen.add(norm)
        groups.append(norm)
    if groups:
        return groups
    fallback = _normalize_keyword_search_text(raw)
    return [fallback] if fallback else []
def _keyword_match_terms(keyword: Any, search_text: Any, exact_match: bool = False, exclude_text: Any = None) -> List[str]:
    keyword_norm = _normalize_keyword_search_text(keyword)
    if not keyword_norm:
        return []
    if _keyword_matches_search(keyword_norm, exclude_text, exact_match=exact_match):
        return []
    groups = _parse_keyword_search_groups(search_text)
    matched: List[str] = []
    for group in groups:
        if not group:
            continue
        if bool(exact_match):
            if keyword_norm == group:
                matched.append(group)
            continue
        tokens = [tok for tok in group.split(" ") if tok]
        if tokens and all(tok in keyword_norm for tok in tokens):
            matched.append(group)
    return matched
def _keyword_matches_search(keyword: Any, search_text: Any, exact_match: bool = False) -> bool:
    keyword_norm = _normalize_keyword_search_text(keyword)
    if not keyword_norm:
        return False
    groups = _parse_keyword_search_groups(search_text)
    for group in groups:
        if not group:
            continue
        if bool(exact_match):
            if keyword_norm == group:
                return True
            continue
        tokens = [tok for tok in group.split(" ") if tok]
        if tokens and all(tok in keyword_norm for tok in tokens):
            return True
    return False
def _collect_powerlink_campaigns_and_adgroups(api_key: str, secret_key: str, cid: str, campaign_ids: List[str] | None = None):
    selected_ids = [str(x or "").strip() for x in (campaign_ids or []) if str(x or "").strip()]
    selected_set = set(selected_ids)
    res_camp, campaign_rows = _fetch_campaigns(api_key, secret_key, cid)
    if res_camp.status_code != 200:
        return res_camp, [], [], [f"캠페인 조회 실패: {res_camp.text}"]
    powerlink_campaigns = []
    skipped_non_powerlink: List[str] = []
    missing_selected = set(selected_ids)
    for row in (campaign_rows or []):
        raw = row.get("raw") if isinstance(row.get("raw"), dict) else {}
        camp_id = str(row.get("id") or raw.get("nccCampaignId") or "").strip()
        camp_name = str(row.get("name") or raw.get("name") or camp_id).strip()
        camp_tp = str(row.get("campaignTp") or raw.get("campaignTp") or "").upper()
        if selected_set and camp_id not in selected_set:
            continue
        if camp_id in missing_selected:
            missing_selected.discard(camp_id)
        if camp_tp == "WEB_SITE":
            powerlink_campaigns.append({"id": camp_id, "name": camp_name})
        elif selected_set and camp_id:
            skipped_non_powerlink.append(f"{camp_name} ({camp_tp or '유형없음'})")
    warnings: List[str] = []
    if missing_selected:
        warnings.append(f"선택 캠페인 {len(missing_selected)}개를 찾지 못했습니다.")
    if skipped_non_powerlink:
        warnings.append(f"파워링크가 아닌 선택 캠페인 {len(skipped_non_powerlink)}개는 제외했습니다.")
    if not powerlink_campaigns:
        return res_camp, [], [], warnings[:10]
    adgroup_contexts: List[Dict[str, Any]] = []
    max_workers = min(max(FAST_IO_WORKERS, 8), max(1, len(powerlink_campaigns)))
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        future_map = {
            ex.submit(_fetch_adgroups, api_key, secret_key, cid, camp["id"], False): camp
            for camp in powerlink_campaigns if camp.get("id")
        }
        for fut in as_completed(future_map):
            camp = future_map[fut]
            camp_id = str(camp.get("id") or "").strip()
            camp_name = str(camp.get("name") or camp_id)
            try:
                res_adg, rows = fut.result()
            except Exception as exc:
                warnings.append(f"캠페인 {camp_name} 광고그룹 조회 실패: {exc}")
                continue
            if res_adg.status_code != 200:
                warnings.append(f"캠페인 {camp_name} 광고그룹 조회 실패: {res_adg.text}")
                continue
            for row in (rows or []):
                raw = row.get("raw") if isinstance(row.get("raw"), dict) else {}
                adg_id = str(row.get("id") or raw.get("nccAdgroupId") or "").strip()
                adg_type = str(row.get("adgroupType") or raw.get("adgroupType") or "").upper()
                if not adg_id or adg_type != "WEB_SITE":
                    continue
                adgroup_contexts.append({
                    "campaign_id": camp_id,
                    "campaign_name": camp_name,
                    "adgroup_id": adg_id,
                    "adgroup_name": str(row.get("name") or raw.get("name") or adg_id),
                    "adgroup_bid": (raw or {}).get("bidAmt"),
                })
    adgroup_contexts.sort(key=lambda x: (str(x.get("campaign_name") or "").casefold(), str(x.get("adgroup_name") or "").casefold()))
    return res_camp, powerlink_campaigns, adgroup_contexts, warnings[:10]
def _scan_powerlink_keywords_by_search(api_key: str, secret_key: str, cid: str, search_text: str, exact_match: bool = False, campaign_ids: List[str] | None = None, adgroup_ids: List[str] | None = None, exclude_text: str = ""):
    _, powerlink_campaigns, adgroup_contexts, warnings = _collect_powerlink_campaigns_and_adgroups(api_key, secret_key, cid, campaign_ids=campaign_ids)
    selected_adgroup_ids = _unique_keep_order([str(x or "").strip() for x in (adgroup_ids or []) if str(x or "").strip()])
    if selected_adgroup_ids:
        selected_adgroup_set = set(selected_adgroup_ids)
        adgroup_contexts = [ctx for ctx in (adgroup_contexts or []) if str(ctx.get("adgroup_id") or "").strip() in selected_adgroup_set]
        selected_campaign_set = {str(ctx.get("campaign_id") or "").strip() for ctx in adgroup_contexts if str(ctx.get("campaign_id") or "").strip()}
        powerlink_campaigns = [row for row in (powerlink_campaigns or []) if str((row or {}).get("id") or "").strip() in selected_campaign_set]
    matched_rows: List[Dict[str, Any]] = []
    update_payload: List[Dict[str, Any]] = []
    updated_adgroup_ids: List[str] = []
    err_details: List[str] = list(warnings)
    cleanup_keys = ['regTm', 'editTm', 'status', 'statusReason', 'inspectStatus', 'delFlag', 'managedKeyword', 'referenceKey']
    search_groups = _parse_keyword_search_groups(search_text)
    exclude_groups = _parse_keyword_search_groups(exclude_text)
    scanned_keyword_count = 0
    matched_count = 0
    if not powerlink_campaigns or not adgroup_contexts:
        return {
            "powerlink_campaigns": powerlink_campaigns,
            "adgroup_contexts": adgroup_contexts,
            "warnings": err_details[:10],
            "matched_rows": matched_rows,
            "update_payload": update_payload,
            "updated_adgroup_ids": updated_adgroup_ids,
            "scanned_keyword_count": 0,
            "matched_count": 0,
            "search_groups": search_groups,
            "exclude_groups": exclude_groups,
        }
    max_workers = min(max(BID_IO_WORKERS, 10), max(1, len(adgroup_contexts)))
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        future_map = {
            ex.submit(_fetch_keywords, api_key, secret_key, cid, str(ctx.get("adgroup_id") or "")): ctx
            for ctx in adgroup_contexts if str(ctx.get("adgroup_id") or "").strip()
        }
        for fut in as_completed(future_map):
            ctx = future_map[fut]
            adg_id = str(ctx.get("adgroup_id") or "").strip()
            adg_name = str(ctx.get("adgroup_name") or adg_id)
            camp_name = str(ctx.get("campaign_name") or "")
            try:
                res_kw, kw_rows = fut.result()
            except Exception as exc:
                if len(err_details) < 10:
                    err_details.append(f"[{camp_name} > {adg_name}] 키워드 조회 실패: {exc}")
                continue
            if res_kw.status_code != 200:
                if len(err_details) < 10:
                    err_details.append(f"[{camp_name} > {adg_name}] 키워드 조회 실패: {res_kw.text}")
                continue
            adgroup_had_match = False
            for kw in (kw_rows or []):
                keyword_text = str((kw or {}).get("keyword") or "").strip()
                if not keyword_text:
                    continue
                scanned_keyword_count += 1
                matched_terms = _keyword_match_terms(keyword_text, search_text, exact_match=exact_match, exclude_text=exclude_text)
                if not matched_terms:
                    continue
                matched_count += 1
                current_bid = _normalize_bid_amt((kw or {}).get("bidAmt")) or 70
                current_use_group = bool((kw or {}).get("useGroupBidAmt"))
                try:
                    current_bid_weight = int((kw or {}).get("bidWeight") or 100)
                except Exception:
                    current_bid_weight = 100
                keyword_enabled = _extract_enabled_from_entity(kw)
                matched_rows.append({
                    "campaign_id": str(ctx.get("campaign_id") or "").strip(),
                    "campaign_name": camp_name,
                    "adgroup_id": adg_id,
                    "adgroup_name": adg_name,
                    "keyword": keyword_text,
                    "current_bid": int(current_bid),
                    "current_use_group": bool(current_use_group),
                    "current_bid_weight": int(current_bid_weight),
                    "current_state": "ON" if keyword_enabled is True else ("OFF" if keyword_enabled is False else "-"),
                    "enabled": keyword_enabled,
                    "matched_terms": matched_terms,
                    "matched_terms_text": ", ".join(matched_terms),
                    "ncc_keyword_id": str((kw or {}).get("nccKeywordId") or "").strip(),
                })
                item = copy.deepcopy(kw)
                item["useGroupBidAmt"] = False
                for key in cleanup_keys:
                    item.pop(key, None)
                update_payload.append(item)
                adgroup_had_match = True
            if adgroup_had_match:
                updated_adgroup_ids.append(adg_id)
    matched_rows.sort(key=lambda x: (str(x.get("campaign_name") or "").casefold(), str(x.get("adgroup_name") or "").casefold(), str(x.get("keyword") or "").casefold()))
    return {
        "powerlink_campaigns": powerlink_campaigns,
        "adgroup_contexts": adgroup_contexts,
        "warnings": err_details[:10],
        "matched_rows": matched_rows,
        "update_payload": update_payload,
        "updated_adgroup_ids": updated_adgroup_ids,
        "scanned_keyword_count": int(scanned_keyword_count),
        "matched_count": int(matched_count),
        "search_groups": search_groups,
    }
def _extract_enabled_from_entity(entity: Dict[str, Any] | None) -> bool | None:
    src = entity or {}
    lock_val = src.get("userLock")
    if isinstance(lock_val, bool):
        return not lock_val
    enable_val = src.get("enable")
    if isinstance(enable_val, bool):
        return enable_val
    paused_val = src.get("paused")
    if isinstance(paused_val, bool):
        return not paused_val
    status = str(src.get("status") or "").upper()
    if status in {"PAUSE", "PAUSED", "STOP", "STOPPED", "LIMITEDBYBUDGET"}:
        return False
    if status:
        return True
    return None
def _set_keyword_state(api_key: str, secret_key: str, cid: str, keyword_id: str, enabled: bool) -> Tuple[bool, str]:
    keyword_id = str(keyword_id or "").strip()
    if not keyword_id:
        return False, "키워드 ID가 비어 있습니다."
    uri = f"/ncc/keywords/{keyword_id}"
    r_get = _do_req("GET", api_key, secret_key, cid, uri)
    if r_get.status_code != 200:
        return False, f"조회 실패: {r_get.text}"
    obj = r_get.json() or {}
    attempts: List[Tuple[str, Dict[str, Any]]] = []
    def _append_attempt(field_name: str, value: Any):
        body = copy.deepcopy(obj)
        body[field_name] = value
        attempts.append((field_name, body))
    if "userLock" in obj:
        _append_attempt("userLock", not enabled)
    if "enable" in obj:
        _append_attempt("enable", enabled)
    if "paused" in obj:
        _append_attempt("paused", not enabled)
    if not attempts:
        _append_attempt("userLock", not enabled)
        _append_attempt("enable", enabled)
    tried: set[str] = set()
    errors: List[str] = []
    for field_name, body in attempts:
        if field_name in tried:
            continue
        tried.add(field_name)
        r_put = _do_req("PUT", api_key, secret_key, cid, uri, params={"fields": field_name}, json_body=body)
        if r_put.status_code in [200, 201]:
            return True, ""
        errors.append(f"{field_name}: {r_put.text}")
    return False, " / ".join(errors) if errors else "상태 변경 실패"
def _build_powerlink_keyword_search_message(search_text: str, exact_match: bool, scan: Dict[str, Any], row_preview_limit: int = 50, exclude_text: str = "") -> str:
    powerlink_campaigns = scan.get("powerlink_campaigns") or []
    adgroup_contexts = scan.get("adgroup_contexts") or []
    matched_rows = scan.get("matched_rows") or []
    err_details = scan.get("warnings") or []
    search_groups = scan.get("search_groups") or _parse_keyword_search_groups(search_text)
    exclude_groups = scan.get("exclude_groups") or _parse_keyword_search_groups(exclude_text)
    scanned_keyword_count = int(scan.get("scanned_keyword_count") or 0)
    matched_count = int(scan.get("matched_count") or 0)
    mode_label = "완전일치" if exact_match else "부분일치"
    lines = [
        f"검색어 기준 파워링크 키워드 조회 완료! ({mode_label})",
        f"검색어: {search_text}",
        (f"제외어: {exclude_text}" if str(exclude_text or "").strip() else "제외어: 없음"),
        f"검색 그룹: {len(search_groups)}개",
        f"조회한 파워링크 캠페인: {len(powerlink_campaigns)}개 | 광고그룹: {len(adgroup_contexts)}개 | 키워드 스캔: {scanned_keyword_count}개",
        f"검색 일치 키워드: {matched_count}개",
    ]
    if matched_rows:
        lines.append("\n[일치 키워드 예시]")
        for row in matched_rows[:row_preview_limit]:
            current_label = f"{int(row.get('current_bid') or 0):,}원"
            if row.get("current_use_group"):
                current_label += " (그룹입찰가 사용)"
            matched_term_label = str(row.get("matched_terms_text") or "")
            if matched_term_label:
                matched_term_label = f" | 매칭: {matched_term_label}"
            lines.append(f"- {row.get('campaign_name')} > {row.get('adgroup_name')} > {row.get('keyword')}{matched_term_label} | 현재 {current_label}")
    if err_details:
        lines.append("\n[상세 내역]")
        lines.extend(err_details[:10])
    return "\n".join(lines)
def _build_powerlink_keyword_export_workbook(rows: List[Dict[str, Any]], search_text: str, exact_match: bool, exclude_text: str = ""):
    wb = Workbook()
    ws = wb.active
    ws.title = "키워드조회결과"
    generated_at = time.strftime("%Y-%m-%d %H:%M:%S")
    mode_label = "완전일치" if exact_match else "부분일치"
    search_groups = _parse_keyword_search_groups(search_text)
    ws["A1"] = "파워링크 키워드 조회 결과"
    ws["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="1D4ED8")
    ws.merge_cells("A1:J1")
    ws["A2"] = f"생성시각: {generated_at}"
    exclude_label = exclude_text if str(exclude_text or "").strip() else "없음"
    ws["A3"] = f"조회조건: {mode_label} / 검색어 {len(search_groups)}개 / {search_text}"
    ws["A4"] = f"제외어: {exclude_label}"
    headers = ["번호", "캠페인명", "광고그룹명", "키워드", "매칭 검색어", "현재 입찰가(원)", "그룹입찰가 사용", "현재 입찰가중치(%)", "키워드 ID", "조회 방식"]
    start_row = 6
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111827")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for idx, row in enumerate(rows, start=1):
        values = [
            idx,
            str(row.get("campaign_name") or ""),
            str(row.get("adgroup_name") or ""),
            str(row.get("keyword") or ""),
            str(row.get("matched_terms_text") or ""),
            int(row.get("current_bid") or 0),
            "Y" if row.get("current_use_group") else "N",
            int(row.get("current_bid_weight") or 100),
            str(row.get("ncc_keyword_id") or ""),
            mode_label,
        ]
        for col_idx, value in enumerate(values, start=1):
            ws.cell(row=start_row + idx, column=col_idx, value=value)
    widths = {1: 8, 2: 24, 3: 24, 4: 32, 5: 24, 6: 16, 7: 14, 8: 18, 9: 22, 10: 14}
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    end_row = start_row + max(1, len(rows))
    for row in ws.iter_rows(min_row=start_row + 1, max_row=end_row, min_col=1, max_col=10):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A7"
    return wb
def _find_powerlink_duplicate_keywords(api_key: str, secret_key: str, cid: str, campaign_ids: List[str]):
    selected_ids = [str(x or "").strip() for x in (campaign_ids or []) if str(x or "").strip()]
    if not selected_ids:
        return {"rows": [], "message": "선택한 캠페인이 없습니다.", "skipped": [], "errors": []}
    res_camp, campaign_rows = _fetch_campaigns(api_key, secret_key, cid)
    if res_camp.status_code != 200:
        raise RuntimeError(f"캠페인 조회 실패: {res_camp.text}")
    campaign_map = {}
    for row in campaign_rows or []:
        camp_id = str(row.get("id") or row.get("nccCampaignId") or "").strip()
        if camp_id:
            campaign_map[camp_id] = row
    powerlink_targets: List[Tuple[str, str]] = []
    skipped: List[str] = []
    for camp_id in selected_ids:
        row = campaign_map.get(camp_id)
        if not row:
            skipped.append(f"{camp_id} (캠페인 조회 실패 또는 없음)")
            continue
        camp_name = str(row.get("name") or camp_id)
        camp_tp = str(row.get("campaignTp") or "").upper()
        if camp_tp != "WEB_SITE":
            skipped.append(f"{camp_name} ({camp_tp or '유형없음'})")
            continue
        powerlink_targets.append((camp_id, camp_name))
    rows_out: List[Dict[str, Any]] = []
    errors: List[str] = []
    for camp_id, camp_name in powerlink_targets:
        res_adg, adgroup_rows = _fetch_adgroups(api_key, secret_key, cid, camp_id, enrich_media=False)
        if res_adg.status_code != 200:
            errors.append(f"[{camp_name}] 광고그룹 조회 실패: {res_adg.text}")
            continue
        adgroups = []
        for row in adgroup_rows or []:
            adgroup_id = str(row.get("id") or row.get("nccAdgroupId") or "").strip()
            if not adgroup_id:
                continue
            adgroup_name = str(row.get("name") or adgroup_id)
            adgroups.append((adgroup_id, adgroup_name))
        keyword_map: Dict[str, Dict[str, Any]] = {}
        max_workers = max(1, min(FAST_IO_WORKERS, len(adgroups) or 1))
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            future_map = {ex.submit(_fetch_keywords, api_key, secret_key, cid, adg_id): (adg_id, adg_name) for adg_id, adg_name in adgroups}
            for fut in as_completed(future_map):
                adg_id, adg_name = future_map[fut]
                try:
                    res_kw, kw_rows = fut.result()
                except Exception as e:
                    errors.append(f"[{camp_name} > {adg_name}] 키워드 조회 실패: {e}")
                    continue
                if res_kw.status_code != 200:
                    errors.append(f"[{camp_name} > {adg_name}] 키워드 조회 실패: {res_kw.text}")
                    continue
                seen_in_group = set()
                for kw in kw_rows or []:
                    keyword = str(kw.get("keyword") or "").strip()
                    if not keyword:
                        continue
                    norm = re.sub(r"\s+", " ", keyword).strip().casefold()
                    if not norm or norm in seen_in_group:
                        continue
                    seen_in_group.add(norm)
                    entry = keyword_map.setdefault(norm, {"keyword": keyword, "adgroups": []})
                    entry["adgroups"].append({"id": adg_id, "name": adg_name})
        for item in keyword_map.values():
            uniq = []
            seen_ids = set()
            for g in item.get("adgroups") or []:
                gid = str(g.get("id") or "")
                if not gid or gid in seen_ids:
                    continue
                seen_ids.add(gid)
                uniq.append(g)
            if len(uniq) < 2:
                continue
            uniq_names = [str(g.get("name") or g.get("id") or "") for g in uniq]
            rows_out.append({
                "campaign_id": camp_id,
                "campaign_name": camp_name,
                "keyword": item.get("keyword") or "",
                "adgroup_count": len(uniq),
                "adgroups": uniq,
                "adgroup_names": uniq_names,
                "adgroup_names_text": ", ".join(uniq_names),
            })
    rows_out.sort(key=lambda x: (str(x.get("campaign_name") or ""), -int(x.get("adgroup_count") or 0), str(x.get("keyword") or "").casefold()))
    msg_parts = []
    if powerlink_targets:
        msg_parts.append(f"파워링크 캠페인 {len(powerlink_targets)}개 기준")
    if rows_out:
        msg_parts.append(f"중복 키워드 {len(rows_out)}개 발견")
    else:
        msg_parts.append("중복 키워드가 없습니다.")
    if skipped:
        msg_parts.append(f"제외 {len(skipped)}개")
    return {
        "rows": rows_out,
        "message": " / ".join(msg_parts),
        "skipped": skipped,
        "errors": errors[:30],
        "checked_campaign_count": len(selected_ids),
        "powerlink_campaign_count": len(powerlink_targets),
    }
def _find_account_powerlink_duplicate_keywords(api_key: str, secret_key: str, cid: str):
    res_camp, campaign_rows = _fetch_campaigns(api_key, secret_key, cid)
    if res_camp.status_code != 200:
        raise RuntimeError(f"캠페인 조회 실패: {res_camp.text}")
    powerlink_ids: List[str] = []
    for row in campaign_rows or []:
        camp_id = str(row.get("id") or row.get("nccCampaignId") or "").strip()
        camp_tp = str(row.get("campaignTp") or "").upper()
        if camp_id and camp_tp == "WEB_SITE":
            powerlink_ids.append(camp_id)
    if not powerlink_ids:
        return {"rows": [], "message": "현재 계정에 파워링크 캠페인이 없습니다.", "skipped": [], "errors": [], "checked_campaign_count": 0, "powerlink_campaign_count": 0}
    result = _find_powerlink_duplicate_keywords(api_key, secret_key, cid, powerlink_ids)
    rows = list(result.get("rows") or [])
    rows.sort(key=lambda x: (-int(x.get("adgroup_count") or 0), str(x.get("keyword") or "").casefold(), str(x.get("campaign_name") or "").casefold()))
    result["rows"] = rows
    found = len(rows)
    result["message"] = f"현재 계정 파워링크 캠페인 {len(powerlink_ids)}개 기준 / 중복 키워드 {found}개 발견" if found else f"현재 계정 파워링크 캠페인 {len(powerlink_ids)}개 기준 / 중복 키워드가 없습니다."
    result["checked_campaign_count"] = len(powerlink_ids)
    result["powerlink_campaign_count"] = len(powerlink_ids)
    return result
def _get_powerlink_keyword_stats(api_key: str, secret_key: str, cid: str, campaign_ids: List[str] | None = None):
    selected_ids = [str(x or "").strip() for x in (campaign_ids or []) if str(x or "").strip()]
    res_camp, campaign_rows = _fetch_campaigns(api_key, secret_key, cid)
    if res_camp.status_code != 200:
        raise RuntimeError(f"캠페인 조회 실패: {res_camp.text}")
    campaign_map: Dict[str, Dict[str, Any]] = {}
    powerlink_targets: List[Tuple[str, str]] = []
    for row in campaign_rows or []:
        camp_id = str(row.get("id") or row.get("nccCampaignId") or "").strip()
        if not camp_id:
            continue
        campaign_map[camp_id] = row
        if str(row.get("campaignTp") or "").upper() == "WEB_SITE":
            powerlink_targets.append((camp_id, str(row.get("name") or camp_id)))
    skipped: List[str] = []
    selected_powerlink_targets: List[Tuple[str, str]] = []
    for camp_id in selected_ids:
        row = campaign_map.get(camp_id)
        if not row:
            skipped.append(f"{camp_id} (캠페인 조회 실패 또는 없음)")
            continue
        camp_name = str(row.get("name") or camp_id)
        camp_tp = str(row.get("campaignTp") or "").upper()
        if camp_tp != "WEB_SITE":
            skipped.append(f"{camp_name} ({camp_tp or '유형없음'})")
            continue
        selected_powerlink_targets.append((camp_id, camp_name))
    campaign_adgroups: Dict[str, List[Dict[str, Any]]] = {}
    errors: List[str] = []
    if powerlink_targets:
        max_workers = max(1, min(FAST_IO_WORKERS, len(powerlink_targets)))
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            future_map = {
                ex.submit(_fetch_adgroups, api_key, secret_key, cid, camp_id, False): (camp_id, camp_name)
                for camp_id, camp_name in powerlink_targets
            }
            for fut in as_completed(future_map):
                camp_id, camp_name = future_map[fut]
                try:
                    res_adg, adgroup_rows = fut.result()
                except Exception as e:
                    errors.append(f"[{camp_name}] 광고그룹 조회 실패: {e}")
                    continue
                if res_adg.status_code != 200:
                    errors.append(f"[{camp_name}] 광고그룹 조회 실패: {res_adg.text}")
                    continue
                campaign_adgroups[camp_id] = adgroup_rows or []
    keyword_tasks: List[Tuple[str, str, str, str]] = []
    total_adgroup_count = 0
    for camp_id, camp_name in powerlink_targets:
        adgroup_rows = campaign_adgroups.get(camp_id) or []
        for row in adgroup_rows:
            adg_id = str(row.get("id") or row.get("nccAdgroupId") or "").strip()
            if not adg_id:
                continue
            adg_name = str(row.get("name") or adg_id)
            keyword_tasks.append((camp_id, camp_name, adg_id, adg_name))
            total_adgroup_count += 1
    campaign_keyword_count: Dict[str, int] = defaultdict(int)
    campaign_adgroup_count: Dict[str, int] = defaultdict(int)
    for camp_id, _, _, _ in keyword_tasks:
        campaign_adgroup_count[camp_id] += 1
    total_keyword_count = 0
    if keyword_tasks:
        max_workers = max(1, min(max(FAST_IO_WORKERS, 12), len(keyword_tasks)))
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            future_map = {
                ex.submit(_fetch_keywords, api_key, secret_key, cid, adg_id): (camp_id, camp_name, adg_id, adg_name)
                for camp_id, camp_name, adg_id, adg_name in keyword_tasks
            }
            for fut in as_completed(future_map):
                camp_id, camp_name, adg_id, adg_name = future_map[fut]
                try:
                    res_kw, kw_rows = fut.result()
                except Exception as e:
                    errors.append(f"[{camp_name} > {adg_name}] 키워드 조회 실패: {e}")
                    continue
                if res_kw.status_code != 200:
                    errors.append(f"[{camp_name} > {adg_name}] 키워드 조회 실패: {res_kw.text}")
                    continue
                kw_cnt = sum(1 for kw in (kw_rows or []) if str((kw or {}).get("keyword") or "").strip())
                campaign_keyword_count[camp_id] += kw_cnt
                total_keyword_count += kw_cnt
    campaign_stats = []
    for camp_id, camp_name in selected_powerlink_targets:
        campaign_stats.append({
            "campaign_id": camp_id,
            "campaign_name": camp_name,
            "keyword_count": int(campaign_keyword_count.get(camp_id) or 0),
            "adgroup_count": int(campaign_adgroup_count.get(camp_id) or 0),
        })
    campaign_stats.sort(key=lambda x: (-int(x.get("keyword_count") or 0), str(x.get("campaign_name") or "").casefold()))
    msg_parts = [
        f"현재 계정 파워링크 캠페인 {len(powerlink_targets)}개",
        f"총 등록 키워드 {total_keyword_count}개",
    ]
    if selected_ids:
        if selected_powerlink_targets:
            msg_parts.append(f"체크한 파워링크 캠페인 {len(selected_powerlink_targets)}개")
        else:
            msg_parts.append("체크한 파워링크 캠페인이 없습니다.")
    if skipped:
        msg_parts.append(f"제외 {len(skipped)}개")
    return {
        "message": " / ".join(msg_parts),
        "total_keyword_count": int(total_keyword_count),
        "total_powerlink_campaign_count": len(powerlink_targets),
        "total_powerlink_adgroup_count": int(total_adgroup_count),
        "checked_campaign_count": len(selected_ids),
        "selected_powerlink_campaign_count": len(selected_powerlink_targets),
        "campaign_stats": campaign_stats,
        "skipped": skipped,
        "errors": errors[:50],
    }
def _fetch_ads(api_key: str, secret_key: str, cid: str, adgroup_id: str):
    res = _do_req("GET", api_key, secret_key, cid, "/ncc/ads", params={"nccAdgroupId": adgroup_id})
    if res.status_code != 200:
        return res, []
    return res, res.json() or []
def _resolve_adgroup_contexts(api_key: str, secret_key: str, cid: str, entity_type: str, entity_ids: List[str]):
    contexts: List[Dict[str, Any]] = []
    warnings: List[str] = []
    seen: set[str] = set()
    entity_type = str(entity_type or "adgroup").strip().lower()
    entity_ids = _unique_keep_order(entity_ids or [])
    def _append_context(adgroup_id: str, obj: Dict[str, Any] | None, fallback_name: str = ""):
        adgroup_id = str(adgroup_id or "").strip()
        if not adgroup_id or adgroup_id in seen:
            return
        obj = obj or {}
        raw = obj.get("raw") if isinstance(obj, dict) and isinstance(obj.get("raw"), dict) else obj
        contexts.append({
            "adgroup_id": adgroup_id,
            "adgroup_type": str((obj or {}).get("adgroupType") or (raw or {}).get("adgroupType") or "").upper(),
            "adgroup_bid": (raw or {}).get("bidAmt"),
            "name": str((obj or {}).get("name") or (raw or {}).get("name") or fallback_name or ""),
        })
        seen.add(adgroup_id)
    if entity_type == "campaign":
        for camp_id in entity_ids:
            camp_id = str(camp_id or "").strip()
            if not camp_id:
                continue
            res, rows = _fetch_adgroups(api_key, secret_key, cid, camp_id)
            if res.status_code != 200:
                warnings.append(f"캠페인 {camp_id} 하위 광고그룹 조회 실패: {res.text}")
                continue
            for row in rows:
                _append_context(str(row.get("id") or row.get("nccAdgroupId") or ""), row, fallback_name=str(row.get("name") or ""))
    else:
        for adg_id in entity_ids:
            adg_id = str(adg_id or "").strip()
            if not adg_id:
                continue
            res, obj = _fetch_adgroup_detail(api_key, secret_key, cid, adg_id)
            if res.status_code != 200 or not obj:
                warnings.append(f"광고그룹 {adg_id} 조회 실패: {res.text}")
                continue
            _append_context(str(obj.get("nccAdgroupId") or adg_id), obj, fallback_name=str(obj.get("name") or adg_id))
    return contexts, warnings
def _adgroup_uses_ad_level_bid(adgroup_type: str) -> bool:
    return str(adgroup_type or "").upper() in SHOPPING_ADGROUP_TYPES_WITH_AD_LEVEL_BID
def _ad_item_has_bid_attr(ad_item: Dict[str, Any] | None) -> bool:
    return str((ad_item or {}).get("type") or "").upper() in SHOPPING_ITEM_BID_AD_TYPES
def _extract_ad_attr(ad_item: Dict[str, Any] | None) -> Dict[str, Any]:
    ad_attr = (ad_item or {}).get("adAttr")
    return copy.deepcopy(ad_attr) if isinstance(ad_attr, dict) else {}
def _resolve_effective_bid(raw_bid: Any, use_group_bid: bool, adgroup_bid: Any) -> Optional[int]:
    if bool(use_group_bid):
        group_bid = _normalize_bid_amt(adgroup_bid)
        if group_bid is not None:
            return group_bid
    return _normalize_bid_amt(raw_bid)
def _put_single_ad_with_ad_attr(api_key: str, secret_key: str, cid: str, ad_item: Dict[str, Any]):
    ad_id = _extract_ad_id(ad_item)
    cleanup_keys = ['regTm', 'editTm', 'status', 'statusReason', 'inspectStatus', 'delFlag', 'referenceKey', 'referenceData', 'referenceKeyData']
    item = copy.deepcopy(ad_item)
    for k in cleanup_keys:
        item.pop(k, None)
    res = _do_req("PUT", api_key, secret_key, cid, f"/ncc/ads/{ad_id}", params={"fields": "adAttr"}, json_body=item)
    if res.status_code in [200, 201]:
        return res
    fallback = _do_req("PUT", api_key, secret_key, cid, "/ncc/ads", params={"fields": "adAttr"}, json_body=[item])
    if fallback.status_code in [200, 201]:
        return fallback
    return res
def _apply_ad_bid_map(api_key: str, secret_key: str, cid: str, adgroup_contexts: List[Dict[str, Any]], bid_map: Dict[str, int], ad_meta: Optional[Dict[str, Dict[str, Any]]] = None):
    success_cnt = fail_cnt = skipped_cnt = 0
    err_details: List[str] = []
    ad_meta = ad_meta or {}
    for ctx in adgroup_contexts:
        if not _adgroup_uses_ad_level_bid(ctx.get("adgroup_type")):
            continue
        adg_id = str(ctx.get("adgroup_id") or "").strip()
        res_ads, ads = _fetch_ads(api_key, secret_key, cid, adg_id)
        if res_ads.status_code != 200:
            fail_cnt += 1
            if len(err_details) < 5:
                err_details.append(f"[광고그룹 {adg_id}] 소재 조회 실패: {res_ads.text}")
            continue
        for ad in (ads or []):
            if not _ad_item_has_bid_attr(ad):
                continue
            ad_id = _extract_ad_id(ad)
            if not ad_id:
                continue
            if ad_id not in bid_map:
                skipped_cnt += 1
                continue
            meta = ad_meta.get(ad_id, {})
            current_use_group = bool(meta.get("use_group_bid", _extract_ad_attr(ad).get("useGroupBidAmt")))
            current_bid = _normalize_bid_amt(meta.get("current_bid"))
            target_bid = int(bid_map[ad_id])
            if current_bid == target_bid and (not current_use_group):
                skipped_cnt += 1
                continue
            item = copy.deepcopy(ad)
            ad_attr = _extract_ad_attr(item)
            ad_attr["useGroupBidAmt"] = False
            ad_attr["bidAmt"] = target_bid
            item["adAttr"] = ad_attr
            r_put = _put_single_ad_with_ad_attr(api_key, secret_key, cid, item)
            if r_put.status_code in [200, 201]:
                success_cnt += 1
            else:
                fail_cnt += 1
                if len(err_details) < 5:
                    name = str(meta.get("name") or ((ad.get("ad") or {}).get("productName") if isinstance(ad.get("ad"), dict) else "") or ad_id)
                    err_details.append(f"[{name}] 변경 실패: {r_put.text}")
    return success_cnt, fail_cnt, skipped_cnt, err_details
def _resolve_shopping_extra_owner_ids(api_key: str, secret_key: str, cid: str, campaign_ids: List[str] | None = None, adgroup_ids: List[str] | None = None):
    campaign_ids = [str(x).strip() for x in (campaign_ids or []) if str(x).strip()]
    adgroup_ids = [str(x).strip() for x in (adgroup_ids or []) if str(x).strip()]
    candidate_adgroup_ids: List[str] = []
    seen_adgroups: set[str] = set()
    warnings: List[str] = []
    for adgroup_id in adgroup_ids:
        if adgroup_id and adgroup_id not in seen_adgroups:
            candidate_adgroup_ids.append(adgroup_id)
            seen_adgroups.add(adgroup_id)
    for campaign_id in campaign_ids:
        res_adg, rows = _fetch_adgroups(api_key, secret_key, cid, campaign_id)
        if res_adg.status_code != 200:
            warnings.append(f"캠페인 {campaign_id} 광고그룹 조회 실패: {res_adg.text}")
            continue
        for row in rows:
            raw = row.get("raw") if isinstance(row, dict) else {}
            adg_type = str((raw or {}).get("adgroupType") or row.get("adgroupType") or "").upper()
            if adg_type != "SHOPPING":
                continue
            adg_id = str(row.get("id") or row.get("nccAdgroupId") or "").strip()
            if adg_id and adg_id not in seen_adgroups:
                candidate_adgroup_ids.append(adg_id)
                seen_adgroups.add(adg_id)
    owner_ids: List[str] = []
    seen_owner_ids: set[str] = set()
    for adgroup_id in candidate_adgroup_ids:
        res_ads, ads = _fetch_ads(api_key, secret_key, cid, adgroup_id)
        if res_ads.status_code != 200:
            warnings.append(f"광고그룹 {adgroup_id} 소재 조회 실패: {res_ads.text}")
            continue
        for ad in (ads or []):
            if not isinstance(ad, dict):
                continue
            ad_type = str(ad.get("type") or ad.get("adType") or "").upper()
            if ad_type not in SHOPPING_EXTRA_AD_TYPES:
                continue
            owner_id = str(ad.get("nccAdId") or ad.get("id") or "").strip()
            if owner_id and owner_id not in seen_owner_ids:
                owner_ids.append(owner_id)
                seen_owner_ids.add(owner_id)
    return owner_ids, warnings
def _fetch_extensions(api_key: str, secret_key: str, cid: str, owner_id: str):
    res = _do_req("GET", api_key, secret_key, cid, "/ncc/ad-extensions", params={"ownerId": owner_id})
    if res.status_code != 200:
        return res, []
    return res, res.json() or []
def _normalize_lookup_scope(value: Any) -> str:
    scope = str(value or "account").strip().lower()
    if scope in {"selected_campaigns", "campaigns", "campaign_only"}:
        return "campaign"
    return "campaign" if scope == "campaign" else "account"
def _collect_asset_scope_adgroups(api_key: str, secret_key: str, cid: str, scope: str, campaign_ids: List[str] | None = None):
    scope = _normalize_lookup_scope(scope)
    selected_campaign_ids = _unique_keep_order([str(x or "").strip() for x in (campaign_ids or []) if str(x or "").strip()])
    res_camp, all_campaigns = _fetch_campaigns(api_key, secret_key, cid)
    if res_camp.status_code != 200:
        return res_camp, [], [], []
    campaign_map = {str((row or {}).get("id") or "").strip(): row for row in (all_campaigns or []) if str((row or {}).get("id") or "").strip()}
    if scope == "campaign":
        if not selected_campaign_ids:
            return _make_fake_response(400, "선택 캠페인 조회를 사용하려면 좌측에서 캠페인을 체크해주세요."), [], [], []
        target_campaigns = []
        missing_campaign_ids: List[str] = []
        for campaign_id in selected_campaign_ids:
            row = campaign_map.get(campaign_id)
            if row:
                target_campaigns.append(row)
            else:
                missing_campaign_ids.append(campaign_id)
        if missing_campaign_ids:
            for campaign_id in missing_campaign_ids:
                target_campaigns.append({"id": campaign_id, "name": campaign_id, "campaignTp": "", "label": "캠페인"})
    else:
        target_campaigns = list(all_campaigns or [])
        missing_campaign_ids = []
    contexts: List[Dict[str, Any]] = []
    warnings: List[str] = []
    if not target_campaigns:
        return _make_fake_response(200, "대상 캠페인 없음"), [], warnings, target_campaigns
    max_workers = min(FAST_IO_WORKERS, max(1, len(target_campaigns)))
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        future_map = {ex.submit(_fetch_adgroups, api_key, secret_key, cid, str(camp.get("id") or ""), False): camp for camp in target_campaigns}
        for fut in as_completed(future_map):
            camp = future_map[fut]
            camp_id = str((camp or {}).get("id") or "").strip()
            camp_name = str((camp or {}).get("name") or camp_id)
            camp_type = str((camp or {}).get("campaignTp") or "")
            try:
                res_adg, adg_rows = fut.result()
            except Exception as e:
                warnings.append(f"캠페인 {camp_name} 광고그룹 조회 실패: {e}")
                continue
            if res_adg.status_code != 200:
                warnings.append(f"캠페인 {camp_name} 광고그룹 조회 실패: {res_adg.text}")
                continue
            for adg in (adg_rows or []):
                adg_id = str((adg or {}).get("id") or (adg or {}).get("nccAdgroupId") or "").strip()
                if not adg_id:
                    continue
                contexts.append({
                    "campaign_id": camp_id,
                    "campaign_name": camp_name,
                    "campaign_type": camp_type,
                    "adgroup_id": adg_id,
                    "adgroup_name": str((adg or {}).get("name") or adg_id),
                    "adgroup_type": str((adg or {}).get("adgroupType") or ""),
                })
    if not contexts and warnings:
        return _make_fake_response(400, "\n".join(warnings[:10])), [], warnings, target_campaigns
    return _make_fake_response(200, "OK"), contexts, warnings, target_campaigns
def _summarize_lookup_ad(ad_item: Dict[str, Any] | None) -> str:
    ad_item = ad_item or {}
    ad_obj = ad_item.get("ad") if isinstance(ad_item.get("ad"), dict) else {}
    pieces = [
        str(ad_obj.get("headline") or "").strip(),
        str(ad_obj.get("productName") or "").strip(),
        str(ad_obj.get("name") or "").strip(),
        str(ad_obj.get("title") or "").strip(),
        str(ad_item.get("name") or "").strip(),
        str(ad_item.get("type") or "").strip(),
    ]
    for value in pieces:
        if value:
            return value
    return json.dumps(ad_item, ensure_ascii=False)[:120]
def _summarize_lookup_extension(ext_item: Dict[str, Any] | None) -> str:
    ext_item = ext_item or {}
    ad_ext = ext_item.get("adExtension") if isinstance(ext_item.get("adExtension"), dict) else {}
    candidates = [
        str(ad_ext.get("headline") or "").strip(),
        str(ad_ext.get("description") or "").strip(),
        str(ad_ext.get("additionalDescription") or "").strip(),
        str(ad_ext.get("promotionTitle") or "").strip(),
        str(ad_ext.get("sellerName") or "").strip(),
        str(ad_ext.get("price") or "").strip(),
        str(ext_item.get("type") or "").strip(),
    ]
    for value in candidates:
        if value:
            return value
    links = ad_ext.get("subLinks") or ad_ext.get("subLinksExtension") or []
    if isinstance(links, list) and links:
        titles = []
        for item in links[:3]:
            if isinstance(item, dict):
                title = str(item.get("title") or item.get("name") or item.get("headline") or "").strip()
                if title:
                    titles.append(title)
        if titles:
            return " / ".join(titles)
    return json.dumps(ext_item, ensure_ascii=False)[:120]
def _collect_lookup_ads_for_contexts(api_key: str, secret_key: str, cid: str, contexts: List[Dict[str, Any]]):
    rows: List[Dict[str, Any]] = []
    warnings: List[str] = []
    if not contexts:
        return rows, warnings
    max_workers = min(FAST_IO_WORKERS, max(1, len(contexts)))
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        future_map = {ex.submit(_fetch_ads, api_key, secret_key, cid, str(ctx.get("adgroup_id") or "")): ctx for ctx in contexts}
        for fut in as_completed(future_map):
            ctx = future_map[fut]
            adgroup_id = str(ctx.get("adgroup_id") or "")
            try:
                res_ads, ads = fut.result()
            except Exception as e:
                warnings.append(f"광고그룹 {adgroup_id} 소재 조회 실패: {e}")
                continue
            if res_ads.status_code != 200:
                warnings.append(f"광고그룹 {adgroup_id} 소재 조회 실패: {res_ads.text}")
                continue
            for ad in (ads or []):
                ad = ad if isinstance(ad, dict) else {}
                ad_id = str(ad.get("nccAdId") or ad.get("id") or "").strip()
                rows.append({
                    "campaignId": str(ctx.get("campaign_id") or ""),
                    "campaignName": str(ctx.get("campaign_name") or ""),
                    "campaignType": str(ctx.get("campaign_type") or ""),
                    "adgroupId": adgroup_id,
                    "adgroupName": str(ctx.get("adgroup_name") or ""),
                    "adgroupType": str(ctx.get("adgroup_type") or ""),
                    "adId": ad_id,
                    "type": str(ad.get("type") or ad.get("adType") or ""),
                    "status": "ON" if _extract_enabled_from_entity(ad) is not False else "OFF",
                    "summary": _summarize_lookup_ad(ad),
                })
    rows.sort(key=lambda x: (x.get("campaignName") or "", x.get("adgroupName") or "", x.get("type") or "", x.get("summary") or ""))
    return rows, warnings

def _collect_lookup_keywords_for_contexts(api_key: str, secret_key: str, cid: str, contexts: List[Dict[str, Any]]):
    rows: List[Dict[str, Any]] = []
    warnings: List[str] = []
    if not contexts:
        return rows, warnings
    max_workers = min(FAST_IO_WORKERS, max(1, len(contexts)))
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        future_map = {ex.submit(_fetch_keywords, api_key, secret_key, cid, str(ctx.get("adgroup_id") or "")): ctx for ctx in contexts}
        for fut in as_completed(future_map):
            ctx = future_map[fut]
            adgroup_id = str(ctx.get("adgroup_id") or "")
            try:
                res_kw, kw_rows = fut.result()
            except Exception as e:
                warnings.append(f"광고그룹 {adgroup_id} 키워드 조회 실패: {e}")
                continue
            if res_kw.status_code != 200:
                warnings.append(f"광고그룹 {adgroup_id} 키워드 조회 실패: {res_kw.text}")
                continue
            for kw in (kw_rows or []):
                kw = kw if isinstance(kw, dict) else {}
                rows.append({
                    "campaignId": str(ctx.get("campaign_id") or ""),
                    "campaignName": str(ctx.get("campaign_name") or ""),
                    "campaignType": str(ctx.get("campaign_type") or ""),
                    "adgroupId": adgroup_id,
                    "adgroupName": str(ctx.get("adgroup_name") or ""),
                    "adgroupType": str(ctx.get("adgroup_type") or ""),
                    "keywordId": str(kw.get("nccKeywordId") or kw.get("id") or "").strip(),
                    "keyword": str(kw.get("keyword") or "").strip(),
                    "status": "ON" if _extract_enabled_from_entity(kw) is not False else "OFF",
                    "bidAmt": kw.get("bidAmt"),
                    "useGroupBidAmt": bool(kw.get("useGroupBidAmt")),
                    "matchType": str(kw.get("matchType") or kw.get("type") or "").strip(),
                })
    rows.sort(key=lambda x: (x.get("campaignName") or "", x.get("adgroupName") or "", x.get("keyword") or ""))
    return rows, warnings
def _collect_lookup_extensions_for_contexts(api_key: str, secret_key: str, cid: str, contexts: List[Dict[str, Any]], ad_rows: List[Dict[str, Any]] | None = None):
    rows: List[Dict[str, Any]] = []
    warnings: List[str] = []
    owner_jobs: List[Tuple[str, Dict[str, Any], str]] = []
    seen_jobs: set[Tuple[str, str]] = set()
    for ctx in contexts or []:
        owner_id = str(ctx.get("adgroup_id") or "").strip()
        if owner_id and (owner_id, "ADGROUP") not in seen_jobs:
            owner_jobs.append((owner_id, ctx, "ADGROUP"))
            seen_jobs.add((owner_id, "ADGROUP"))
    for ad_row in (ad_rows or []):
        owner_id = str(ad_row.get("adId") or "").strip()
        if not owner_id or (owner_id, "AD") in seen_jobs:
            continue
        ctx = {
            "campaign_id": ad_row.get("campaignId"),
            "campaign_name": ad_row.get("campaignName"),
            "campaign_type": ad_row.get("campaignType"),
            "adgroup_id": ad_row.get("adgroupId"),
            "adgroup_name": ad_row.get("adgroupName"),
            "adgroup_type": ad_row.get("adgroupType"),
        }
        owner_jobs.append((owner_id, ctx, "AD"))
        seen_jobs.add((owner_id, "AD"))
    if not owner_jobs:
        return rows, warnings
    max_workers = min(FAST_IO_WORKERS, max(1, len(owner_jobs)))
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        future_map = {ex.submit(_fetch_extensions, api_key, secret_key, cid, owner_id): (owner_id, ctx, owner_scope) for owner_id, ctx, owner_scope in owner_jobs}
        for fut in as_completed(future_map):
            owner_id, ctx, owner_scope = future_map[fut]
            try:
                res_ext, ext_rows = fut.result()
            except Exception as e:
                warnings.append(f"owner {owner_id} 확장소재 조회 실패: {e}")
                continue
            if res_ext.status_code != 200:
                warnings.append(f"owner {owner_id} 확장소재 조회 실패: {res_ext.text}")
                continue
            for ext in (ext_rows or []):
                ext = ext if isinstance(ext, dict) else {}
                rows.append({
                    "campaignId": str(ctx.get("campaign_id") or ""),
                    "campaignName": str(ctx.get("campaign_name") or ""),
                    "campaignType": str(ctx.get("campaign_type") or ""),
                    "adgroupId": str(ctx.get("adgroup_id") or ""),
                    "adgroupName": str(ctx.get("adgroup_name") or ""),
                    "adgroupType": str(ctx.get("adgroup_type") or ""),
                    "ownerId": owner_id,
                    "ownerScope": "소재" if owner_scope == "AD" else "광고그룹",
                    "adExtensionId": str(ext.get("adExtensionId") or ext.get("id") or "").strip(),
                    "type": str(ext.get("type") or ext.get("adExtensionType") or ""),
                    "status": "ON" if _extract_enabled_from_entity(ext) is not False else "OFF",
                    "summary": _summarize_lookup_extension(ext),
                })
    rows.sort(key=lambda x: (x.get("campaignName") or "", x.get("adgroupName") or "", x.get("ownerScope") or "", x.get("type") or "", x.get("summary") or ""))
    return rows, warnings
def _build_asset_lookup_workbook(rows: List[Dict[str, Any]], title: str, scope_label: str, columns: List[Tuple[str, str]]):
    def _format_lookup_excel_value(key: str, label: str, value: Any) -> Any:
        if key == "type" or (isinstance(label, str) and "유형" in label):
            raw = str(value or "").strip()
            if not raw:
                return value
            normalized_ad_type = _normalize_ad_type(raw)
            ad_label = AD_TYPE_LABELS.get(normalized_ad_type)
            if ad_label:
                return ad_label
            normalized_ext_type = _normalize_extension_type(raw)
            ext_label = AD_EXTENSION_TYPE_LABELS.get(normalized_ext_type)
            if ext_label:
                return ext_label
            fallback_map = {
                "TEXT_45": "기본소재(제목+설명)",
                "RSA_AD": "반응형 검색소재",
                "SHOPPING_PRODUCT_AD": "상품소재",
                "CATALOG_PRODUCT_AD": "카탈로그 상품소재",
                "CATALOG_AD": "카탈로그 소재",
                "SHOPPING_BRAND_AD": "쇼핑브랜드 소재",
                "DESCRIPTION": "추가 설명문구",
                "DESCRIPTION_EXTRA": "설명 확장문구",
                "HEADLINE": "추가 제목",
                "IMAGE_SUB_LINKS": "이미지 서브링크",
                "POWER_LINK_IMAGE": "파워링크 이미지",
                "WEBSITE_INFO": "웹사이트 정보",
                "SUB_LINKS": "서브링크",
                "PHONE": "전화번호",
                "LOCATION": "위치정보",
                "PROMOTION": "프로모션",
                "PRICE_LINKS": "가격링크",
                "SHOPPING_PROMO_TEXT": "쇼핑 추가홍보문구",
                "SHOPPING_EXTRA": "쇼핑상품부가정보",
            }
            return fallback_map.get(raw.upper(), value)
        return value

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]
    generated_at = time.strftime("%Y-%m-%d %H:%M:%S")
    last_col = max(1, len(columns))
    ws.cell(row=1, column=1, value=title)
    ws.cell(row=1, column=1).font = Font(size=14, bold=True, color="FFFFFF")
    ws.cell(row=1, column=1).fill = PatternFill("solid", fgColor="1D4ED8")
    if last_col > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    ws.cell(row=2, column=1, value=f"생성시각: {generated_at}")
    ws.cell(row=3, column=1, value=f"조회범위: {scope_label}")
    start_row = 5
    for col_idx, (_, label) in enumerate(columns, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=label)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="111827")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, (key, label) in enumerate(columns, start=1):
            value = _format_lookup_excel_value(key, label, row.get(key))
            ws.cell(row=start_row + row_idx, column=col_idx, value=value)
    widths = {
        "campaignName": 22, "adgroupName": 22, "summary": 36, "type": 16, "status": 10,
        "adId": 24, "adExtensionId": 24, "ownerId": 24, "campaignId": 20, "adgroupId": 22, "ownerScope": 12,
    }
    for col_idx, (key, label) in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(key, max(12, min(40, len(label) + 4)))
    ws.freeze_panes = "A6"
    return wb
def _fetch_target_restrict_object(api_key: str, secret_key: str, cid: str, owner_id: str):
    res = _do_req("GET", api_key, secret_key, cid, "/ncc/targets", params={"ownerId": owner_id, "types": "RESTRICT_KEYWORD_TARGET"})
    if res.status_code != 200:
        return res, None
    data = res.json() or []
    if isinstance(data, dict):
        data = [data]
    for item in data:
        if isinstance(item, dict) and item.get("targetTp") == "RESTRICT_KEYWORD_TARGET":
            return res, item
    return res, None
def _fetch_target_object(api_key: str, secret_key: str, cid: str, owner_id: str, target_type: str):
    res = _do_req("GET", api_key, secret_key, cid, "/ncc/targets", params={"ownerId": owner_id, "types": target_type})
    if res.status_code != 200:
        return res, None
    data = res.json() or []
    if isinstance(data, dict):
        data = [data]
    for item in data:
        if isinstance(item, dict) and str(item.get("targetTp") or "").upper() == str(target_type or "").upper():
            return res, item
    return res, None
def _fetch_non_search_keyword_target(api_key: str, secret_key: str, cid: str, owner_id: str):
    return _fetch_target_object(api_key, secret_key, cid, owner_id, "NON_SEARCH_KEYWORD_TARGET")
def _get_pc_mobile_tuple(media_type: Any) -> Tuple[bool, bool]:
    media = str(media_type or "ALL").strip().upper()
    if media == "PC":
        return True, False
    if media == "MOBILE":
        return False, True
    return True, True
def _normalize_media_detail(media_detail: Any) -> Dict[str, bool]:
    raw = media_detail if isinstance(media_detail, dict) else {}
    return {
        "search_naver": _boolish(raw.get("search_naver"), True),
        "search_partner": _boolish(raw.get("search_partner"), True),
        "contents_naver": _boolish(raw.get("contents_naver"), True),
        "contents_partner": _boolish(raw.get("contents_partner"), True),
    }
def _build_media_target_payload(media_detail: Any) -> Dict[str, Any]:
    detail = _normalize_media_detail(media_detail)
    search: List[str] = []
    contents: List[str] = []
    if detail["search_naver"]:
        search.append("naver")
    if detail["search_partner"]:
        search.append("partner")
    if detail["contents_naver"]:
        contents.append("naver")
    if detail["contents_partner"]:
        contents.append("partner")
    if len(search) == 2 and len(contents) == 2:
        return {
            "type": 1,
            "search": [],
            "contents": [],
            "white": {"media": None, "mediaGroup": None},
            "black": {"media": None, "mediaGroup": None},
        }
    if not search and not contents:
        raise ValueError("세부 매체는 최소 1개 이상 선택해야 합니다.")
    return {
        "type": 2,
        "search": search,
        "contents": contents,
        "black": {"media": [], "mediaGroup": []},
        "white": {"media": None, "mediaGroup": None},
    }
def _update_pc_mobile_target(api_key: str, secret_key: str, cid: str, owner_id: str, media_type: Any):
    pc, mobile = _get_pc_mobile_tuple(media_type)
    res_target, target_obj = _fetch_target_object(api_key, secret_key, cid, owner_id, "PC_MOBILE_TARGET")
    if res_target.status_code == 200 and target_obj and target_obj.get("nccTargetId"):
        payload = {
            "nccTargetId": target_obj.get("nccTargetId"),
            "ownerId": owner_id,
            "targetTp": "PC_MOBILE_TARGET",
            "target": {"pc": pc, "mobile": mobile},
            "delFlag": False,
        }
        res_put = _do_req("PUT", api_key, secret_key, cid, f"/ncc/targets/{target_obj.get('nccTargetId')}", json_body=payload)
        if res_put.status_code in [200, 201]:
            return True, "PC_MOBILE_TARGET 업데이트 완료"
        fallback_detail = res_put.text
    else:
        fallback_detail = res_target.text if res_target is not None else "PC_MOBILE_TARGET 조회 실패"
    # 하위 호환용 fallback
    res_get = _do_req("GET", api_key, secret_key, cid, f"/ncc/adgroups/{owner_id}")
    if res_get.status_code != 200:
        return False, f"타겟/광고그룹 조회 실패: {fallback_detail}"
    obj = res_get.json() or {}
    obj["pcDevice"], obj["mobileDevice"] = pc, mobile
    res_put2 = _do_req("PUT", api_key, secret_key, cid, f"/ncc/adgroups/{owner_id}", params={"fields": "pcDevice,mobileDevice"}, json_body=obj)
    if res_put2.status_code in [200, 201]:
        return True, "pcDevice/mobileDevice fallback 업데이트 완료"
    return False, f"타겟/광고그룹 업데이트 실패: {fallback_detail} | {res_put2.text}"
def _update_media_target(api_key: str, secret_key: str, cid: str, owner_id: str, media_detail: Any):
    try:
        target_payload = _build_media_target_payload(media_detail)
    except ValueError as e:
        return False, str(e)
    res_target, target_obj = _fetch_target_object(api_key, secret_key, cid, owner_id, "MEDIA_TARGET")
    if res_target.status_code == 200 and target_obj and target_obj.get("nccTargetId"):
        payload = {
            "nccTargetId": target_obj.get("nccTargetId"),
            "ownerId": owner_id,
            "targetTp": "MEDIA_TARGET",
            "target": target_payload,
            "delFlag": False,
        }
        res_put = _do_req("PUT", api_key, secret_key, cid, f"/ncc/targets/{target_obj.get('nccTargetId')}", json_body=payload)
        if res_put.status_code in [200, 201]:
            return True, "MEDIA_TARGET 업데이트 완료"
        return False, res_put.text
    fallback_detail = res_target.text if res_target is not None else "MEDIA_TARGET 조회 실패"
    return False, f"MEDIA_TARGET 조회 실패: {fallback_detail}"
def _update_non_search_keyword_target(api_key: str, secret_key: str, cid: str, owner_id: str, excluded: bool):
    res_target, target_obj = _fetch_non_search_keyword_target(api_key, secret_key, cid, owner_id)
    if res_target.status_code != 200:
        return False, f"NON_SEARCH_KEYWORD_TARGET 조회 실패: {res_target.text}"
    if not target_obj or not target_obj.get("nccTargetId"):
        return False, "NON_SEARCH_KEYWORD_TARGET 정보를 찾지 못했습니다. 광고센터에서 해당 쇼핑 광고그룹의 제외키워드 탭을 한 번 저장한 뒤 다시 시도해주세요."
    payload = {
        "nccTargetId": target_obj.get("nccTargetId"),
        "ownerId": owner_id,
        "targetTp": "NON_SEARCH_KEYWORD_TARGET",
        "target": {"excluded": bool(excluded)},
        "delFlag": False,
    }
    res_put = _do_req("PUT", api_key, secret_key, cid, f"/ncc/targets/{target_obj.get('nccTargetId')}", json_body=payload)
    if res_put.status_code in [200, 201]:
        return True, "검색어 없는 경우 광고 노출 제외 설정 완료"
    return False, res_put.text
def _copy_target_payload_exact(api_key: str, secret_key: str, cid: str, source_owner_id: str, target_owner_id: str, target_type: str):
    res_src, src_target_obj = _fetch_target_object(api_key, secret_key, cid, source_owner_id, target_type)
    if res_src.status_code != 200:
        return False, f"{target_type} 원본 조회 실패: {res_src.text}"
    if not src_target_obj:
        return False, f"{target_type} 원본 설정 없음"
    src_target = copy.deepcopy(src_target_obj.get("target"))
    if src_target is None:
        return False, f"{target_type} 원본 target 비어 있음"
    res_dst, dst_target_obj = _fetch_target_object(api_key, secret_key, cid, target_owner_id, target_type)
    if res_dst.status_code == 200 and dst_target_obj and dst_target_obj.get("nccTargetId"):
        payload = {
            "customerId": int(cid),
            "nccTargetId": dst_target_obj.get("nccTargetId"),
            "ownerId": target_owner_id,
            "targetTp": target_type,
            "target": src_target,
            "delFlag": False,
        }
        res_put = _do_req("PUT", api_key, secret_key, cid, f"/ncc/targets/{dst_target_obj.get('nccTargetId')}", json_body=payload)
        if res_put.status_code in [200, 201]:
            return True, f"{target_type} 복사 완료"
        return False, f"{target_type} 적용 실패: {res_put.text}"
    if target_type == "PC_MOBILE_TARGET" and isinstance(src_target, dict):
        pc = bool(src_target.get("pc", True))
        mobile = bool(src_target.get("mobile", True))
        if pc and mobile:
            media_type = "ALL"
        elif pc:
            media_type = "PC"
        elif mobile:
            media_type = "MOBILE"
        else:
            return False, "PC_MOBILE_TARGET 원본이 모두 비활성화 상태입니다."
        return _update_pc_mobile_target(api_key, secret_key, cid, target_owner_id, media_type)
    create_payload = {
        "customerId": int(cid),
        "ownerId": str(target_owner_id),
        "targetTp": str(target_type),
        "target": src_target,
        "delFlag": False,
    }
    res_post = _do_req("POST", api_key, secret_key, cid, "/ncc/targets", json_body=create_payload)
    if res_post.status_code in [200, 201]:
        return True, f"{target_type} 복사 완료"
    detail = res_dst.text if res_dst is not None else '알 수 없는 오류'
    return False, f"{target_type} 생성/적용 실패: {res_post.text if res_post is not None else detail}"
def _fetch_all_target_objects(api_key: str, secret_key: str, cid: str, owner_id: str):
    res = _do_req("GET", api_key, secret_key, cid, "/ncc/targets", params={"ownerId": owner_id})
    if res.status_code != 200:
        return res, []
    data = res.json() or []
    if isinstance(data, dict):
        data = [data]
    return res, [item for item in data if isinstance(item, dict)]
def _normalize_target_type_name(target_type: Any) -> str:
    return str(target_type or "").strip().upper()
def _is_extra_copy_target_type(target_type: Any) -> bool:
    tp = _normalize_target_type_name(target_type)
    if not tp:
        return False
    if tp in {"PC_MOBILE_TARGET", "MEDIA_TARGET", "RESTRICT_KEYWORD_TARGET", "NON_SEARCH_KEYWORD_TARGET"}:
        return False
    return True
def _extract_schedule_codes_from_payload(payload: Any) -> List[str]:
    found: List[str] = []
    seen = set()
    day_token_to_num = {
        "MON": 1, "MONDAY": 1, "MO": 1, "월": 1,
        "TUE": 2, "TUESDAY": 2, "TU": 2, "화": 2,
        "WED": 3, "WEDNESDAY": 3, "WE": 3, "수": 3,
        "THU": 4, "THURSDAY": 4, "TH": 4, "목": 4,
        "FRI": 5, "FRIDAY": 5, "FR": 5, "금": 5,
        "SAT": 6, "SATURDAY": 6, "SA": 6, "토": 6,
        "SUN": 7, "SUNDAY": 7, "SU": 7, "일": 7,
    }

    def _append_code(code: Any):
        value = str(code or "").strip().upper()
        if re.fullmatch(r"SD[A-Z]{2}\d{4}", value) and value not in seen:
            seen.add(value)
            found.append(value)

    def _coerce_hour(value: Any):
        if value is None:
            return None
        if isinstance(value, str):
            m = re.search(r"(\d{1,2})", value)
            if not m:
                return None
            value = m.group(1)
        try:
            n = int(value)
        except Exception:
            return None
        return n if 0 <= n <= 24 else None

    def _coerce_day_num(value: Any):
        if value is None:
            return None
        if isinstance(value, (int, float)):
            n = int(value)
            return n if n in DAY_NUM_TO_CODE else None
        token = str(value or "").strip().upper()
        if not token:
            return None
        if token.isdigit():
            n = int(token)
            return n if n in DAY_NUM_TO_CODE else None
        return day_token_to_num.get(token)

    def visit(node: Any):
        if isinstance(node, dict):
            for k in ("dictionaryCode", "code"):
                if k in node:
                    _append_code(node.get(k))
            day_num = None
            for dk in ("day", "dayOfWeek", "weekDay", "weekday", "dayCode"):
                if dk in node:
                    day_num = _coerce_day_num(node.get(dk))
                    if day_num:
                        break
            start_h = None
            for sk in ("start", "startHour", "startTm", "startTime", "beginHour"):
                if sk in node:
                    start_h = _coerce_hour(node.get(sk))
                    if start_h is not None:
                        break
            end_h = None
            for ek in ("end", "endHour", "endTm", "endTime", "finishHour"):
                if ek in node:
                    end_h = _coerce_hour(node.get(ek))
                    if end_h is not None:
                        break
            if day_num and start_h is not None and end_h is not None and 0 <= start_h <= 23 and 1 <= end_h <= 24 and end_h > start_h:
                day_code = DAY_NUM_TO_CODE.get(day_num)
                if day_code:
                    for hour in range(start_h, end_h):
                        _append_code(f"SD{day_code}{hour:02d}{hour + 1:02d}")
            for value in node.values():
                visit(value)
            return
        if isinstance(node, list):
            for item in node:
                visit(item)
            return
        if isinstance(node, str):
            _append_code(node)

    visit(payload)
    return found

def _fetch_schedule_entries(api_key: str, secret_key: str, cid: str, owner_id: str):
    owner_id = str(owner_id or "").strip()
    if not owner_id:
        return _make_fake_response(400, "ownerId가 없습니다."), []
    res = _do_req("GET", api_key, secret_key, cid, f"/ncc/criterion/{owner_id}/SD")
    if res.status_code == 200:
        data = res.json() or []
        if isinstance(data, dict):
            data = [data]
        return res, [item for item in data if isinstance(item, dict)]
    if res.status_code == 405:
        res_adg, adgroup_obj = _fetch_adgroup_detail(api_key, secret_key, cid, owner_id)
        if res_adg.status_code == 200 and isinstance(adgroup_obj, dict):
            fallback_codes = _extract_schedule_codes_from_payload(adgroup_obj)
            if fallback_codes:
                fallback_rows = [{"dictionaryCode": code, "bidWeight": 100} for code in fallback_codes]
                return _make_fake_response(200, "SCHEDULE adgroup fallback"), fallback_rows
    return res, []
def _copy_schedule_criterion_exact(api_key: str, secret_key: str, cid: str, source_owner_id: str, target_owner_id: str):
    res_src, src_rows = _fetch_schedule_entries(api_key, secret_key, cid, source_owner_id)
    if res_src.status_code == 404:
        return False, "SCHEDULE 원본 설정 없음"
    if res_src.status_code != 200:
        return False, f"SCHEDULE 원본 조회 실패: {res_src.text}"
    codes: List[str] = []
    bid_weight_map: Dict[str, int] = {}
    for row in (src_rows or []):
        code = str(row.get("dictionaryCode") or row.get("code") or "").strip()
        if not code:
            continue
        codes.append(code)
        try:
            bid_weight = int(row.get("bidWeight") or 100)
        except Exception:
            bid_weight = 100
        bid_weight_map[code] = bid_weight
    codes = _unique_keep_order(codes)
    if not codes:
        return False, "SCHEDULE 원본 설정 없음"
    body = [{"customerId": int(cid), "ownerId": str(target_owner_id), "dictionaryCode": code, "type": "SD"} for code in codes]
    res_put = _do_req("PUT", api_key, secret_key, cid, f"/ncc/criterion/{target_owner_id}/SD", json_body=body)
    if res_put.status_code != 200:
        return False, f"SCHEDULE 적용 실패: {res_put.text}"
    weight_groups: Dict[int, List[str]] = defaultdict(list)
    for code in codes:
        try:
            weight = int(bid_weight_map.get(code, 100) or 100)
        except Exception:
            weight = 100
        weight_groups[weight].append(code)
    for bid_weight, group_codes in weight_groups.items():
        if int(bid_weight) == 100 or not group_codes:
            continue
        for i in range(0, len(group_codes), 50):
            chunk = group_codes[i:i + 50]
            res_bw = _do_req(
                "PUT",
                api_key,
                secret_key,
                cid,
                f"/ncc/criterion/{target_owner_id}/bidWeight",
                params={"codes": ",".join(chunk), "bidWeight": int(bid_weight)},
            )
            if res_bw.status_code != 200:
                return False, f"SCHEDULE 입찰가중치 적용 실패: {res_bw.text}"
    return True, "SCHEDULE 복사 완료"
def _copy_adgroup_extra_target_settings(api_key: str, secret_key: str, cid: str, source_adgroup_id: str, target_adgroup_id: str):
    messages: List[str] = []
    overall_ok = True
    res_src, src_targets = _fetch_all_target_objects(api_key, secret_key, cid, source_adgroup_id)
    if res_src.status_code == 200:
        copied_types: List[str] = []
        for src_target in (src_targets or []):
            target_type = _normalize_target_type_name(src_target.get("targetTp"))
            if not _is_extra_copy_target_type(target_type):
                continue
            ok, msg = _copy_target_payload_exact(api_key, secret_key, cid, source_adgroup_id, target_adgroup_id, target_type)
            if ok:
                copied_types.append(target_type)
            else:
                overall_ok = False
                messages.append(msg)
        if copied_types:
            messages.append("추가 타겟 복사 완료: " + ", ".join(copied_types))
    elif res_src.status_code != 404:
        overall_ok = False
        messages.append(f"추가 타겟 조회 실패: {res_src.text}")
    ok_schedule, msg_schedule = _copy_schedule_criterion_exact(api_key, secret_key, cid, source_adgroup_id, target_adgroup_id)
    if msg_schedule and msg_schedule != "SCHEDULE 원본 설정 없음":
        messages.append(msg_schedule)
    if not ok_schedule and msg_schedule not in {"", "SCHEDULE 원본 설정 없음"}:
        overall_ok = False
    return overall_ok, messages
def _resolve_bulk_target_adgroup_ids(api_key: str, secret_key: str, cid: str, target_scope: str, campaign_ids: List[str], adgroup_ids: List[str]):
    scope = str(target_scope or "adgroup").strip().lower()
    if scope == "campaign":
        resolved: List[str] = []
        warnings: List[str] = []
        for camp_id in _unique_keep_order([str(x).strip() for x in (campaign_ids or []) if str(x).strip()]):
            r_adgs, rows = _fetch_adgroups(api_key, secret_key, cid, camp_id, enrich_media=False)
            if r_adgs.status_code == 200:
                resolved.extend([str((row or {}).get("id") or (row or {}).get("nccAdgroupId") or "").strip() for row in (rows or [])])
            else:
                warnings.append(f"[{camp_id}] 하위 광고그룹 조회 실패: {r_adgs.text}")
        return _unique_keep_order([x for x in resolved if x]), warnings
    return _unique_keep_order([str(x).strip() for x in (adgroup_ids or []) if str(x).strip()]), []
def _copy_adgroup_extra_targets_only(api_key: str, secret_key: str, cid: str, source_adgroup_id: str, target_adgroup_id: str):
    messages: List[str] = []
    overall_ok = True
    res_src, src_targets = _fetch_all_target_objects(api_key, secret_key, cid, source_adgroup_id)
    if res_src.status_code == 200:
        copied_types: List[str] = []
        for src_target in (src_targets or []):
            target_type = _normalize_target_type_name(src_target.get("targetTp"))
            if not _is_extra_copy_target_type(target_type):
                continue
            ok, msg = _copy_target_payload_exact(api_key, secret_key, cid, source_adgroup_id, target_adgroup_id, target_type)
            if ok:
                copied_types.append(target_type)
            else:
                overall_ok = False
                messages.append(msg)
        if copied_types:
            messages.append("추가 타겟 복사 완료: " + ", ".join(copied_types))
    elif res_src.status_code != 404:
        overall_ok = False
        messages.append(f"추가 타겟 조회 실패: {res_src.text}")
    return overall_ok, messages
def _copy_adgroup_search_option_settings(api_key: str, secret_key: str, cid: str, source_adgroup_id: str, target_adgroup_id: str):
    try:
        res_get, src_obj = _fetch_adgroup_detail(api_key, secret_key, cid, source_adgroup_id)
    except Exception as e:
        return False, f"검색옵션 원본 조회 실패: {e}"
    if res_get.status_code != 200 or not isinstance(src_obj, dict):
        detail = res_get.text if res_get is not None else "광고그룹 조회 실패"
        return False, f"검색옵션 원본 조회 실패: {detail}"
    if str(src_obj.get("adgroupType") or "").upper() != "WEB_SITE":
        return True, "파워링크 그룹이 아니어서 검색옵션 복사 건너뜀"
    use_keyword_plus = src_obj.get("useKeywordPlus") if "useKeywordPlus" in src_obj else None
    keyword_plus_weight = src_obj.get("keywordPlusWeight") if "keywordPlusWeight" in src_obj else None
    use_close_variant = src_obj.get("useCloseVariant") if "useCloseVariant" in src_obj else None
    if use_keyword_plus is None and keyword_plus_weight is None and use_close_variant is None:
        return True, "원본 검색옵션 값이 없어 기본값 유지"
    ok, msg = _update_adgroup_search_options(
        api_key, secret_key, cid, target_adgroup_id,
        use_keyword_plus=None if use_keyword_plus is None else bool(use_keyword_plus),
        keyword_plus_weight=keyword_plus_weight,
        use_close_variant=None if use_close_variant is None else bool(use_close_variant),
    )
    return ok, msg
def _copy_adgroup_media_settings(api_key: str, secret_key: str, cid: str, source_adgroup_id: str, target_adgroup_id: str):
    messages: List[str] = []
    ok_pm, msg_pm = _copy_target_payload_exact(api_key, secret_key, cid, source_adgroup_id, target_adgroup_id, "PC_MOBILE_TARGET")
    if not ok_pm:
        try:
            res_get, src_obj = _fetch_adgroup_detail(api_key, secret_key, cid, source_adgroup_id)
            if res_get.status_code == 200 and isinstance(src_obj, dict):
                pc = bool(src_obj.get("pcDevice", True))
                mobile = bool(src_obj.get("mobileDevice", True))
                media_type = "ALL" if (pc and mobile) else ("PC" if pc else ("MOBILE" if mobile else "ALL"))
                ok_pm, msg_pm = _update_pc_mobile_target(api_key, secret_key, cid, target_adgroup_id, media_type)
        except Exception:
            pass
    if msg_pm:
        messages.append(msg_pm)
    ok_media, msg_media = _copy_target_payload_exact(api_key, secret_key, cid, source_adgroup_id, target_adgroup_id, "MEDIA_TARGET")
    if msg_media:
        messages.append(msg_media)
    return ok_pm and ok_media, messages
def _update_adgroup_search_options(api_key: str, secret_key: str, cid: str, adgroup_id: str, use_keyword_plus: Optional[bool] = None, keyword_plus_weight: Optional[int] = None, use_close_variant: Optional[bool] = None):
    if use_keyword_plus is None and keyword_plus_weight is None and use_close_variant is None:
        return True, "변경사항 없음"
    res_get = _do_req("GET", api_key, secret_key, cid, f"/ncc/adgroups/{adgroup_id}")
    if res_get.status_code != 200:
        return False, f"광고그룹 조회 실패: {res_get.text}"
    obj = res_get.json() or {}
    fields: List[str] = []
    ignored_msgs: List[str] = []
    if use_keyword_plus is not None:
        obj["useKeywordPlus"] = bool(use_keyword_plus)
        fields.append("useKeywordPlus")
        if keyword_plus_weight is None:
            keyword_plus_weight = int(obj.get("keywordPlusWeight") or 100)
        try:
            keyword_plus_weight = max(1, min(999999, int(keyword_plus_weight)))
        except Exception:
            keyword_plus_weight = 100
        obj["keywordPlusWeight"] = int(keyword_plus_weight)
        fields.append("keywordPlusWeight")
    elif keyword_plus_weight is not None:
        try:
            keyword_plus_weight = max(1, min(999999, int(keyword_plus_weight)))
        except Exception:
            keyword_plus_weight = 100
        obj["keywordPlusWeight"] = int(keyword_plus_weight)
        fields.append("keywordPlusWeight")
    if use_close_variant is not None:
        obj["useCloseVariant"] = bool(use_close_variant)
        fields.append("useCloseVariant")
    fields = list(dict.fromkeys(fields))
    if not fields:
        return True, " / ".join(ignored_msgs) if ignored_msgs else "변경사항 없음"
    def _put_with_fields(active_fields: List[str]):
        return _do_req(
            "PUT", api_key, secret_key, cid, f"/ncc/adgroups/{adgroup_id}",
            params={"fields": ",".join(active_fields)}, json_body=obj
        )
    res_put = _put_with_fields(fields)
    if res_put.status_code in [200, 201]:
        msg = "광고그룹 검색옵션 업데이트 완료"
        if ignored_msgs:
            msg += " / " + " / ".join(ignored_msgs)
        return True, msg
    detail = res_put.text if res_put is not None else ""
    if use_close_variant is not None and ("Not support modify field" in detail or '3726' in detail):
        retry_fields = [x for x in fields if x != "useCloseVariant"]
        if retry_fields:
            res_retry = _put_with_fields(retry_fields)
            if res_retry.status_code in [200, 201]:
                return True, "광고그룹 검색옵션 업데이트 완료 / 일치검색은 API 수정 미지원으로 적용 제외"
        return False, detail
    return False, detail
def _extract_restricted_rows_from_target(target_obj: Dict[str, Any]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    if not isinstance(target_obj, dict):
        return rows
    for item in target_obj.get("target") or []:
        if not isinstance(item, dict):
            continue
        kw = str(item.get("keyword") or "").strip()
        if not kw:
            continue
        rows.append({
            "keyword": kw,
            "type": item.get("type"),
            "matchType": _label_negative_type(item.get("type")),
            "ownerId": target_obj.get("ownerId"),
            "nccTargetId": target_obj.get("nccTargetId"),
        })
    return rows
def _fetch_restricted_keywords(api_key: str, secret_key: str, cid: str, adgroup_id: str):
    detail_res, adgroup_obj = _fetch_adgroup_detail(api_key, secret_key, cid, adgroup_id)
    is_shopping = _is_shopping_adgroup(adgroup_obj)
    if is_shopping:
        res_target, target_obj = _fetch_target_restrict_object(api_key, secret_key, cid, adgroup_id)
        if res_target.status_code == 200:
            rows = _extract_restricted_rows_from_target(target_obj) if target_obj else []
            rows.sort(key=lambda x: (str(x.get("keyword") or "").lower(), str(x.get("matchType") or "")))
            return res_target, rows
        return res_target, []
    rows: List[Dict[str, Any]] = []
    res = _do_req("GET", api_key, secret_key, cid, f"/ncc/adgroups/{adgroup_id}/restricted-keywords", params={"type": "KEYWORD_PLUS_RESTRICT"})
    if res.status_code == 200:
        for item in res.json() or []:
            if isinstance(item, dict):
                kw = str(item.get("keyword") or item.get("restrictedKeyword") or "").strip()
                if kw:
                    rows.append({
                        "keyword": kw,
                        "type": item.get("type") or "KEYWORD_PLUS_RESTRICT",
                        "matchType": _label_negative_type(item.get("type") or "KEYWORD_PLUS_RESTRICT"),
                        "ownerId": item.get("ownerId") or item.get("nccAdgroupId") or adgroup_id,
                    })
            elif isinstance(item, str) and item.strip():
                rows.append({"keyword": item.strip(), "type": "KEYWORD_PLUS_RESTRICT", "matchType": "일치", "ownerId": adgroup_id})
        rows.sort(key=lambda x: (str(x.get("keyword") or "").lower(), str(x.get("matchType") or "")))
        return res, rows
    if detail_res.status_code == 200:
        return _make_fake_response(200, "[]"), []
    return res, []
def _normalize_ext_compare_type(value: Any) -> str:
    s = str(value or "").strip().upper()
    if s == "SHOPPING_PROMO_TEXT":
        return "PROMOTION"
    if s == "쇼핑상품부가정보".upper():
        return "SHOPPING_EXTRA"
    return s
def _extension_matches(ext_item: Dict[str, Any], ext_type: str, data: Dict[str, Any]) -> bool:
    if not isinstance(ext_item, dict):
        return False
    item_type = _normalize_ext_compare_type(ext_item.get("type"))
    target_type = _normalize_ext_compare_type(ext_type)
    if item_type != target_type:
        return False
    ext = ext_item.get("adExtension")
    if target_type == "HEADLINE":
        return str((ext or {}).get("headline") or "").strip() == str(data.get("headline") or "").strip()
    if target_type in {"DESCRIPTION_EXTRA", "DESCRIPTION"}:
        return str((ext or {}).get("description") or "").strip() == str(data.get("description") or "").strip()
    if target_type == "PROMOTION":
        ext = ext or {}
        return (
            str(ext.get("basicText") or "").strip() == str(data.get("basicText") or "").strip()
            and str(ext.get("additionalText") or "").strip() == str(data.get("additionalText") or "").strip()
        )
    if target_type == "SUB_LINKS":
        if not isinstance(ext, list):
            return False
        left = sorted(
            (str(x.get("name") or "").strip(), str(x.get("final") or "").strip())
            for x in ext if isinstance(x, dict)
        )
        right = sorted(
            (str(x.get("name") or "").strip(), str(x.get("final") or "").strip())
            for x in (data.get("links") or []) if isinstance(x, dict)
        )
        return left == right
    if target_type == "SHOPPING_EXTRA":
        return True
    return False
def _find_existing_extension(api_key: str, secret_key: str, cid: str, owner_id: str, ext_type: str, data: Dict[str, Any]):
    res = _do_req("GET", api_key, secret_key, cid, "/ncc/ad-extensions", params={"ownerId": owner_id})
    if res.status_code != 200:
        return None
    try:
        items = res.json() or []
    except Exception:
        return None
    for item in items:
        if _extension_matches(item, ext_type, data):
            return item
    return None
def _build_extension_payload(owner_id: str, ext_type: str, data: Dict[str, Any], customer_id: int, position: int | None = None) -> Dict[str, Any]:
    ext_type = _normalize_extension_type(ext_type)
    payload: Dict[str, Any] = {"ownerId": owner_id, "customerId": customer_id, "type": ext_type}
    if ext_type == "HEADLINE":
        ad_ext = {"headline": str(data.get("headline") or "").strip()}
        if position in {1, 2}:
            # 공식 이슈 답변 기준: HEADLINE 노출 위치는 top-level priority/displayOrder가 아니라
            # adExtension.pin 값(1 또는 2)으로 지정합니다.
            ad_ext["pin"] = int(position)
        payload["adExtension"] = ad_ext
    elif ext_type in {"DESCRIPTION_EXTRA", "DESCRIPTION"}:
        payload["adExtension"] = {"description": str(data.get("description") or "").strip()}
    elif ext_type == "PROMOTION":
        promo = {"basicText": str(data.get("basicText") or "").strip()}
        additional = str(data.get("additionalText") or "").strip()
        if additional:
            promo["additionalText"] = additional
        payload["adExtension"] = promo
    elif ext_type == "SUB_LINKS":
        payload["adExtension"] = [
            {"name": str(x.get("name") or "").strip(), "final": str(x.get("final") or "").strip()}
            for x in (data.get("links") or []) if str(x.get("name") or "").strip() and str(x.get("final") or "").strip()
        ]
    elif ext_type == "SHOPPING_EXTRA":
        # 쇼핑상품부가정보는 소재(ownerId=nccAdId)에 매핑되며 adExtension 본문 없이 타입만 요청하는 방식이 가장 호환성이 높습니다.
        pass
    return payload
def _extract_extension_item(obj: Any) -> Dict[str, Any] | None:
    if isinstance(obj, dict):
        return obj
    if isinstance(obj, list):
        for item in obj:
            if isinstance(item, dict):
                return item
    return None
def _extract_headline_pin(ext_item: Dict[str, Any] | None) -> int | None:
    if not isinstance(ext_item, dict):
        return None
    ext = ext_item.get('adExtension')
    candidates: List[Any] = []
    if isinstance(ext, dict):
        candidates.extend([ext.get('pin'), ext.get('Pin'), ext.get('position'), ext.get('priority'), ext.get('displayOrder')])
    candidates.extend([ext_item.get('pin'), ext_item.get('Pin'), ext_item.get('position'), ext_item.get('priority'), ext_item.get('displayOrder')])
    for value in candidates:
        try:
            pin = int(str(value).strip())
        except Exception:
            continue
        if pin in {1, 2}:
            return pin
    return None
def _headline_pin_label(value: int | None) -> str:
    return {1: '위치1', 2: '위치2'}.get(value, '모든 위치')
def _verify_headline_pin(api_key: str, secret_key: str, cid: str, owner_id: str, headline: str, expected_position: int | None):
    existing = _find_existing_extension(api_key, secret_key, cid, owner_id, 'HEADLINE', {'headline': headline})
    if not existing:
        return {
            'ok': False,
            'item': None,
            'expected_pin': expected_position,
            'actual_pin': None,
            'detail': '등록 후 재조회에서 추가제목을 찾지 못했습니다.'
        }
    actual_pin = _extract_headline_pin(existing)
    ok = actual_pin == expected_position
    if ok:
        detail = f"재조회 검증 완료 ({_headline_pin_label(actual_pin)})"
    else:
        detail = f"재조회 검증 불일치 (요청: {_headline_pin_label(expected_position)} / 실제: {_headline_pin_label(actual_pin)})"
    return {
        'ok': ok,
        'item': existing,
        'expected_pin': expected_position,
        'actual_pin': actual_pin,
        'detail': detail,
    }
def _build_headline_pin_update_candidates(ext_item: Dict[str, Any], position: int | None, customer_id: int) -> List[Dict[str, Any]]:
    owner_id = str(ext_item.get("ownerId") or "").strip()
    ext_id = str(ext_item.get("adExtensionId") or ext_item.get("id") or "").strip()
    headline = str(((ext_item.get("adExtension") or {}) if isinstance(ext_item.get("adExtension"), dict) else {}).get("headline") or "").strip()
    if not owner_id or not ext_id or not headline:
        return []
    ad_ext: Dict[str, Any] = {"headline": headline}
    if position in {1, 2}:
        ad_ext["pin"] = int(position)
    common = {
        "ownerId": owner_id,
        "customerId": int(customer_id),
        "type": "HEADLINE",
        "adExtension": ad_ext,
    }
    for k in ("pcChannelId", "mobileChannelId", "schedule", "usePeriod", "userLock", "status"):
        if k in ext_item and ext_item.get(k) is not None:
            common[k] = ext_item.get(k)
    candidates: List[Dict[str, Any]] = []
    seen = set()
    def add(payload: Dict[str, Any]):
        key = json.dumps(payload, ensure_ascii=False, sort_keys=True)
        if key not in seen:
            seen.add(key)
            candidates.append(payload)
    add(common)
    add({
        "ownerId": owner_id,
        "customerId": int(customer_id),
        "type": "HEADLINE",
        "adExtension": ad_ext,
    })
    if position in {1, 2}:
        add({**common, "priority": int(position)})
        add({**common, "displayOrder": int(position)})
        add({**common, "priority": int(position), "displayOrder": int(position)})
        add({**common, "adExtension": {"headline": headline, "headlinePin": position == 1}})
        add({**common, "priority": int(position), "adExtension": {"headline": headline, "headlinePin": position == 1}})
        add({**common, "displayOrder": int(position), "adExtension": {"headline": headline, "headlinePin": position == 1}})
    return candidates
def _apply_headline_pin_best_effort(api_key: str, secret_key: str, cid: str, ext_item: Dict[str, Any] | None, position: int | None):
    if position not in {None, 1, 2} or not isinstance(ext_item, dict):
        return {"ok": False, "detail": "추가제목 위치 보정 대상이 아닙니다."}
    ext_id = str(ext_item.get("adExtensionId") or ext_item.get("id") or "").strip()
    if not ext_id:
        return {"ok": False, "detail": "확장소재 ID를 찾지 못했습니다."}
    tried: List[Dict[str, Any]] = []
    for payload in _build_headline_pin_update_candidates(ext_item, position, int(cid)):
        res = _do_req("PUT", api_key, secret_key, cid, f"/ncc/ad-extensions/{ext_id}", json_body=payload)
        tried.append({"status": res.status_code, "detail": res.text, "payload": payload})
        if res.status_code in [200, 201, 204]:
            item = ext_item
            try:
                parsed = res.json()
                item = _extract_extension_item(parsed) or item
            except Exception:
                pass
            label = _headline_pin_label(position)
            return {"ok": True, "item": item, "detail": f"{label} 반영 성공", "tried": tried}
    last = tried[-1] if tried else {}
    label = _headline_pin_label(position)
    return {"ok": False, "detail": last.get("detail") or f"{label} 반영 실패", "tried": tried}
def _apply_headline_position_best_effort(api_key: str, secret_key: str, cid: str, ext_item: Dict[str, Any] | None, position: int | None):
    return _apply_headline_pin_best_effort(api_key, secret_key, cid, ext_item, position)
def _shopping_extra_payload_candidates(owner_id: str, customer_id: int) -> List[Dict[str, Any]]:
    base = {"ownerId": owner_id, "customerId": int(customer_id), "type": "SHOPPING_EXTRA"}
    candidates: List[Dict[str, Any]] = [
        dict(base),
        {**base, "adExtension": {"showReviewCount": True, "showReviewRating": True, "showWishCount": True, "showPurchaseCount": True}},
        {**base, "adExtension": {"useReviewCount": True, "useReviewRating": True, "useWishCount": True, "usePurchaseCount": True}},
        {**base, "adExtension": {"reviewCount": True, "reviewRating": True, "wishCount": True, "purchaseCount": True}},
        {**base, "adExtension": {"reviewCnt": True, "reviewScore": True, "wishCnt": True, "purchaseCnt": True}},
        {**base, "adExtension": {"itemTypes": ["REVIEW_COUNT", "REVIEW_RATING", "WISH_COUNT", "PURCHASE_COUNT"]}},
        {**base, "adExtension": {"items": ["REVIEW_COUNT", "REVIEW_RATING", "WISH_COUNT", "PURCHASE_COUNT"]}},
        {**base, "adExtension": {"enabled": True}},
        {**base, "adExtension": {}},
    ]
    uniq: List[Dict[str, Any]] = []
    seen = set()
    for payload in candidates:
        key = json.dumps(payload, ensure_ascii=False, sort_keys=True)
        if key in seen:
            continue
        seen.add(key)
        uniq.append(payload)
    return uniq
def _create_shopping_extra_with_fallbacks(api_key: str, secret_key: str, cid: str, owner_id: str):
    attempts: List[Dict[str, Any]] = []
    last_res = None
    for idx, payload in enumerate(_shopping_extra_payload_candidates(owner_id, int(cid)), start=1):
        res = _do_req("POST", api_key, secret_key, cid, "/ncc/ad-extensions", params={"ownerId": owner_id}, json_body=payload)
        last_res = res
        if res.status_code in [200, 201]:
            return True, res, payload, attempts
        attempts.append({
            "try": idx,
            "status": res.status_code,
            "detail": (res.text or "")[:500],
            "payload": payload,
        })
        if res.status_code == 400 and "A record with the same name already exists." in (res.text or ""):
            return False, res, payload, attempts
    return False, last_res, (attempts[-1]["payload"] if attempts else None), attempts
def _bulk_create_campaigns(api_key: str, secret_key: str, cid: str, rows: List[Dict[str, Any]]):
    results = []
    success = fail = 0
    for idx, row in enumerate(rows, start=1):
        payload = _prepare_payload_row(row, "campaign", cid)
        name = str(payload.get("name") or f"{idx}행")
        if not payload.get("name") or not payload.get("campaignTp"):
            fail += 1
            results.append(_result_item(idx, False, name, "name / campaignTp는 필수입니다."))
            continue
        res = _do_req("POST", api_key, secret_key, cid, "/ncc/campaigns", json_body=payload)
        if res.status_code in [200, 201]:
            success += 1
            results.append(_result_item(idx, True, name, "생성 완료"))
        else:
            fail += 1
            results.append(_result_item(idx, False, name, res.text))
    return success, fail, results
def _bulk_create_adgroups(api_key: str, secret_key: str, cid: str, rows: List[Dict[str, Any]]):
    results = []
    success = fail = 0
    for idx, row in enumerate(rows, start=1):
        payload = _prepare_payload_row(row, "adgroup", cid)
        name = str(payload.get("name") or f"{idx}행")
        if not payload.get("nccCampaignId") or not payload.get("name") or not payload.get("adgroupType"):
            fail += 1
            results.append(_result_item(idx, False, name, "nccCampaignId / name / adgroupType는 필수입니다."))
            continue
        res = _do_req("POST", api_key, secret_key, cid, "/ncc/adgroups", json_body=payload)
        if res.status_code in [200, 201]:
            success += 1
            results.append(_result_item(idx, True, name, "생성 완료"))
        else:
            fail += 1
            results.append(_result_item(idx, False, name, res.text))
    return success, fail, results
def _bulk_create_keywords(api_key: str, secret_key: str, cid: str, rows: List[Dict[str, Any]]):
    grouped: Dict[str, List[Tuple[int, Dict[str, Any]]]] = defaultdict(list)
    results = []
    success = fail = 0
    for idx, row in enumerate(rows, start=1):
        payload = _prepare_payload_row(row, "keyword", cid)
        adg_id = str(payload.get("nccAdgroupId") or "").strip()
        name = str(payload.get("keyword") or f"{idx}행")
        if not adg_id or not payload.get("keyword"):
            fail += 1
            results.append(_result_item(idx, False, name, "nccAdgroupId / keyword는 필수입니다."))
            continue
        grouped[adg_id].append((idx, payload))
    batch_size = 30
    for adg_id, items in grouped.items():
        for start_idx in range(0, len(items), batch_size):
            chunk = items[start_idx:start_idx + batch_size]
            payloads = [x[1] for x in chunk]
            res = _do_req("POST", api_key, secret_key, cid, "/ncc/keywords", params={"nccAdgroupId": adg_id}, json_body=payloads)
            if res.status_code in [200, 201]:
                for idx, payload in chunk:
                    success += 1
                    results.append(_result_item(idx, True, str(payload.get("keyword")), "생성 완료"))
                continue
            with ThreadPoolExecutor(max_workers=min(8, len(chunk))) as ex:
                future_map = {
                    ex.submit(_do_req, "POST", api_key, secret_key, cid, "/ncc/keywords", {"nccAdgroupId": adg_id}, payload): (idx, payload)
                    for idx, payload in chunk
                }
                for fut in as_completed(future_map):
                    idx, payload = future_map[fut]
                    try:
                        single = fut.result()
                    except Exception as e:
                        fail += 1
                        results.append(_result_item(idx, False, str(payload.get("keyword")), str(e)))
                        continue
                    if single.status_code in [200, 201]:
                        success += 1
                        results.append(_result_item(idx, True, str(payload.get("keyword")), "생성 완료"))
                    else:
                        fail += 1
                        results.append(_result_item(idx, False, str(payload.get("keyword")), single.text))
    return success, fail, sorted(results, key=lambda x: x["row_no"])
def _post_one_ad(api_key: str, secret_key: str, cid: str, row_no: int, payload: Dict[str, Any]):
    adg_id = str(payload.get("nccAdgroupId") or "").strip()
    ad_type = str(payload.get("type") or "")
    name = str((payload.get("ad") or {}).get("headline") if isinstance(payload.get("ad"), dict) else payload.get("type") or f"{row_no}행")
    if not adg_id or not ad_type:
        return _result_item(row_no, False, name, "nccAdgroupId / type는 필수입니다.")
    body = copy.deepcopy(payload)
    if ad_type == "TEXT_45":
        ad = body.get("ad") or {}
        if not isinstance(ad, dict):
            return _result_item(row_no, False, name, "TEXT_45 소재는 ad JSON 객체가 필요합니다.")
        if isinstance(ad.get("pc"), dict) and ad["pc"].get("final") and not ad["pc"].get("display"):
            ad["pc"]["display"] = _display_url_from_final(ad["pc"]["final"])
        if isinstance(ad.get("mobile"), dict) and ad["mobile"].get("final") and not ad["mobile"].get("display"):
            ad["mobile"]["display"] = _display_url_from_final(ad["mobile"]["final"])
        missing = []
        if not ad.get("headline"):
            missing.append("headline")
        if not ad.get("description"):
            missing.append("description")
        if not (ad.get("pc") or {}).get("final"):
            missing.append("pc.final")
        if not (ad.get("mobile") or {}).get("final"):
            missing.append("mobile.final")
        if missing:
            return _result_item(row_no, False, name, f"TEXT_45 필수값 누락: {', '.join(missing)}")
        body["ad"] = ad
    if ad_type in SHOPPING_AD_TYPES:
        if "ad" not in body:
            body["ad"] = {}
        res = _do_req("POST", api_key, secret_key, cid, "/ncc/ads", params={"nccAdgroupId": adg_id, "isList": "true"}, json_body=[body])
    else:
        res = _do_req("POST", api_key, secret_key, cid, "/ncc/ads", params={"nccAdgroupId": adg_id}, json_body=body)
    if res.status_code in [200, 201]:
        return _result_item(row_no, True, name, "생성 완료")
    return _result_item(row_no, False, name, res.text)
def _bulk_create_ads(api_key: str, secret_key: str, cid: str, rows: List[Dict[str, Any]]):
    prepared = [(idx, _prepare_payload_row(row, "ad", cid)) for idx, row in enumerate(rows, start=1)]
    results = []
    with ThreadPoolExecutor(max_workers=5) as executor:
        futures = [executor.submit(_post_one_ad, api_key, secret_key, cid, idx, payload) for idx, payload in prepared]
        for future in as_completed(futures):
            results.append(future.result())
    success = sum(1 for x in results if x["ok"])
    fail = sum(1 for x in results if not x["ok"])
    return success, fail, sorted(results, key=lambda x: x["row_no"])
def _bulk_create_extensions(api_key: str, secret_key: str, cid: str, rows: List[Dict[str, Any]]):
    results = []
    success = fail = 0
    for idx, row in enumerate(rows, start=1):
        payload = _prepare_payload_row(row, "ad_extension", cid)
        owner_id = str(payload.get("ownerId") or "").strip()
        ext_type = str(payload.get("type") or "")
        name = str(payload.get("title") or ext_type or f"{idx}행")
        if not owner_id or not ext_type:
            fail += 1
            results.append(_result_item(idx, False, name, "ownerId / type는 필수입니다."))
            continue
        res = _do_req("POST", api_key, secret_key, cid, "/ncc/ad-extensions", params={"ownerId": owner_id}, json_body=payload)
        if res.status_code in [200, 201]:
            success += 1
            results.append(_result_item(idx, True, name, "생성 완료"))
        else:
            fail += 1
            results.append(_result_item(idx, False, name, res.text))
    return success, fail, results
def _merge_restricted_keyword_rows(existing_rows: List[Dict[str, Any]], new_rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    merged: List[Dict[str, Any]] = []
    seen = set()
    for item in (existing_rows or []):
        kw = str(item.get("keyword") or "").strip()
        if not kw:
            continue
        tp = _normalize_negative_type(item.get("type") or item.get("matchType") or 1)
        key = (kw.lower(), tp)
        if key not in seen:
            seen.add(key)
            merged.append({"keyword": kw, "type": tp})
    for item in (new_rows or []):
        kw = str(item.get("keyword") or "").strip()
        if not kw:
            continue
        tp = _normalize_negative_type(item.get("type") or item.get("matchType") or 1)
        key = (kw.lower(), tp)
        if key not in seen:
            seen.add(key)
            merged.append({"keyword": kw, "type": tp})
    return merged
def _upsert_shopping_restricted_keywords(api_key: str, secret_key: str, cid: str, adgroup_id: str, rows: List[Dict[str, Any]]):
    res_target, target_obj = _fetch_target_restrict_object(api_key, secret_key, cid, adgroup_id)
    if res_target.status_code != 200:
        return res_target
    if not target_obj or not target_obj.get("nccTargetId"):
        return _make_fake_response(400, "쇼핑 제외키워드 Target 정보를 찾지 못했습니다. 광고센터에서 해당 광고그룹의 제외키워드를 한 번 저장한 뒤 다시 시도해주세요.")
    merged = _merge_restricted_keyword_rows(_extract_restricted_rows_from_target(target_obj), rows)
    payload = {
        "nccTargetId": target_obj.get("nccTargetId"),
        "ownerId": adgroup_id,
        "targetTp": "RESTRICT_KEYWORD_TARGET",
        "target": merged,
        "delFlag": False,
    }
    return _do_req("PUT", api_key, secret_key, cid, f"/ncc/targets/{target_obj.get('nccTargetId')}", json_body=payload)
def _bulk_create_restricted_keywords(api_key: str, secret_key: str, cid: str, rows: List[Dict[str, Any]]):
    grouped: Dict[str, List[Tuple[int, Dict[str, Any]]]] = defaultdict(list)
    results = []
    success = fail = 0
    for idx, row in enumerate(rows, start=1):
        payload = _prepare_payload_row(row, "restricted_keyword", cid)
        adg_id = str(payload.get("nccAdgroupId") or "").strip()
        keyword = str(payload.get("keyword") or "").strip()
        if not adg_id or not keyword:
            fail += 1
            results.append(_result_item(idx, False, keyword or f"{idx}행", "nccAdgroupId / keyword는 필수입니다."))
            continue
        grouped[adg_id].append((idx, payload))
    for adg_id, items in grouped.items():
        _, adgroup_obj = _fetch_adgroup_detail(api_key, secret_key, cid, adg_id)
        is_shopping = _is_shopping_adgroup(adgroup_obj)
        if is_shopping:
            res = _upsert_shopping_restricted_keywords(api_key, secret_key, cid, adg_id, [x[1] for x in items])
            if res.status_code in [200, 201, 204]:
                for idx, payload in items:
                    success += 1
                    results.append(_result_item(idx, True, str(payload.get("keyword")), "등록 완료"))
            else:
                for idx, payload in items:
                    fail += 1
                    results.append(_result_item(idx, False, str(payload.get("keyword")), res.text))
            continue
        payloads = []
        for _, payload in items:
            payloads.append({
                "nccAdgroupId": adg_id,
                "keyword": str(payload.get("keyword") or "").strip(),
                "description": "노출 제한 키워드 추가",
                "type": "KEYWORD_PLUS_RESTRICT",
            })
        res = _do_req("POST", api_key, secret_key, cid, f"/ncc/adgroups/{adg_id}/restricted-keywords", json_body=payloads)
        if res.status_code in [200, 201, 204]:
            for idx, payload in items:
                success += 1
                results.append(_result_item(idx, True, str(payload.get("keyword")), "등록 완료"))
        else:
            for idx, payload in items:
                single = _do_req("POST", api_key, secret_key, cid, f"/ncc/adgroups/{adg_id}/restricted-keywords", json_body=[{
                    "nccAdgroupId": adg_id,
                    "keyword": str(payload.get("keyword") or "").strip(),
                    "description": "노출 제한 키워드 추가",
                    "type": "KEYWORD_PLUS_RESTRICT",
                }])
                if single.status_code in [200, 201, 204]:
                    success += 1
                    results.append(_result_item(idx, True, str(payload.get("keyword")), "등록 완료"))
                else:
                    fail += 1
                    results.append(_result_item(idx, False, str(payload.get("keyword")), single.text))
    return success, fail, sorted(results, key=lambda x: x["row_no"])
def _fetch_entity_detail(api_key: str, secret_key: str, cid: str, entity_type: str, entity_id: str):
    entity_id = str(entity_id or "").strip()
    if not entity_id:
        return _make_fake_response(400, "ID가 없습니다."), None
    uri_map = {
        "ad": f"/ncc/ads/{entity_id}",
        "ad_extension": f"/ncc/ad-extensions/{entity_id}",
    }
    uri = uri_map.get(entity_type)
    if not uri:
        return _make_fake_response(400, f"지원하지 않는 조회 유형: {entity_type}"), None
    res = _do_req("GET", api_key, secret_key, cid, uri)
    if res.status_code == 200:
        try:
            obj = res.json()
            if isinstance(obj, list):
                for one in obj:
                    if isinstance(one, dict):
                        return res, one
            if isinstance(obj, dict):
                return res, obj
        except Exception:
            pass
    return res, None
def _build_copy_name(base_name: str, copy_index: int, total_count: int, suffix: str = "_복사본") -> str:
    base = str(base_name or "").strip()
    if total_count > 1:
        return f"{base}{copy_index}" if copy_index > 1 else base
    return base + str(suffix or "")
def _parse_multiline_names(raw_value: Any) -> List[str]:
    values: List[str] = []
    if isinstance(raw_value, list):
        iterable = raw_value
    else:
        iterable = str(raw_value or "").replace("\r\n", "\n").replace("\r", "\n").split("\n")
    for item in iterable:
        name = str(item or "").strip()
        if name:
            values.append(name)
    return values
def _find_duplicate_names(names: List[str]) -> List[str]:
    seen: set[str] = set()
    dupes: List[str] = []
    for name in names:
        key = str(name or "").strip()
        if not key:
            continue
        if key in seen and key not in dupes:
            dupes.append(key)
            continue
        seen.add(key)
    return dupes
def _resolve_copy_names(source_names: List[str], copy_count: int, suffix: str, custom_names_raw: Any = None) -> List[str]:
    custom_names = _parse_multiline_names(custom_names_raw)
    total_needed = max(0, len(source_names)) * max(1, int(copy_count or 1))
    if custom_names:
        if len(custom_names) != total_needed:
            raise ValueError(f"직접 지정 이름은 총 {total_needed}개가 필요합니다. 현재 {len(custom_names)}개 입력됨")
        return custom_names
    resolved: List[str] = []
    for source_name in source_names:
        for idx in range(1, max(1, int(copy_count or 1)) + 1):
            resolved.append(_build_copy_name(source_name, idx, max(1, int(copy_count or 1)), suffix))
    return resolved
def _validate_name_candidates(candidates: List[str], existing_names: Iterable[str], label: str) -> Optional[str]:
    cleaned = [str(x or "").strip() for x in candidates if str(x or "").strip()]
    dupes = _find_duplicate_names(cleaned)
    if dupes:
        return f"중복된 {label} 이름이 있습니다: {', '.join(dupes[:10])}"
    existing = {str(x or "").strip() for x in existing_names if str(x or "").strip()}
    conflicts = [name for name in cleaned if name in existing]
    if conflicts:
        uniq_conflicts: List[str] = []
        for name in conflicts:
            if name not in uniq_conflicts:
                uniq_conflicts.append(name)
        return f"이미 존재하는 {label} 이름과 겹칩니다: {', '.join(uniq_conflicts[:10])}"
    return None
def _normalize_text_ad_payload(body: Dict[str, Any] | None) -> Dict[str, Any]:
    payload = copy.deepcopy(body) if isinstance(body, dict) else {}
    ad = payload.get("ad") if isinstance(payload.get("ad"), dict) else {}
    headline = str(ad.get("headline") or payload.get("headline") or "").strip()
    description = str(ad.get("description") or payload.get("description") or "").strip()
    pc = ad.get("pc") if isinstance(ad.get("pc"), dict) else {}
    mobile = ad.get("mobile") if isinstance(ad.get("mobile"), dict) else {}
    pc_final = str(pc.get("final") or payload.get("pcFinalUrl") or payload.get("finalUrl") or "").strip()
    mobile_final = str(mobile.get("final") or payload.get("mobileFinalUrl") or payload.get("mobileUrl") or pc_final).strip()
    payload["ad"] = {
        "headline": headline,
        "description": description,
        "pc": {
            "final": pc_final,
            "display": str(pc.get("display") or _display_url_from_final(pc_final)).strip(),
        },
        "mobile": {
            "final": mobile_final,
            "display": str(mobile.get("display") or _display_url_from_final(mobile_final)).strip(),
        },
    }
    return _strip_empty(payload)
def _make_ad_signature(ad_item: Dict[str, Any] | None) -> str:
    ad = ad_item if isinstance(ad_item, dict) else {}
    ad_type = _normalize_ad_type(ad.get("type") or ad.get("adType"))
    if ad_type == "TEXT_45":
        ad_body = ad.get("ad") if isinstance(ad.get("ad"), dict) else {}
        pc = ad_body.get("pc") if isinstance(ad_body.get("pc"), dict) else {}
        mobile = ad_body.get("mobile") if isinstance(ad_body.get("mobile"), dict) else {}
        return "|".join([
            ad_type,
            str(ad_body.get("headline") or "").strip(),
            str(ad_body.get("description") or "").strip(),
            str(pc.get("final") or ad.get("pcFinalUrl") or "").strip(),
            str(mobile.get("final") or ad.get("mobileFinalUrl") or pc.get("final") or "").strip(),
        ])
    if ad_type in SHOPPING_AD_TYPES:
        return "|".join([ad_type, str(ad.get("referenceKey") or _extract_reference_key_from_ad(ad) or "").strip()])
    try:
        return "|".join([ad_type, json.dumps(ad.get("ad") or ad, ensure_ascii=False, sort_keys=True)])
    except Exception:
        return "|".join([ad_type, str(ad)])
def _collect_ad_signatures(api_key: str, secret_key: str, cid: str, adgroup_id: str) -> set[str]:
    res_ads, ads = _fetch_ads(api_key, secret_key, cid, str(adgroup_id))
    if res_ads.status_code != 200:
        return set()
    return {_make_ad_signature(ad) for ad in (ads or []) if isinstance(ad, dict)}
def _retry_missing_ads(api_key: str, secret_key: str, cid: str, new_adg_id: str, payloads: List[Dict[str, Any]], existing_signatures: set[str] | None = None) -> Tuple[int, List[str]]:
    signatures = set(existing_signatures or set())
    success = 0
    errors: List[str] = []
    for payload in payloads or []:
        item = copy.deepcopy(payload) if isinstance(payload, dict) else {}
        if _normalize_ad_type(item.get("type")) == "TEXT_45":
            item = _normalize_text_ad_payload(item)
        sig = _make_ad_signature(item)
        if sig in signatures:
            continue
        ad_type = _normalize_ad_type(item.get("type"))
        if ad_type in SHOPPING_AD_TYPES:
            res = _do_req("POST", api_key, secret_key, cid, "/ncc/ads", params={"nccAdgroupId": new_adg_id, "isList": "true"}, json_body=[item])
        else:
            res = _do_req("POST", api_key, secret_key, cid, "/ncc/ads", params={"nccAdgroupId": new_adg_id}, json_body=item)
        if res.status_code in [200, 201]:
            success += 1
            signatures.add(sig)
        else:
            ad_name = str(((item.get("ad") or {}).get("headline") if isinstance(item.get("ad"), dict) else "") or item.get("type") or "")
            errors.append(f"소재 재시도 에러({ad_name}): {res.text}")
    return success, errors
def _build_copy_ad_payload(api_key: str, secret_key: str, cid: str, ad_item: Dict[str, Any], new_adg_id: str) -> Dict[str, Any]:
    ad_item = ad_item if isinstance(ad_item, dict) else {}
    ad_id = str(ad_item.get("nccAdId") or ad_item.get("id") or "").strip()
    detail = None
    if ad_id:
        _, detail = _fetch_entity_detail(api_key, secret_key, cid, "ad", ad_id)
    src = detail if isinstance(detail, dict) and detail else copy.deepcopy(ad_item)
    payload = _prepare_payload_row(src, "ad", cid)
    payload["nccAdgroupId"] = str(new_adg_id)
    payload["customerId"] = int(cid)
    ad_type = _normalize_ad_type(payload.get("type") or src.get("type"))
    payload["type"] = ad_type
    ref_data = src.get("referenceData") if isinstance(src.get("referenceData"), dict) else {}
    if ad_type in SHOPPING_AD_TYPES:
        payload["ad"] = payload.get("ad") if isinstance(payload.get("ad"), dict) else {}
        ref_key = (
            payload.get("referenceKey")
            or src.get("referenceKey")
            or ref_data.get("mallProductId")
            or ref_data.get("id")
        )
        if ref_key:
            payload["referenceKey"] = str(ref_key)
    else:
        payload.pop("referenceKey", None)
    payload.pop("nccAdId", None)
    payload.pop("id", None)
    payload.pop("adId", None)
    return _strip_empty(payload)
def _build_copy_extension_payload(api_key: str, secret_key: str, cid: str, ext_item: Dict[str, Any], new_owner_id: str, biz_channel_id: str | None = None) -> Dict[str, Any]:
    ext_item = ext_item if isinstance(ext_item, dict) else {}
    ext_id = str(ext_item.get("adExtensionId") or ext_item.get("id") or "").strip()
    detail = None
    if ext_id:
        _, detail = _fetch_entity_detail(api_key, secret_key, cid, "ad_extension", ext_id)
    src = detail if isinstance(detail, dict) and detail else copy.deepcopy(ext_item)
    payload = _prepare_payload_row(src, "ad_extension", cid)
    payload["ownerId"] = str(new_owner_id)
    payload["customerId"] = int(cid)
    ext_type = _normalize_extension_type(payload.get("type") or src.get("type"))
    payload["type"] = ext_type
    if biz_channel_id and biz_channel_id not in ["keep", "undefined"] and ext_type in ["SUB_LINKS", "POWER_LINK_IMAGE", "WEBSITE_INFO", "IMAGE_SUB_LINKS"]:
        payload["pcChannelId"] = str(biz_channel_id)
        payload["mobileChannelId"] = str(biz_channel_id)
    payload.pop("adExtensionId", None)
    payload.pop("id", None)
    if ext_type == "SHOPPING_EXTRA":
        payload.pop("adExtension", None)
    return _strip_empty(payload)
def _extract_created_ad_id_from_response(res: Any) -> str:
    if res is None:
        return ""
    try:
        data = res.json()
    except Exception:
        return ""
    def _walk(obj: Any) -> str:
        if isinstance(obj, dict):
            for key in ("nccAdId", "adId", "id"):
                value = str(obj.get(key) or "").strip()
                if value and (value.startswith("nad-") or value.startswith("ad-") or value.startswith("ncc") or len(value) >= 10):
                    return value
            for value in obj.values():
                found = _walk(value)
                if found:
                    return found
        elif isinstance(obj, list):
            for item in obj:
                found = _walk(item)
                if found:
                    return found
        return ""
    return _walk(data)
def _extract_reference_key_from_ad(ad_item: Dict[str, Any] | None) -> str:
    ad_item = ad_item if isinstance(ad_item, dict) else {}
    ref_data = ad_item.get("referenceData") if isinstance(ad_item.get("referenceData"), dict) else {}
    candidates = [
        ad_item.get("referenceKey"),
        ref_data.get("mallProductId"),
        ref_data.get("id"),
        ref_data.get("parentId"),
        ref_data.get("productId"),
        ref_data.get("mallProductNo"),
        ref_data.get("channelProductNo"),
        ref_data.get("productNo"),
    ]
    for value in candidates:
        s = str(value or "").strip()
        if s:
            return s
    return ""
def _classify_copy_ad_strategy(ads: List[Dict[str, Any]] | None) -> str:
    ads = [x for x in (ads or []) if isinstance(x, dict)]
    if not ads:
        return "none"
    shopping = 0
    standard = 0
    for ad in ads:
        if _normalize_ad_type(ad.get("type") or ad.get("adType")) in SHOPPING_AD_TYPES:
            shopping += 1
        else:
            standard += 1
    if shopping and standard:
        return "mixed"
    if shopping:
        return "shopping"
    return "standard"
def _find_created_ad_id_by_reference(api_key: str, secret_key: str, cid: str, adgroup_id: str, reference_key: str, before_ids: set[str] | None = None) -> str:
    reference_key = str(reference_key or "").strip()
    res_ads, ads = _fetch_ads(api_key, secret_key, cid, str(adgroup_id))
    if res_ads.status_code != 200:
        return ""
    before_ids = {str(x).strip() for x in (before_ids or set()) if str(x).strip()}
    for ad in (ads or []):
        if not isinstance(ad, dict):
            continue
        ad_id = _extract_ad_id(ad)
        if before_ids and ad_id in before_ids:
            continue
        if _extract_reference_key_from_ad(ad) == reference_key:
            return ad_id
    return ""
def _build_copy_summary(source_adgroup_id: str, target_adgroup_id: str) -> Dict[str, Any]:
    return {
        "source_adgroup_id": str(source_adgroup_id or "").strip(),
        "target_adgroup_id": str(target_adgroup_id or "").strip(),
        "strategy": "none",
        "keywords": {"source": 0, "success": 0, "fail": 0},
        "ads": {"source": 0, "success": 0, "fail": 0, "standard_source": 0, "shopping_source": 0},
        "group_extensions": {"source": 0, "success": 0, "fail": 0},
        "ad_extensions": {"source": 0, "success": 0, "fail": 0},
        "negatives": {"source": 0, "success": 0, "fail": 0},
        "notes": [],
    }
def _format_copy_summary(summary: Dict[str, Any]) -> str:
    if not isinstance(summary, dict):
        return "복사 요약 없음"
    strategy = str(summary.get("strategy") or "none")
    kw = summary.get("keywords") or {}
    ads = summary.get("ads") or {}
    gext = summary.get("group_extensions") or {}
    aext = summary.get("ad_extensions") or {}
    neg = summary.get("negatives") or {}
    notes = [str(x).strip() for x in (summary.get("notes") or []) if str(x).strip()]
    base = (
        f"전략={strategy} | 키워드 {int(kw.get('success') or 0)}/{int(kw.get('source') or 0)}"
        f" | 소재 {int(ads.get('success') or 0)}/{int(ads.get('source') or 0)}"
        f" (일반 {int(ads.get('standard_source') or 0)}, 쇼핑 {int(ads.get('shopping_source') or 0)})"
        f" | 그룹확장 {int(gext.get('success') or 0)}/{int(gext.get('source') or 0)}"
        f" | 소재하위확장 {int(aext.get('success') or 0)}/{int(aext.get('source') or 0)}"
        f" | 제외검색어 {int(neg.get('success') or 0)}/{int(neg.get('source') or 0)}"
    )
    if notes:
        base += " | 참고: " + "; ".join(notes[:3])
    return base
def _copy_ad_owner_extensions(api_key: str, secret_key: str, cid: str, old_owner_id: str, new_owner_id: str, biz_channel_id: str | None = None) -> List[str]:
    errors: List[str] = []
    old_owner_id = str(old_owner_id or "").strip()
    new_owner_id = str(new_owner_id or "").strip()
    if not old_owner_id or not new_owner_id:
        return errors
    r_ext = _do_req("GET", api_key, secret_key, cid, "/ncc/ad-extensions", params={"ownerId": old_owner_id})
    if r_ext.status_code != 200:
        if r_ext.status_code not in [404]:
            errors.append(f"소재 확장소재 조회 실패({old_owner_id}): {r_ext.text}")
        return errors
    for ext in (r_ext.json() or []):
        if not isinstance(ext, dict):
            continue
        item = _build_copy_extension_payload(api_key, secret_key, cid, ext, new_owner_id, biz_channel_id)
        res = _do_req("POST", api_key, secret_key, cid, "/ncc/ad-extensions", params={"ownerId": new_owner_id}, json_body=item)
        if res.status_code not in [200, 201] and "4003" not in res.text:
            ext_type = _normalize_extension_type(item.get("type"))
            errors.append(f"소재 확장소재 에러({ext_type}): {res.text}")
    return errors
def _copy_adgroup_children(api_key, secret_key, cid, old_adg_id, new_adg_id, biz_channel_id, include_keywords=True, include_ads=True, include_extensions=True, include_negatives=True, return_summary=False):
    errors: List[str] = []
    summary = _build_copy_summary(str(old_adg_id), str(new_adg_id))
    if include_keywords:
        r_kw = _do_req("GET", api_key, secret_key, cid, "/ncc/keywords", params={"nccAdgroupId": old_adg_id})
    else:
        r_kw = None
    if r_kw is not None and r_kw.status_code == 200:
        new_kws = []
        for kw in r_kw.json() or []:
            item = copy.deepcopy(kw)
            for k in ['nccKeywordId', 'regTm', 'editTm', 'status', 'statusReason', 'inspectStatus', 'delFlag', 'managedKeyword', 'referenceKey']:
                item.pop(k, None)
            item.update({"nccAdgroupId": str(new_adg_id), "customerId": int(cid)})
            new_kws.append(item)
        summary["keywords"]["source"] = len(new_kws)
        if new_kws:
            for i in range(0, len(new_kws), 100):
                batch = new_kws[i:i + 100]
                res = _do_req("POST", api_key, secret_key, cid, "/ncc/keywords", params={"nccAdgroupId": new_adg_id}, json_body=batch)
                if res.status_code in [200, 201]:
                    summary["keywords"]["success"] += len(batch)
                else:
                    for item in batch:
                        r_single = _do_req("POST", api_key, secret_key, cid, "/ncc/keywords", params={"nccAdgroupId": new_adg_id}, json_body=item)
                        if r_single.status_code in [200, 201]:
                            summary["keywords"]["success"] += 1
                        else:
                            summary["keywords"]["fail"] += 1
                            errors.append(f"키워드 에러: {r_single.text}")
    elif r_kw is not None and r_kw.status_code not in [200, 404]:
        errors.append(f"키워드 조회 에러: {r_kw.text}")
        summary["notes"].append("키워드 조회 실패")
    if include_ads:
        r_ad = _do_req("GET", api_key, secret_key, cid, "/ncc/ads", params={"nccAdgroupId": old_adg_id})
    else:
        r_ad = None
    copied_ad_owner_pairs: List[Tuple[str, str]] = []
    if r_ad is not None and r_ad.status_code == 200:
        ads = [x for x in (r_ad.json() or []) if isinstance(x, dict)]
        summary["strategy"] = _classify_copy_ad_strategy(ads)
        summary["ads"]["source"] = len(ads)
        summary["ads"]["shopping_source"] = sum(1 for ad in ads if _normalize_ad_type(ad.get("type") or ad.get("adType")) in SHOPPING_AD_TYPES)
        summary["ads"]["standard_source"] = int(summary["ads"]["source"] or 0) - int(summary["ads"]["shopping_source"] or 0)
        target_signatures_before = _collect_ad_signatures(api_key, secret_key, cid, str(new_adg_id))
        expected_payloads: List[Dict[str, Any]] = []
        def _post_ad(ad):
            src_ad = ad if isinstance(ad, dict) else {}
            src_ad_id = _extract_ad_id(src_ad)
            item = _build_copy_ad_payload(api_key, secret_key, cid, src_ad, str(new_adg_id))
            ad_type = _normalize_ad_type(item.get("type"))
            item.setdefault("userLock", False)
            expected_payloads.append(copy.deepcopy(item))
            before_ids: set[str] = set()
            ref_key = _extract_reference_key_from_ad(src_ad) or str(item.get("referenceKey") or "").strip()
            if ad_type in SHOPPING_AD_TYPES:
                res_before, before_ads = _fetch_ads(api_key, secret_key, cid, str(new_adg_id))
                if res_before.status_code == 200:
                    before_ids = {_extract_ad_id(x) for x in (before_ads or []) if isinstance(x, dict) and _extract_ad_id(x)}
                res = _do_req("POST", api_key, secret_key, cid, "/ncc/ads", params={"nccAdgroupId": new_adg_id, "isList": "true"}, json_body=[item])
            else:
                res = _do_req("POST", api_key, secret_key, cid, "/ncc/ads", params={"nccAdgroupId": new_adg_id}, json_body=item)
            if res.status_code not in [200, 201]:
                ad_name = str(((item.get("ad") or {}).get("headline") if isinstance(item.get("ad"), dict) else "") or item.get("type") or "")
                return {"error": f"소재 에러({ad_name}): {res.text}", "source_ad_id": src_ad_id, "created_ad_id": ""}
            created_ad_id = _extract_created_ad_id_from_response(res)
            if not created_ad_id and ref_key:
                created_ad_id = _find_created_ad_id_by_reference(api_key, secret_key, cid, str(new_adg_id), ref_key, before_ids)
            if ad_type in SHOPPING_AD_TYPES and not created_ad_id:
                ad_name = str(item.get("type") or "쇼핑소재")
                return {"error": f"소재 에러({ad_name}): 생성 후 소재 ID 확인 실패", "source_ad_id": src_ad_id, "created_ad_id": ""}
            return {"error": None, "source_ad_id": src_ad_id, "created_ad_id": created_ad_id}
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = [executor.submit(_post_ad, ad) for ad in ads]
            for f in as_completed(futures):
                result = f.result() or {}
                err_msg = result.get("error")
                if err_msg:
                    summary["ads"]["fail"] += 1
                    errors.append(err_msg)
                    continue
                summary["ads"]["success"] += 1
                src_ad_id = str(result.get("source_ad_id") or "").strip()
                created_ad_id = str(result.get("created_ad_id") or "").strip()
                if src_ad_id and created_ad_id:
                    copied_ad_owner_pairs.append((src_ad_id, created_ad_id))
        target_signatures_after = _collect_ad_signatures(api_key, secret_key, cid, str(new_adg_id))
        missing_payloads = []
        for payload in expected_payloads:
            sig = _make_ad_signature(payload)
            if sig and sig not in target_signatures_after:
                missing_payloads.append(payload)
        if missing_payloads:
            retry_success, retry_errors = _retry_missing_ads(
                api_key, secret_key, cid, str(new_adg_id), missing_payloads,
                existing_signatures=(target_signatures_after or target_signatures_before),
            )
            summary["ads"]["success"] += int(retry_success or 0)
            summary["ads"]["fail"] += len(retry_errors)
            errors.extend(retry_errors)
            if retry_success:
                summary["notes"].append(f"누락 소재 재시도 성공: {retry_success}건")
    elif r_ad is not None and r_ad.status_code not in [200, 404]:
        errors.append(f"소재 조회 에러: {r_ad.text}")
        summary["notes"].append("소재 조회 실패")
    if include_extensions:
        r_ext = _do_req("GET", api_key, secret_key, cid, "/ncc/ad-extensions", params={"ownerId": old_adg_id})
    else:
        r_ext = None
    if r_ext is not None and r_ext.status_code == 200:
        group_exts = [x for x in (r_ext.json() or []) if isinstance(x, dict)]
        summary["group_extensions"]["source"] = len(group_exts)
        for ext in group_exts:
            item = _build_copy_extension_payload(api_key, secret_key, cid, ext, str(new_adg_id), biz_channel_id)
            res = _do_req("POST", api_key, secret_key, cid, "/ncc/ad-extensions", params={"ownerId": new_adg_id}, json_body=item)
            if res.status_code not in [200, 201] and "4003" not in res.text:
                summary["group_extensions"]["fail"] += 1
                ext_type = _normalize_extension_type(item.get("type"))
                errors.append(f"확장소재 에러({ext_type}): {res.text}")
            else:
                summary["group_extensions"]["success"] += 1
    elif r_ext is not None and r_ext.status_code not in [200, 404]:
        errors.append(f"광고그룹 확장소재 조회 에러: {r_ext.text}")
        summary["notes"].append("광고그룹 확장소재 조회 실패")
    if include_extensions and copied_ad_owner_pairs:
        seen_owner_pairs = set()
        for old_owner_id, new_owner_id in copied_ad_owner_pairs:
            pair = (str(old_owner_id), str(new_owner_id))
            if not pair[0] or not pair[1] or pair in seen_owner_pairs:
                continue
            seen_owner_pairs.add(pair)
            old_res, old_items = _fetch_extensions(api_key, secret_key, cid, pair[0])
            if old_res.status_code == 200:
                summary["ad_extensions"]["source"] += len([x for x in (old_items or []) if isinstance(x, dict)])
            child_errors = _copy_ad_owner_extensions(api_key, secret_key, cid, pair[0], pair[1], biz_channel_id)
            if child_errors:
                summary["ad_extensions"]["fail"] += len(child_errors)
                errors.extend(child_errors)
            elif old_res.status_code == 200:
                summary["ad_extensions"]["success"] += len([x for x in (old_items or []) if isinstance(x, dict)])
    if summary["ad_extensions"]["source"] < summary["ad_extensions"]["success"]:
        summary["ad_extensions"]["source"] = summary["ad_extensions"]["success"]
    if include_negatives:
        res_rk, rows_rk = _fetch_restricted_keywords(api_key, secret_key, cid, str(old_adg_id))
        if res_rk.status_code == 200 and rows_rk:
            post_rows = []
            for rk in rows_rk:
                if not isinstance(rk, dict):
                    continue
                kw = str(rk.get("keyword") or rk.get("restrictedKeyword") or "").strip()
                if not kw:
                    continue
                post_rows.append({
                    "nccAdgroupId": str(new_adg_id),
                    "customerId": int(cid),
                    "keyword": kw,
                    "type": _normalize_negative_type(rk.get("type") or rk.get("matchType") or 1),
                })
            summary["negatives"]["source"] = len(post_rows)
            if post_rows:
                neg_success, neg_fail, neg_results = _bulk_create_restricted_keywords(api_key, secret_key, cid, post_rows)
                summary["negatives"]["success"] = int(neg_success or 0)
                summary["negatives"]["fail"] = int(neg_fail or 0)
                if neg_fail > 0:
                    fail_msgs = [str(x.get("message") or "") for x in (neg_results or []) if not x.get("success")]
                    if fail_msgs:
                        errors.extend([f"제외검색어 에러: {msg}" for msg in fail_msgs[:10]])
                    else:
                        errors.append("제외검색어 에러: 일부 제외검색어 복사 실패")
        elif res_rk.status_code not in [200, 404]:
            errors.append(f"제외검색어 조회 에러: {res_rk.text}")
            summary["notes"].append("제외검색어 조회 실패")
    dedup_errors = list(dict.fromkeys([str(x) for x in errors if str(x).strip()]))
    if return_summary:
        return dedup_errors, summary
    return dedup_errors
def _extract_adgroup(src, target_camp_id, cid, biz_channel_id):
    res = {
        "nccCampaignId": str(target_camp_id),
        "customerId": int(cid),
        "name": src.get("name"),
        "adgroupType": src.get("adgroupType"),
        "useDailyBudget": src.get("useDailyBudget", False),
        "dailyBudget": src.get("dailyBudget", 0),
        "bidAmt": src.get("bidAmt", 70),
    }
    adgroup_type = src.get("adgroupType", "")
    if biz_channel_id and biz_channel_id not in ["keep", "undefined"] and adgroup_type == "WEB_SITE":
        res["pcChannelId"] = res["mobileChannelId"] = str(biz_channel_id)
    else:
        if src.get("pcChannelId"):
            res["pcChannelId"] = str(src.get("pcChannelId"))
        if src.get("mobileChannelId"):
            res["mobileChannelId"] = str(src.get("mobileChannelId"))
    for k in ["useStoreUrl", "nccProductGroupId", "contentsNetworkBidAmt", "keywordPlusFlag", "contractId"]:
        if k in src:
            res[k] = src[k]
    if str(src.get("adgroupType") or "").upper() == "WEB_SITE":
        for k in ["useKeywordPlus", "keywordPlusWeight", "useCloseVariant"]:
            if k in src:
                res[k] = src.get(k)
    return res
def _delete_entity_by_id(api_key: str, secret_key: str, cid: str, entity_type: str, entity_id: str):
    entity_id = str(entity_id or "").strip()
    if not entity_id:
        return _make_fake_response(400, "ID가 없습니다.")
    uri_map = {
        "campaign": f"/ncc/campaigns/{entity_id}",
        "adgroup": f"/ncc/adgroups/{entity_id}",
        "keyword": f"/ncc/keywords/{entity_id}",
        "ad": f"/ncc/ads/{entity_id}",
        "ad_extension": f"/ncc/ad-extensions/{entity_id}",
    }
    uri = uri_map.get(entity_type)
    if not uri:
        return _make_fake_response(400, f"지원하지 않는 삭제 유형: {entity_type}")
    return _do_req("DELETE", api_key, secret_key, cid, uri)
@app.route("/")
def index():
    accounts = []
    csv_path = os.path.join(BASE_DIR, "accounts.csv")
    if os.path.exists(csv_path):
        try:
            df = pd.read_csv(csv_path, encoding="utf-8-sig")
        except Exception:
            df = pd.read_csv(csv_path, encoding="cp949")
        cols = {c.lower().strip(): c for c in df.columns}
        cid_col = cols.get("customer_id") or cols.get("customerid") or df.columns[0]
        name_col = cols.get("account_name") or cols.get("name") or (df.columns[1] if len(df.columns) > 1 else df.columns[0])
        df2 = df[[cid_col, name_col]].copy()
        df2.columns = ["customer_id", "account_name"]
        df2["customer_id"] = df2["customer_id"].astype(str).str.strip()
        df2["account_name"] = df2["account_name"].astype(str).str.strip()
        accounts = df2.to_dict(orient="records")
    return render_template("index.html", accounts=accounts)
@app.route("/health", methods=["GET"])
def health():
    return jsonify({"ok": True, "service": "naver-bulk-manager-refresh"})
@app.route("/get_campaigns", methods=["POST"])
def get_campaigns():
    d = request.json or {}
    api_key = d.get("api_key")
    secret_key = d.get("secret_key")
    cid = d.get("customer_id")
    force = bool(d.get("force"))
    cache_key = _stable_cache_key(api_key, secret_key, cid, "campaigns")
    if not force:
        cached = _cache_get(_CAMPAIGN_CACHE, cache_key)
        if cached is not None:
            return jsonify(cached)
    res, rows = _fetch_campaigns(api_key, secret_key, cid)
    if res.status_code == 200:
        _cache_set(_CAMPAIGN_CACHE, cache_key, rows)
        return jsonify(rows)
    return jsonify({"error": "캠페인 조회 실패", "details": res.text}), 400
@app.route("/get_adgroups", methods=["POST"])
def get_adgroups():
    d = request.json or {}
    api_key = d.get("api_key")
    secret_key = d.get("secret_key")
    cid = d.get("customer_id")
    campaign_id = str(d.get("campaign_id") or "").strip()
    force = bool(d.get("force"))
    if not campaign_id:
        return jsonify([])
    cache_key = _stable_cache_key(api_key, secret_key, cid, f"adgroups::{campaign_id}")
    if not force:
        cached = _cache_get(_ADGROUP_CACHE, cache_key)
        if cached is not None:
            return jsonify(cached)
    res, rows = _fetch_adgroups(api_key, secret_key, cid, campaign_id, enrich_media=False)
    if res.status_code == 200:
        _cache_set(_ADGROUP_CACHE, cache_key, rows)
        return jsonify(rows)
    return jsonify({"error": "광고그룹 조회 실패", "details": res.text}), 400
@app.route("/get_biz_channels", methods=["POST"])
def get_biz_channels():
    d = request.json or {}
    api_key = d.get("api_key")
    secret_key = d.get("secret_key")
    cid = d.get("customer_id")
    force = bool(d.get("force"))
    cache_key = _stable_cache_key(api_key, secret_key, cid, "channels")
    if not force:
        cached = _cache_get(_CHANNEL_CACHE, cache_key)
        if cached is not None:
            return jsonify(cached)
    res = _do_req("GET", api_key, secret_key, cid, "/ncc/channels")
    if res.status_code == 200:
        normalized = [_normalize_channel_item(item) for item in (res.json() or [])]
        _cache_set(_CHANNEL_CACHE, cache_key, normalized)
        return jsonify(normalized)
    return jsonify({"error": "비즈채널 조회 실패", "details": res.text}), 400
@app.route("/get_keywords", methods=["POST"])
def get_keywords():
    d = request.json or {}
    res, rows = _fetch_keywords(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), d.get("adgroup_id"))
    if res.status_code == 200:
        return jsonify(rows)
    return jsonify({"error": "키워드 조회 실패", "details": res.text}), 400
@app.route("/get_ads", methods=["POST"])
def get_ads():
    d = request.json or {}
    res, rows = _fetch_ads(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), d.get("adgroup_id"))
    if res.status_code == 200:
        return jsonify(rows)
    return jsonify({"error": "소재 조회 실패", "details": res.text}), 400
@app.route("/get_ad_extensions", methods=["POST"])
def get_ad_extensions():
    d = request.json or {}
    res, rows = _fetch_extensions(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), d.get("owner_id"))
    if res.status_code == 200:
        return jsonify(rows)
    return jsonify({"error": "확장소재 조회 실패", "details": res.text}), 400
@app.route("/query_account_ads", methods=["POST"])
def query_account_ads():
    d = request.json or {}
    api_key = str(d.get("api_key") or "").strip()
    secret_key = str(d.get("secret_key") or "").strip()
    cid = str(d.get("customer_id") or "").strip()
    scope = _normalize_lookup_scope(d.get("scope") or d.get("search_scope"))
    campaign_ids = [str(x or "").strip() for x in (d.get("campaign_ids") or []) if str(x or "").strip()]
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API 정보 및 광고주를 선택해주세요."}), 400
    res_ctx, contexts, warnings, _ = _collect_asset_scope_adgroups(api_key, secret_key, cid, scope, campaign_ids)
    if res_ctx.status_code != 200:
        return jsonify({"error": "소재 조회 실패", "details": res_ctx.text}), 400
    rows, row_warnings = _collect_lookup_ads_for_contexts(api_key, secret_key, cid, contexts)
    warnings.extend(row_warnings)
    return jsonify({
        "ok": True,
        "scope": scope,
        "total": len(rows),
        "preview_limit": 200,
        "rows": rows,
        "warnings": warnings[:20],
        "message": f"소재 {len(rows):,}건 조회 완료",
    })
@app.route("/query_account_extensions", methods=["POST"])
def query_account_extensions():
    d = request.json or {}
    api_key = str(d.get("api_key") or "").strip()
    secret_key = str(d.get("secret_key") or "").strip()
    cid = str(d.get("customer_id") or "").strip()
    scope = _normalize_lookup_scope(d.get("scope") or d.get("search_scope"))
    campaign_ids = [str(x or "").strip() for x in (d.get("campaign_ids") or []) if str(x or "").strip()]
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API 정보 및 광고주를 선택해주세요."}), 400
    res_ctx, contexts, warnings, _ = _collect_asset_scope_adgroups(api_key, secret_key, cid, scope, campaign_ids)
    if res_ctx.status_code != 200:
        return jsonify({"error": "확장소재 조회 실패", "details": res_ctx.text}), 400
    ad_rows, ad_warnings = _collect_lookup_ads_for_contexts(api_key, secret_key, cid, contexts)
    warnings.extend(ad_warnings)
    rows, row_warnings = _collect_lookup_extensions_for_contexts(api_key, secret_key, cid, contexts, ad_rows=ad_rows)
    warnings.extend(row_warnings)
    return jsonify({
        "ok": True,
        "scope": scope,
        "total": len(rows),
        "preview_limit": 200,
        "rows": rows,
        "warnings": warnings[:20],
        "message": f"확장소재 {len(rows):,}건 조회 완료",
    })
@app.route("/query_account_keywords", methods=["POST"])
def query_account_keywords():
    d = request.json or {}
    api_key = str(d.get("api_key") or "").strip()
    secret_key = str(d.get("secret_key") or "").strip()
    cid = str(d.get("customer_id") or "").strip()
    scope = _normalize_lookup_scope(d.get("scope") or d.get("search_scope"))
    campaign_ids = [str(x or "").strip() for x in (d.get("campaign_ids") or []) if str(x or "").strip()]
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API 정보 및 광고주를 선택해주세요."}), 400
    res_ctx, contexts, warnings, _ = _collect_asset_scope_adgroups(api_key, secret_key, cid, scope, campaign_ids)
    if res_ctx.status_code != 200:
        return jsonify({"error": "키워드 조회 실패", "details": res_ctx.text}), 400
    rows, row_warnings = _collect_lookup_keywords_for_contexts(api_key, secret_key, cid, contexts)
    warnings.extend(row_warnings)
    return jsonify({
        "ok": True,
        "scope": scope,
        "total": len(rows),
        "preview_limit": 200,
        "rows": rows,
        "warnings": warnings[:20],
        "message": f"등록 키워드 {len(rows):,}건 조회 완료",
    })

@app.route("/export_account_keywords_excel", methods=["POST"])
def export_account_keywords_excel():
    d = request.json or {}
    api_key = str(d.get("api_key") or "").strip()
    secret_key = str(d.get("secret_key") or "").strip()
    cid = str(d.get("customer_id") or "").strip()
    scope = _normalize_lookup_scope(d.get("scope") or d.get("search_scope"))
    campaign_ids = [str(x or "").strip() for x in (d.get("campaign_ids") or []) if str(x or "").strip()]
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API 정보 및 광고주를 선택해주세요."}), 400
    res_ctx, contexts, warnings, _ = _collect_asset_scope_adgroups(api_key, secret_key, cid, scope, campaign_ids)
    if res_ctx.status_code != 200:
        return jsonify({"error": "키워드 엑셀 다운로드 실패", "details": res_ctx.text}), 400
    rows, row_warnings = _collect_lookup_keywords_for_contexts(api_key, secret_key, cid, contexts)
    warnings.extend(row_warnings)
    if not rows:
        return jsonify({"error": "내보낼 등록 키워드가 없습니다."}), 400
    scope_label = "선택 캠페인만" if scope == "campaign" else "계정 전체"
    wb = _build_asset_lookup_workbook(rows, "계정 등록 키워드 조회", scope_label, [
        ("campaignName", "캠페인명"), ("adgroupName", "광고그룹명"), ("keyword", "키워드"), ("status", "상태"),
        ("bidAmt", "입찰가"), ("useGroupBidAmt", "그룹입찰가사용"), ("matchType", "매치유형"),
        ("keywordId", "키워드 ID"), ("campaignId", "캠페인 ID"), ("adgroupId", "광고그룹 ID"),
    ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"account_keywords_{scope}_{stamp}.xlsx")

@app.route("/export_account_ads_excel", methods=["POST"])
def export_account_ads_excel():
    d = request.json or {}
    api_key = str(d.get("api_key") or "").strip()
    secret_key = str(d.get("secret_key") or "").strip()
    cid = str(d.get("customer_id") or "").strip()
    scope = _normalize_lookup_scope(d.get("scope") or d.get("search_scope"))
    campaign_ids = [str(x or "").strip() for x in (d.get("campaign_ids") or []) if str(x or "").strip()]
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API 정보 및 광고주를 선택해주세요."}), 400
    res_ctx, contexts, warnings, _ = _collect_asset_scope_adgroups(api_key, secret_key, cid, scope, campaign_ids)
    if res_ctx.status_code != 200:
        return jsonify({"error": "소재 엑셀 다운로드 실패", "details": res_ctx.text}), 400
    rows, row_warnings = _collect_lookup_ads_for_contexts(api_key, secret_key, cid, contexts)
    warnings.extend(row_warnings)
    if not rows:
        return jsonify({"error": "내보낼 소재가 없습니다."}), 400
    scope_label = "선택 캠페인만" if scope == "campaign" else "계정 전체"
    wb = _build_asset_lookup_workbook(rows, "계정 등록 소재 조회", scope_label, [
        ("campaignName", "캠페인명"), ("adgroupName", "광고그룹명"), ("type", "소재유형"), ("status", "상태"),
        ("summary", "요약"), ("adId", "소재 ID"), ("campaignId", "캠페인 ID"), ("adgroupId", "광고그룹 ID"),
    ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"account_ads_{scope}_{stamp}.xlsx")
@app.route("/export_account_extensions_excel", methods=["POST"])
def export_account_extensions_excel():
    d = request.json or {}
    api_key = str(d.get("api_key") or "").strip()
    secret_key = str(d.get("secret_key") or "").strip()
    cid = str(d.get("customer_id") or "").strip()
    scope = _normalize_lookup_scope(d.get("scope") or d.get("search_scope"))
    campaign_ids = [str(x or "").strip() for x in (d.get("campaign_ids") or []) if str(x or "").strip()]
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API 정보 및 광고주를 선택해주세요."}), 400
    res_ctx, contexts, warnings, _ = _collect_asset_scope_adgroups(api_key, secret_key, cid, scope, campaign_ids)
    if res_ctx.status_code != 200:
        return jsonify({"error": "확장소재 엑셀 다운로드 실패", "details": res_ctx.text}), 400
    ad_rows, ad_warnings = _collect_lookup_ads_for_contexts(api_key, secret_key, cid, contexts)
    warnings.extend(ad_warnings)
    rows, row_warnings = _collect_lookup_extensions_for_contexts(api_key, secret_key, cid, contexts, ad_rows=ad_rows)
    warnings.extend(row_warnings)
    if not rows:
        return jsonify({"error": "내보낼 확장소재가 없습니다."}), 400
    scope_label = "선택 캠페인만" if scope == "campaign" else "계정 전체"
    wb = _build_asset_lookup_workbook(rows, "계정 등록 확장소재 조회", scope_label, [
        ("campaignName", "캠페인명"), ("adgroupName", "광고그룹명"), ("ownerScope", "적용대상"), ("type", "확장소재유형"),
        ("status", "상태"), ("summary", "요약"), ("adExtensionId", "확장소재 ID"), ("ownerId", "owner ID"),
        ("campaignId", "캠페인 ID"), ("adgroupId", "광고그룹 ID"),
    ])
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    stamp = time.strftime("%Y%m%d_%H%M%S")
    return send_file(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", as_attachment=True, download_name=f"account_extensions_{scope}_{stamp}.xlsx")
@app.route("/get_restricted_keywords", methods=["POST"])
def get_restricted_keywords():
    d = request.json or {}
    res, rows = _fetch_restricted_keywords(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), d.get("adgroup_id"))
    if res.status_code == 200:
        return jsonify(rows)
    return jsonify({"error": "제외키워드 조회 실패", "details": res.text}), 400
@app.route("/find_powerlink_duplicate_keywords", methods=["POST"])
def find_powerlink_duplicate_keywords():
    d = request.json or {}
    api_key = d.get("api_key")
    secret_key = d.get("secret_key")
    cid = d.get("customer_id")
    campaign_ids = d.get("campaign_ids") or []
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API 정보 및 광고주를 선택해주세요."}), 400
    if not campaign_ids:
        return jsonify({"error": "캠페인을 1개 이상 선택해주세요."}), 400
    try:
        result = _find_powerlink_duplicate_keywords(api_key, secret_key, cid, campaign_ids)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"error": "중복 키워드 조회 실패", "details": str(e)}), 400
@app.route("/find_account_powerlink_duplicate_keywords", methods=["POST"])
def find_account_powerlink_duplicate_keywords():
    d = request.json or {}
    api_key = d.get("api_key")
    secret_key = d.get("secret_key")
    cid = d.get("customer_id")
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API 정보 및 광고주를 선택해주세요."}), 400
    try:
        result = _find_account_powerlink_duplicate_keywords(api_key, secret_key, cid)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"error": "계정 중복 키워드 조회 실패", "details": str(e)}), 400
@app.route("/get_powerlink_keyword_stats", methods=["POST"])
def get_powerlink_keyword_stats():
    d = request.json or {}
    api_key = d.get("api_key")
    secret_key = d.get("secret_key")
    cid = d.get("customer_id")
    campaign_ids = d.get("campaign_ids") or []
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API 정보 및 광고주를 선택해주세요."}), 400
    try:
        result = _get_powerlink_keyword_stats(api_key, secret_key, cid, campaign_ids)
        return jsonify({"ok": True, **result})
    except Exception as e:
        return jsonify({"error": "파워링크 키워드 개수 조회 실패", "details": str(e)}), 400
@app.route("/search_powerlink_keywords", methods=["POST"])
def search_powerlink_keywords():
    d = request.json or {}
    api_key = str(d.get("api_key") or "").strip()
    secret_key = str(d.get("secret_key") or "").strip()
    cid = str(d.get("customer_id") or "").strip()
    search_text = str(d.get("search_text") or d.get("keyword_query") or "").strip()
    exact_match = _boolish(d.get("exact_match"), False)
    exclude_text = str(d.get("exclude_text") or d.get("exclude_keyword_query") or "").strip()
    search_scope = str(d.get("search_scope") or "account").strip().lower()
    campaign_ids = [str(x or "").strip() for x in (d.get("campaign_ids") or []) if str(x or "").strip()]
    adgroup_ids = [str(x or "").strip() for x in (d.get("adgroup_ids") or []) if str(x or "").strip()]
    adgroup_ids = [str(x or "").strip() for x in (d.get("adgroup_ids") or []) if str(x or "").strip()]
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API 정보 및 광고주를 선택해주세요."}), 400
    if not search_text:
        return jsonify({"error": "검색어를 입력해주세요."}), 400
    if search_scope not in {"account", "selected_campaigns", "selected_adgroups"}:
        search_scope = "account"
    if search_scope == "selected_campaigns" and not campaign_ids:
        return jsonify({"error": "선택 캠페인 조회를 사용하려면 좌측에서 캠페인을 체크해주세요."}), 400
    if search_scope == "selected_adgroups" and not adgroup_ids:
        return jsonify({"error": "선택 그룹내 조회를 사용하려면 좌측에서 광고그룹을 체크해주세요."}), 400
    try:
        scan = _scan_powerlink_keywords_by_search(
            api_key,
            secret_key,
            cid,
            search_text,
            exact_match=exact_match,
            campaign_ids=(campaign_ids if search_scope == "selected_campaigns" else None),
            adgroup_ids=(adgroup_ids if search_scope == "selected_adgroups" else None),
            exclude_text=exclude_text,
        )
        if search_scope == "selected_campaigns":
            selected_set = set(campaign_ids)
            filtered_rows = [row for row in (scan.get("matched_rows") or []) if str(row.get("campaign_id") or "").strip() in selected_set]
            scan = dict(scan)
            scan["matched_rows"] = filtered_rows
            scan["matched_count"] = len(filtered_rows)
            scan["powerlink_campaigns"] = [row for row in (scan.get("powerlink_campaigns") or []) if str(row.get("id") or "").strip() in selected_set]
            scan["adgroup_contexts"] = [row for row in (scan.get("adgroup_contexts") or []) if str(row.get("campaign_id") or "").strip() in selected_set]
        elif search_scope == "selected_adgroups":
            selected_adgroup_set = set(adgroup_ids)
            filtered_rows = [row for row in (scan.get("matched_rows") or []) if str(row.get("adgroup_id") or "").strip() in selected_adgroup_set]
            scan = dict(scan)
            scan["matched_rows"] = filtered_rows
            scan["matched_count"] = len(filtered_rows)
            scan["adgroup_contexts"] = [row for row in (scan.get("adgroup_contexts") or []) if str(row.get("adgroup_id") or "").strip() in selected_adgroup_set]
            selected_campaign_set = {str(row.get("campaign_id") or "").strip() for row in (scan.get("adgroup_contexts") or []) if str(row.get("campaign_id") or "").strip()}
            scan["powerlink_campaigns"] = [row for row in (scan.get("powerlink_campaigns") or []) if str(row.get("id") or "").strip() in selected_campaign_set]
        preview_token = hashlib.sha256(f"{cid}|{search_scope}|{'/'.join(campaign_ids)}|{'/'.join(adgroup_ids)}|{search_text}|{exclude_text}|{'1' if exact_match else '0'}|{int(scan.get('matched_count') or 0)}".encode('utf-8')).hexdigest()[:20]
        scope_label = "선택 캠페인 기준" if search_scope == "selected_campaigns" else ("선택 그룹 기준" if search_scope == "selected_adgroups" else "계정 전체 기준")
        return jsonify({
            "ok": True,
            "message": f"[{scope_label}]\n" + _build_powerlink_keyword_search_message(search_text, exact_match, scan, row_preview_limit=80, exclude_text=exclude_text),
            "search_text": search_text,
            "search_groups": scan.get("search_groups") or _parse_keyword_search_groups(search_text),
            "exact_match": bool(exact_match),
            "exclude_text": exclude_text,
            "matched_count": int(scan.get("matched_count") or 0),
            "scanned_keyword_count": int(scan.get("scanned_keyword_count") or 0),
            "total_powerlink_campaign_count": len(scan.get("powerlink_campaigns") or []),
            "total_powerlink_adgroup_count": len(scan.get("adgroup_contexts") or []),
            "rows": (scan.get("matched_rows") or []),
            "warnings": (scan.get("warnings") or [])[:10],
            "preview_token": preview_token,
            "search_scope": search_scope,
            "selected_campaign_count": len(campaign_ids) if search_scope == "selected_campaigns" else 0,
            "selected_adgroup_count": len(adgroup_ids) if search_scope == "selected_adgroups" else 0,
        })
    except Exception as e:
        return jsonify({"error": "검색어 기준 키워드 조회 실패", "details": str(e)}), 400
@app.route("/export_powerlink_keywords_excel", methods=["POST"])
def export_powerlink_keywords_excel():
    d = request.json or {}
    api_key = str(d.get("api_key") or "").strip()
    secret_key = str(d.get("secret_key") or "").strip()
    cid = str(d.get("customer_id") or "").strip()
    search_text = str(d.get("search_text") or d.get("keyword_query") or "").strip()
    exact_match = _boolish(d.get("exact_match"), False)
    exclude_text = str(d.get("exclude_text") or d.get("exclude_keyword_query") or "").strip()
    search_scope = str(d.get("search_scope") or "account").strip().lower()
    campaign_ids = [str(x or "").strip() for x in (d.get("campaign_ids") or []) if str(x or "").strip()]
    adgroup_ids = [str(x or "").strip() for x in (d.get("adgroup_ids") or []) if str(x or "").strip()]
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API 정보 및 광고주를 선택해주세요."}), 400
    if not search_text:
        return jsonify({"error": "검색어를 입력해주세요."}), 400
    if search_scope not in {"account", "selected_campaigns", "selected_adgroups"}:
        search_scope = "account"
    if search_scope == "selected_campaigns" and not campaign_ids:
        return jsonify({"error": "선택 캠페인 조회를 사용하려면 좌측에서 캠페인을 체크해주세요."}), 400
    if search_scope == "selected_adgroups" and not adgroup_ids:
        return jsonify({"error": "선택 그룹내 조회를 사용하려면 좌측에서 광고그룹을 체크해주세요."}), 400
    try:
        scan = _scan_powerlink_keywords_by_search(
            api_key,
            secret_key,
            cid,
            search_text,
            exact_match=exact_match,
            campaign_ids=(campaign_ids if search_scope == "selected_campaigns" else None),
            adgroup_ids=(adgroup_ids if search_scope == "selected_adgroups" else None),
            exclude_text=exclude_text,
        )
        rows = scan.get("matched_rows") or []
        if not rows:
            return jsonify({"error": "내보낼 일치 키워드가 없습니다."}), 400
        wb = _build_powerlink_keyword_export_workbook(rows, search_text, exact_match, exclude_text=exclude_text)
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        stamp = time.strftime("%Y%m%d_%H%M%S")
        mode_label = "exact" if exact_match else "partial"
        filename = f"powerlink_keyword_search_{mode_label}_{stamp}.xlsx"
        return send_file(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except Exception as e:
        return jsonify({"error": "키워드 엑셀 다운로드 실패", "details": str(e)}), 400
@app.route("/get_action_logs", methods=["GET"])
def get_action_logs():
    try:
        limit_raw = str(request.args.get("limit") or "120").strip()
        limit = int(limit_raw)
        if limit <= 0:
            limit = 120
        limit = min(limit, 500)
        return jsonify(_read_action_logs(limit))
    except Exception as e:
        return jsonify({"error": "작업 로그 조회 실패", "details": str(e)}), 500
@app.route("/clear_action_logs", methods=["POST"])
def clear_action_logs():
    try:
        with _ACTION_LOG_LOCK:
            os.makedirs(LOG_DIR, exist_ok=True)
            with open(ACTION_LOG_PATH, 'w', encoding='utf-8') as fp:
                fp.write("")
        return jsonify({"ok": True, "message": "작업 로그를 비웠습니다."})
    except Exception as e:
        return jsonify({"error": "작업 로그 비우기 실패", "details": str(e)}), 500
@app.route("/create_campaign", methods=["POST"])
def create_campaign():
    d = request.json or {}
    payload = {
        "customerId": int(d.get("customer_id")),
        "name": str(d.get("name") or "").strip(),
        "campaignTp": _normalize_campaign_tp(d.get("campaign_tp")),
        "useDailyBudget": bool(d.get("use_daily_budget", True)),
        "dailyBudget": int(d.get("daily_budget") or 0),
    }
    if not payload["name"]:
        return jsonify({"error": "캠페인명은 필수입니다."}), 400
    res = _do_req("POST", d.get("api_key"), d.get("secret_key"), d.get("customer_id"), "/ncc/campaigns", json_body=payload)
    if res.status_code in [200, 201]:
        _cache_invalidate(d.get("api_key"), d.get("secret_key"), d.get("customer_id"))
        return jsonify({"ok": True, "item": res.json(), "message": "캠페인 생성 완료"})
    return jsonify({"error": "캠페인 생성 실패", "details": res.text}), 400
@app.route("/create_adgroup_simple", methods=["POST"])
def create_adgroup_simple():
    d = request.json or {}
    campaign_id = str(d.get("campaign_id") or "").strip()
    name = str(d.get("name") or "").strip()
    requested_campaign_tp = _normalize_campaign_tp(d.get("campaign_tp"))
    requested_adgroup_tp = _normalize_adgroup_tp(d.get("adgroup_type"), requested_campaign_tp)
    biz_channel_id = str(d.get("biz_channel_id") or "").strip()
    media_type = str(d.get("media_type") or "ALL").strip().upper()
    media_detail = _normalize_media_detail(d.get("media_detail"))
    use_keyword_plus = d.get("use_keyword_plus")
    use_close_variant = d.get("use_close_variant")
    keyword_plus_weight = d.get("keyword_plus_weight")
    if not campaign_id or not name:
        return jsonify({"error": "캠페인과 광고그룹명은 필수입니다."}), 400
    campaign_detail = _fetch_campaign_detail(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), campaign_id)
    actual_campaign_tp = _normalize_campaign_tp((campaign_detail or {}).get("campaignTp") or requested_campaign_tp)
    # 중요: 최신 버전에서 쇼핑 캠페인 유형을 실제 campaignTp(CATALOG/SHOPPING_BRAND 등)로 강제 승격하면서
    # 기존 3/4번 파일 대비 생성 실패가 늘어날 수 있다. 생성 자체는 예전처럼 "요청값 우선"으로 시도하고,
    # 필요할 때만 실제 campaignTp 기반 fallback 을 추가로 시도한다.
    is_requested_shopping = _is_shopping_campaign_type(requested_campaign_tp) or _is_shopping_campaign_type(requested_adgroup_tp)
    is_actual_shopping = _is_shopping_campaign_type(actual_campaign_tp)
    is_web_campaign = not (is_requested_shopping or is_actual_shopping) and actual_campaign_tp == "WEB_SITE"
    legacy_adgroup_tp = _normalize_adgroup_tp(d.get("adgroup_type"), requested_campaign_tp or actual_campaign_tp)
    inferred_adgroup_tp = _normalize_adgroup_tp(d.get("adgroup_type"), actual_campaign_tp or requested_campaign_tp)
    def _base_payload(adgroup_tp: str, attach_biz_channel: bool) -> Dict[str, Any]:
        payload = {
            "customerId": int(d.get("customer_id")),
            "nccCampaignId": campaign_id,
            "name": name,
            "adgroupType": adgroup_tp,
            "useDailyBudget": bool(d.get("use_daily_budget", True)),
            "dailyBudget": int(d.get("daily_budget") or 0),
            "bidAmt": int(d.get("bid_amt") or 70),
        }
        if attach_biz_channel and biz_channel_id:
            payload["pcChannelId"] = biz_channel_id
            payload["mobileChannelId"] = biz_channel_id
        return payload
    attempts: List[Dict[str, Any]] = []
    candidate_payloads: List[Tuple[str, Dict[str, Any]]] = []
    if is_web_campaign:
        if not biz_channel_id:
            biz_channel_id = _fetch_first_biz_channel_id(d.get("api_key"), d.get("secret_key"), d.get("customer_id"))
            if not biz_channel_id:
                return jsonify({"error": "파워링크 광고그룹은 비즈채널이 필요합니다. 비즈채널을 먼저 선택하거나 생성해주세요."}), 400
        candidate_payloads.append(("web-primary", _base_payload(legacy_adgroup_tp, True)))
    else:
        seen_signatures = set()
        def _append_candidate(label: str, adgroup_tp: str):
            tp = str(adgroup_tp or "").strip().upper()
            if not tp:
                return
            payload = _base_payload(tp, False)
            sig = json.dumps(payload, ensure_ascii=False, sort_keys=True)
            if sig in seen_signatures:
                return
            seen_signatures.add(sig)
            candidate_payloads.append((label, payload))
        # 1차: 예전 3/4번 파일처럼 요청값 우선(가장 호환성이 높음)
        _append_candidate("shopping-legacy", legacy_adgroup_tp)
        # 2차: 실제 캠페인 유형 기반 fallback. 다만 SHOPPING_BRAND 는 상품그룹 ID가 없으면 실패 확률이 높아 자동 강제하지 않는다.
        actual_upper = str(actual_campaign_tp or "").strip().upper()
        if actual_upper == "CATALOG":
            _append_candidate("shopping-catalog", "CATALOG")
        elif actual_upper == "SHOPPING_BRAND":
            if str(d.get("ncc_product_group_id") or d.get("nccProductGroupId") or "").strip():
                payload = _base_payload("SHOPPING_BRAND", False)
                payload["nccProductGroupId"] = str(d.get("ncc_product_group_id") or d.get("nccProductGroupId") or "").strip()
                if int(payload.get("bidAmt") or 0) < 300:
                    payload["bidAmt"] = 300
                sig = json.dumps(payload, ensure_ascii=False, sort_keys=True)
                if sig not in seen_signatures:
                    seen_signatures.add(sig)
                    candidate_payloads.append(("shopping-brand", payload))
        elif inferred_adgroup_tp and inferred_adgroup_tp != legacy_adgroup_tp:
            _append_candidate("shopping-inferred", inferred_adgroup_tp)
    final_res = None
    final_payload = None
    final_label = ""
    for label, payload in candidate_payloads:
        attempts.append({"label": label, "payload": payload})
        res = _do_req("POST", d.get("api_key"), d.get("secret_key"), d.get("customer_id"), "/ncc/adgroups", json_body=payload)
        if res.status_code in [200, 201]:
            final_res = res
            final_payload = payload
            final_label = label
            break
        final_res = res
        final_payload = payload
        final_label = label
    res = final_res
    payload = final_payload or _base_payload(legacy_adgroup_tp, is_web_campaign and bool(biz_channel_id))
    if res and res.status_code in [200, 201]:
        item = res.json() or {}
        new_adgroup_id = str(item.get("nccAdgroupId") or item.get("id") or "").strip()
        warnings: List[str] = []
        if new_adgroup_id:
            ok_media_pm, detail_media_pm = _update_pc_mobile_target(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), new_adgroup_id, media_type)
            if not ok_media_pm:
                warnings.append(f"PC/모바일 매체 적용 실패: {detail_media_pm}")
            ok_media_network, detail_media_network = _update_media_target(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), new_adgroup_id, media_detail)
            if not ok_media_network:
                warnings.append(f"세부 매체 적용 실패: {detail_media_network}")
            if is_web_campaign:
                ukp = None if use_keyword_plus is None else bool(use_keyword_plus)
                ucv = None if use_close_variant is None else bool(use_close_variant)
                kwp = None
                if str(keyword_plus_weight or "").strip() != "":
                    try:
                        kwp = int(keyword_plus_weight)
                    except Exception:
                        kwp = None
                ok_opts, detail_opts = _update_adgroup_search_options(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), new_adgroup_id, use_keyword_plus=ukp, keyword_plus_weight=kwp, use_close_variant=ucv)
                if not ok_opts:
                    warnings.append(f"파워링크 검색옵션 적용 실패: {detail_opts}")
        message = "광고그룹 생성 완료"
        if warnings:
            message += " (일부 후처리 실패)"
        if final_label and final_label != "web-primary":
            message += f" [{final_label}]"
        return jsonify({"ok": True, "item": item, "message": message, "warnings": warnings})
    detail = res.text if res is not None else "알 수 없는 오류"
    try:
        j = res.json() if res is not None else {}
        detail = j.get("title") or j.get("message") or detail
        if j.get("detail"):
            detail = f"{detail} | {j.get('detail')}"
    except Exception:
        pass
    hint = None
    if str(actual_campaign_tp or "").strip().upper() == "SHOPPING_BRAND" and not str(d.get("ncc_product_group_id") or d.get("nccProductGroupId") or "").strip():
        hint = "현재 캠페인이 SHOPPING_BRAND 로 보입니다. 이 유형은 상품그룹 ID(nccProductGroupId) 없이는 생성이 실패할 수 있습니다."
    return jsonify({
        "error": "광고그룹 생성 실패",
        "details": detail,
        "payload": payload,
        "hint": hint,
        "debug": {
            "requested_campaign_tp": requested_campaign_tp,
            "actual_campaign_tp": actual_campaign_tp,
            "requested_adgroup_tp": requested_adgroup_tp,
            "legacy_adgroup_tp": legacy_adgroup_tp,
            "inferred_adgroup_tp": inferred_adgroup_tp,
            "biz_channel_id": biz_channel_id,
            "campaign_id": campaign_id,
            "is_web_campaign": is_web_campaign,
            "is_requested_shopping": is_requested_shopping,
            "is_actual_shopping": is_actual_shopping,
            "campaign_detail": campaign_detail,
            "attempted_payloads": attempts,
            "last_attempt_label": final_label,
        },
    }), 400
@app.route("/create_keywords_simple", methods=["POST"])
def create_keywords_simple():
    d = request.json or {}
    keyword_batches = d.get("keyword_batches") if isinstance(d.get("keyword_batches"), list) else None
    rows: List[Dict[str, Any]] = []
    batch_summary: List[Dict[str, Any]] = []
    errors: List[str] = []
    def _append_keywords(adgroup_id: str, keywords_text: str, slot_index: Optional[int] = None):
        adgroup_id_local = str(adgroup_id or "").strip()
        keywords_local = [x.strip() for x in str(keywords_text or "").replace(",", "\n").splitlines() if x.strip()]
        if not keywords_local:
            return
        if not adgroup_id_local:
            errors.append(f"광고그룹을 선택하지 않은 슬롯이 있습니다{'' if slot_index is None else f' (슬롯 {slot_index})'}.")
            return
        for kw in keywords_local:
            rows.append({
                "nccAdgroupId": adgroup_id_local,
                "keyword": kw,
                "useGroupBidAmt": bool(d.get("use_group_bid_amt", True)),
                "bidAmt": int(d.get("bid_amt") or 70),
                "userLock": bool(d.get("user_lock", False)),
            })
        batch_summary.append({
            "slot_index": int(slot_index or len(batch_summary) + 1),
            "adgroup_id": adgroup_id_local,
            "keyword_count": len(keywords_local),
        })
    if keyword_batches:
        for idx, batch in enumerate(keyword_batches, start=1):
            if not isinstance(batch, dict):
                continue
            _append_keywords(
                str(batch.get("adgroup_id") or "").strip(),
                batch.get("keywords") or "",
                int(batch.get("slot_index") or idx),
            )
    else:
        _append_keywords(str(d.get("adgroup_id") or "").strip(), d.get("keywords") or "", 1)
    if errors:
        return jsonify({"error": " / ".join(errors)}), 400
    if not rows:
        return jsonify({"error": "광고그룹과 키워드는 필수입니다."}), 400
    success, fail, results = _bulk_create_keywords(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), rows)
    return jsonify({
        "ok": True,
        "total": len(rows),
        "success": success,
        "fail": fail,
        "results": results,
        "batch_summary": batch_summary,
    })
def _parse_target_ids(payload: Dict[str, Any], single_key: str, multi_key: str) -> List[str]:
    ids: List[str] = []
    raw_multi = payload.get(multi_key)
    if isinstance(raw_multi, list):
        ids.extend([str(x).strip() for x in raw_multi if str(x).strip()])
    elif isinstance(raw_multi, str) and raw_multi.strip():
        ids.extend([x.strip() for x in raw_multi.replace(",", "\n").splitlines() if x.strip()])
    raw_single = str(payload.get(single_key) or "").strip()
    if raw_single:
        ids.append(raw_single)
    uniq: List[str] = []
    seen = set()
    for x in ids:
        if x not in seen:
            seen.add(x)
            uniq.append(x)
    return uniq
def _count_keyword_insertions(text_value: str) -> int:
    return len(re.findall(r"\{키워드(?::[^}]+)?\}", str(text_value or "")))
def _apply_keyword_insertion_template(text_value: str, replace_keyword: str) -> str:
    text_value = str(text_value or "").strip()
    if not text_value:
        return text_value
    if "{키워드}" in text_value:
        if not replace_keyword:
            raise ValueError("키워드삽입 사용 시 대체키워드를 입력해야 합니다.")
        text_value = text_value.replace("{키워드}", f"{{키워드:{replace_keyword}}}")
    return text_value
def _visible_keyword_length(text_value: str) -> int:
    text_value = str(text_value or "").strip()
    text_value = re.sub(r"\{키워드:[^}]+\}", "{키워드}", text_value)
    return len(text_value)
def _text_ad_length_errors(headline: str, description: str) -> List[str]:
    errors: List[str] = []
    headline = str(headline or "").strip()
    description = str(description or "").strip()
    headline_visible_len = _visible_keyword_length(headline)
    description_visible_len = _visible_keyword_length(description)
    if not (1 <= headline_visible_len <= 15):
        errors.append(f"제목은 대체키워드 제외 기준 1~15자여야 합니다. (현재 {headline_visible_len}자)")
    if not (20 <= description_visible_len <= 45):
        errors.append(f"설명은 대체키워드 제외 기준 20~45자여야 합니다. (현재 {description_visible_len}자)")
    return errors
def _parse_extension_position(value: Any) -> int | None:
    try:
        pos = int(str(value or "").strip())
    except Exception:
        return None
    return pos if pos in {1, 2} else None
def _normalize_headline_pin_input(value: Any) -> int | None:
    s = str(value or '').strip()
    if not s:
        return None
    lowered = s.lower().replace(' ', '')
    if lowered in {'all', '전체', '모든위치', '모든위치노출', '전체노출', 'allpositions'}:
        return None
    if lowered in {'1', '1번', '1st', 'pin1', 'position1', '위치1', '위치1만', '위치1만노출', '위치1에만노출', '위치1만노출가능'}:
        return 1
    if lowered in {'2', '2번', '2nd', 'pin2', 'position2', '위치2', '위치2만', '위치2만노출', '위치2에만노출', '위치2만노출가능'}:
        return 2
    m = re.search(r'([12])', lowered)
    if m:
        return int(m.group(1))
    raise ValueError('pin 값은 비우거나 1 / 2 로 입력해주세요.')
def _row_pick_value(row: Dict[str, Any], candidate_keys: List[str]) -> Any:
    lookup: Dict[str, Any] = {}
    for k, v in (row or {}).items():
        nk = re.sub(r"\s+", "", str(k or "")).lower()
        lookup[nk] = v
    for key in candidate_keys:
        nk = re.sub(r"\s+", "", str(key or "")).lower()
        if nk in lookup:
            return lookup[nk]
    return None
def _row_has_any_key(row: Dict[str, Any], candidate_keys: List[str]) -> bool:
    lookup = {re.sub(r"\s+", "", str(k or "")).lower() for k in (row or {}).keys()}
    for key in candidate_keys:
        nk = re.sub(r"\s+", "", str(key or "")).lower()
        if nk in lookup:
            return True
    return False
def _boolish(value: Any, default: bool = False) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {'1', 'true', 'y', 'yes', 'on'}
def _create_text_ad_for_adgroup(d: Dict[str, Any], adgroup_id: str, headline: str, description: str, pc_url: str, mobile_url: str) -> Dict[str, Any]:
    payload = {
        "customerId": int(d.get("customer_id")),
        "nccAdgroupId": adgroup_id,
        "type": "TEXT_45",
        "userLock": bool(d.get("user_lock", False)),
        "ad": {
            "headline": headline,
            "description": description,
            "pc": {"final": pc_url, "display": _display_url_from_final(pc_url)},
            "mobile": {"final": mobile_url, "display": _display_url_from_final(mobile_url)},
        },
    }
    res = _do_req("POST", d.get("api_key"), d.get("secret_key"), d.get("customer_id"), "/ncc/ads", params={"nccAdgroupId": adgroup_id}, json_body=payload)
    ok = res is not None and res.status_code in [200, 201]
    return {
        "ok": ok,
        "adgroup_id": adgroup_id,
        "detail": (res.text if res is not None else "응답 없음"),
        "item": (res.json() if ok else None),
    }
def _create_shopping_ad_for_adgroup(d: Dict[str, Any], adgroup_id: str, reference_key: str, ad_type: str) -> Dict[str, Any]:
    payload = {
        "customerId": int(d.get("customer_id")),
        "nccAdgroupId": adgroup_id,
        "type": ad_type,
        "referenceKey": reference_key,
        "useGroupBidAmt": bool(d.get("use_group_bid_amt", True)),
        "bidAmt": int(d.get("bid_amt") or 0),
        "userLock": bool(d.get("user_lock", False)),
        "ad": {},
    }
    res = _do_req("POST", d.get("api_key"), d.get("secret_key"), d.get("customer_id"), "/ncc/ads", params={"nccAdgroupId": adgroup_id, "isList": "true"}, json_body=[payload])
    ok = res is not None and res.status_code in [200, 201]
    return {
        "ok": ok,
        "adgroup_id": adgroup_id,
        "detail": (res.text if res is not None else "응답 없음"),
        "item": (res.json() if ok else None),
    }
def _create_extension_for_owner(d: Dict[str, Any], owner_id: str, ext_type: str, data: Dict[str, Any], position: int | None = None) -> Dict[str, Any]:
    payload = _build_extension_payload(owner_id, ext_type, data, int(d.get("customer_id")), position=position)
    if ext_type != "SHOPPING_EXTRA" and not payload.get("adExtension"):
        return {"ok": False, "owner_id": owner_id, "detail": "확장소재 내용이 비어 있습니다."}
    res = _do_req("POST", d.get("api_key"), d.get("secret_key"), d.get("customer_id"), "/ncc/ad-extensions", params={"ownerId": owner_id}, json_body=payload)
    ok = res is not None and res.status_code in [200, 201]
    return {
        "ok": ok,
        "owner_id": owner_id,
        "detail": (res.text if res is not None else "응답 없음"),
        "item": (res.json() if ok else None),
    }
def _bulk_upload_one_text_ad(api_key: str, secret_key: str, cid: str, row_no: int, adgroup_id: str, headline: str, description: str, pc_url: str, mobile_url: str) -> Dict[str, Any]:
    result = _create_text_ad_for_adgroup({
        "api_key": api_key,
        "secret_key": secret_key,
        "customer_id": cid,
    }, adgroup_id, headline, description, pc_url, mobile_url)
    if result.get("ok"):
        return _result_item(row_no, True, headline or adgroup_id, "생성 완료")
    return _result_item(row_no, False, headline or adgroup_id, result.get("detail") or "소재 생성 실패")
@app.route("/bulk_upload_text_ads", methods=["POST"])
def bulk_upload_text_ads():
    api_key = str(request.form.get("api_key") or "").strip()
    secret_key = str(request.form.get("secret_key") or "").strip()
    cid = str(request.form.get("customer_id") or "").strip()
    upload = request.files.get("file")
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API Key / Secret Key / Customer ID가 필요합니다."}), 400
    if upload is None or not getattr(upload, "filename", ""):
        return jsonify({"error": "업로드할 파일을 선택해주세요."}), 400
    try:
        rows = _read_uploaded_table(upload)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"업로드 파일 처리 실패: {e}"}), 400
    if not rows:
        return jsonify({"error": "파일에서 등록할 데이터가 없습니다."}), 400
    prepared: List[Tuple[int, str, str, str, str, str]] = []
    precheck_results: List[Dict[str, Any]] = []
    for idx, row in enumerate(rows, start=1):
        try:
            payload = _prepare_payload_row(row, "ad", cid)
        except Exception as e:
            precheck_results.append(_result_item(idx, False, f"{idx}행", f"행 파싱 실패: {e}"))
            continue
        adgroup_id = str(payload.get("nccAdgroupId") or "").strip()
        ad_obj = payload.get("ad") if isinstance(payload.get("ad"), dict) else {}
        raw_headline = str((ad_obj or {}).get("headline") or payload.get("headline") or "").strip()
        raw_description = str((ad_obj or {}).get("description") or payload.get("description") or "").strip()
        pc_url = _ensure_final_url((ad_obj.get("pc") or {}).get("final") if isinstance(ad_obj, dict) else payload.get("pcFinalUrl"))
        mobile_url = _ensure_final_url((ad_obj.get("mobile") or {}).get("final") if isinstance(ad_obj, dict) else payload.get("mobileFinalUrl")) or pc_url
        replace_keyword = str(row.get("replace_keyword") or row.get("대체키워드") or row.get("replaceKeyword") or "").strip()
        if not adgroup_id or not raw_headline or not raw_description or not pc_url:
            precheck_results.append(_result_item(idx, False, raw_headline or f"{idx}행", "그룹ID, 소재제목, 설명, PC 연결 URL은 필수입니다."))
            continue
        try:
            headline = _apply_keyword_insertion_template(raw_headline, replace_keyword)
            description = _apply_keyword_insertion_template(raw_description, replace_keyword)
        except ValueError as e:
            precheck_results.append(_result_item(idx, False, raw_headline or f"{idx}행", str(e)))
            continue
        headline_insert_cnt = _count_keyword_insertions(headline)
        desc_insert_cnt = _count_keyword_insertions(description)
        if headline_insert_cnt > 1:
            precheck_results.append(_result_item(idx, False, raw_headline or f"{idx}행", "키워드삽입은 제목에 1회까지만 사용할 수 있습니다."))
            continue
        if desc_insert_cnt > 2:
            precheck_results.append(_result_item(idx, False, raw_headline or f"{idx}행", "키워드삽입은 설명에 2회까지만 사용할 수 있습니다."))
            continue
        length_errors = _text_ad_length_errors(headline, description)
        if length_errors:
            precheck_results.append(_result_item(idx, False, raw_headline or f"{idx}행", " / ".join(length_errors)))
            continue
        prepared.append((idx, adgroup_id, headline, description, pc_url, mobile_url))
    exec_results: List[Dict[str, Any]] = []
    if prepared:
        with ThreadPoolExecutor(max_workers=5) as executor:
            futures = [
                executor.submit(_bulk_upload_one_text_ad, api_key, secret_key, cid, idx, adgroup_id, headline, description, pc_url, mobile_url)
                for idx, adgroup_id, headline, description, pc_url, mobile_url in prepared
            ]
            for future in as_completed(futures):
                exec_results.append(future.result())
    results = sorted(precheck_results + exec_results, key=lambda x: x.get("row_no", 0))
    success = sum(1 for item in results if item.get("ok"))
    fail = len(results) - success
    return jsonify({
        "ok": True,
        "total": len(results),
        "success": success,
        "fail": fail,
        "results": results,
        "message": f"소재 대량등록 완료 · 성공 {success}건 / 실패 {fail}건",
    })
@app.route("/create_text_ad_simple", methods=["POST"])
def create_text_ad_simple():
    d = request.json or {}
    adgroup_ids = _parse_target_ids(d, "adgroup_id", "adgroup_ids")
    raw_headline = str(d.get("headline") or d.get("headline_template") or "").strip()
    raw_description = str(d.get("description") or d.get("description_template") or "").strip()
    replace_keyword = str(d.get("replace_keyword") or "").strip()
    pc_url = str(d.get("pc_url") or "").strip()
    mobile_url = str(d.get("mobile_url") or "").strip() or pc_url
    if not adgroup_ids or not raw_headline or not raw_description or not pc_url or not mobile_url:
        return jsonify({"error": "광고그룹, 제목, 설명, PC URL, 모바일 URL은 필수입니다."}), 400
    try:
        headline = _apply_keyword_insertion_template(raw_headline, replace_keyword)
        description = _apply_keyword_insertion_template(raw_description, replace_keyword)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    headline_insert_cnt = _count_keyword_insertions(headline)
    desc_insert_cnt = _count_keyword_insertions(description)
    if headline_insert_cnt > 1:
        return jsonify({"error": "키워드삽입은 제목에 1회까지만 사용할 수 있습니다."}), 400
    if desc_insert_cnt > 2:
        return jsonify({"error": "키워드삽입은 설명에 2회까지만 사용할 수 있습니다."}), 400
    length_errors = _text_ad_length_errors(headline, description)
    if length_errors:
        return jsonify({"error": " / ".join(length_errors), "headline": headline, "description": description}), 400
    results = [_create_text_ad_for_adgroup(d, adg_id, headline, description, pc_url, mobile_url) for adg_id in adgroup_ids]
    success = sum(1 for r in results if r.get("ok"))
    fail = len(results) - success
    if len(adgroup_ids) == 1:
        r = results[0]
        if r.get("ok"):
            return jsonify({"ok": True, "item": r.get("item"), "message": "기본소재 생성 완료"})
        payload = {
            "customerId": int(d.get("customer_id")),
            "nccAdgroupId": adgroup_ids[0],
            "type": "TEXT_45",
            "userLock": bool(d.get("user_lock", False)),
            "ad": {
                "headline": headline,
                "description": description,
                "pc": {"final": pc_url, "display": _display_url_from_final(pc_url)},
                "mobile": {"final": mobile_url, "display": _display_url_from_final(mobile_url)},
            },
        }
        return jsonify({"error": "소재 생성 실패", "details": r.get("detail"), "payload": payload}), 400
    return jsonify({
        "ok": True,
        "total": len(results),
        "success": success,
        "fail": fail,
        "results": results,
        "message": f"기본소재 등록 완료 · 성공 {success}건 / 실패 {fail}건",
    })
@app.route("/create_ad_advanced", methods=["POST"])
def create_ad_advanced():
    d = request.json or {}
    adgroup_id = str(d.get("adgroup_id") or "").strip()
    raw_json = str(d.get("raw_json") or "").strip()
    if not adgroup_id or not raw_json:
        return jsonify({"error": "광고그룹과 JSON 입력이 필요합니다."}), 400
    try:
        payload = json.loads(raw_json)
    except Exception as e:
        return jsonify({"error": f"JSON 파싱 실패: {e}"}), 400
    payload.setdefault("nccAdgroupId", adgroup_id)
    payload.setdefault("customerId", int(d.get("customer_id")))
    ad_type = str(payload.get("type") or "")
    if ad_type in SHOPPING_AD_TYPES:
        res = _do_req("POST", d.get("api_key"), d.get("secret_key"), d.get("customer_id"), "/ncc/ads", params={"nccAdgroupId": adgroup_id, "isList": "true"}, json_body=[payload])
    else:
        res = _do_req("POST", d.get("api_key"), d.get("secret_key"), d.get("customer_id"), "/ncc/ads", params={"nccAdgroupId": adgroup_id}, json_body=payload)
    if res.status_code in [200, 201]:
        return jsonify({"ok": True, "item": res.json(), "message": "고급 소재 생성 완료"})
    return jsonify({"error": "고급 소재 생성 실패", "details": res.text, "payload": payload}), 400
@app.route("/create_extension_simple", methods=["POST"])
def create_extension_simple():
    d = request.json or {}
    owner_ids = _parse_target_ids(d, "owner_id", "owner_ids")
    adgroup_ids = _parse_target_ids(d, "adgroup_id", "adgroup_ids")
    campaign_ids = _parse_target_ids(d, "campaign_id", "campaign_ids")
    ext_type = _normalize_extension_type(d.get("type"))
    if not ext_type:
        return jsonify({"error": "확장소재 유형은 필수입니다."}), 400
    if ext_type not in {"HEADLINE", "DESCRIPTION_EXTRA", "DESCRIPTION", "SUB_LINKS", "PROMOTION", "SHOPPING_EXTRA"}:
        return jsonify({"error": "현재 간편 등록은 추가제목 / 홍보문구 / 추가설명 / 서브링크 / 쇼핑 추가홍보문구 / 쇼핑상품부가정보만 지원합니다."}), 400
    data: Dict[str, Any] = {}
    position = _parse_extension_position(d.get("position"))
    warnings: List[str] = []
    if ext_type == "HEADLINE":
        headline = str(d.get("headline") or "").strip()
        if not headline or len(headline) > 15:
            return jsonify({"error": "추가제목은 1~15자로 입력해야 합니다."}), 400
        data["headline"] = headline
    elif ext_type == "DESCRIPTION_EXTRA":
        description = str(d.get("description") or "").strip()
        if not description or len(description) > 14:
            return jsonify({"error": "홍보문구는 1~14자로 입력해야 합니다."}), 400
        data["description"] = description
    elif ext_type == "PROMOTION":
        basic_text = str(d.get("basic_text") or "").strip()
        additional_text = str(d.get("additional_text") or "").strip()
        if not basic_text:
            return jsonify({"error": "문구 1은 필수입니다."}), 400
        if len(basic_text) > 10:
            return jsonify({"error": "문구 1은 10자 이내로 입력해야 합니다."}), 400
        if additional_text and len(additional_text) > 30:
            return jsonify({"error": "문구 2는 30자 이내로 입력해야 합니다."}), 400
        data["basicText"] = basic_text
        if additional_text:
            data["additionalText"] = additional_text
    elif ext_type == "DESCRIPTION":
        description = str(d.get("description") or "").strip()
        if not description:
            return jsonify({"error": "추가설명은 필수입니다."}), 400
        data["description"] = description
    elif ext_type == "SUB_LINKS":
        raw_links = d.get("links") or []
        links = []
        for item in raw_links:
            if not isinstance(item, dict):
                continue
            name = str(item.get("name") or "").strip()
            final = str(item.get("final") or "").strip()
            if not name and not final:
                continue
            if len(name) > 6:
                return jsonify({"error": f"서브링크명 '{name}' 은(는) 6자 이내여야 합니다."}), 400
            if not re.match(r"^https?://", final, re.I):
                return jsonify({"error": f"서브링크 URL은 http:// 또는 https:// 로 시작해야 합니다. ({final})"}), 400
            links.append({"name": name, "final": final})
        if len(links) < 3 or len(links) > 4:
            return jsonify({"error": "서브링크는 최소 3개, 최대 4개까지 등록할 수 있습니다."}), 400
        data["links"] = links
    elif ext_type == "SHOPPING_EXTRA":
        resolved_owner_ids, warnings = _resolve_shopping_extra_owner_ids(
            d.get("api_key"), d.get("secret_key"), d.get("customer_id"),
            campaign_ids=campaign_ids, adgroup_ids=(adgroup_ids or [])
        )
        if resolved_owner_ids:
            owner_ids = resolved_owner_ids
        elif not owner_ids:
            return jsonify({
                "error": "쇼핑상품부가정보를 추가할 쇼핑 상품소재를 찾지 못했습니다.",
                "warnings": warnings,
            }), 400
    if not owner_ids:
        return jsonify({"error": "대상 광고그룹 또는 캠페인을 선택해주세요."}), 400
    results = []
    success = fail = 0
    for owner_id in owner_ids:
        payload = _build_extension_payload(owner_id, ext_type, data, int(d.get("customer_id")), position=position)
        if ext_type != "SHOPPING_EXTRA" and not payload.get("adExtension"):
            results.append({"ok": False, "owner_id": owner_id, "detail": "확장소재 내용이 비어 있습니다."})
            fail += 1
            continue
        existing = _find_existing_extension(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), owner_id, ext_type, data)
        if existing:
            detail_msg = "이미 동일한 확장소재가 등록되어 있습니다."
            if ext_type == "HEADLINE" and position in {1, 2}:
                upd = _apply_headline_position_best_effort(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), existing, position)
                if upd.get("ok"):
                    existing = upd.get("item") or existing
                    verify = _verify_headline_pin(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), owner_id, data.get("headline"), position)
                    if verify.get("ok"):
                        existing = verify.get("item") or existing
                        detail_msg = f"이미 등록된 추가제목의 노출 위치를 {position}번으로 반영했습니다. · {verify.get('detail')}"
                    else:
                        warnings.append(verify.get('detail'))
                        detail_msg = f"이미 등록된 추가제목의 노출 위치를 {position}번으로 반영 요청했습니다."
                else:
                    warnings.append(f"추가제목 위치 {position}번 반영 미확인: {upd.get('detail')}")
            success += 1
            results.append({"ok": True, "owner_id": owner_id, "detail": detail_msg, "item": existing, "exists": True})
            continue
        if ext_type == "SHOPPING_EXTRA":
            ok_created, res, used_payload, attempts = _create_shopping_extra_with_fallbacks(
                d.get("api_key"), d.get("secret_key"), d.get("customer_id"), owner_id
            )
            if ok_created and res is not None:
                success += 1
                item = None
                try:
                    item = res.json()
                except Exception:
                    item = None
                results.append({"ok": True, "owner_id": owner_id, "detail": "등록 완료", "item": item, "payload": used_payload})
                continue
            payload = used_payload or payload
        else:
            res = _do_req("POST", d.get("api_key"), d.get("secret_key"), d.get("customer_id"), "/ncc/ad-extensions", params={"ownerId": owner_id}, json_body=payload)
            attempts = []
        if res.status_code in [200, 201]:
            success += 1
            item = None
            try:
                item = res.json()
            except Exception:
                item = None
            item_one = _extract_extension_item(item)
            detail_msg = "등록 완료"
            if ext_type == "HEADLINE" and position in {1, 2} and item_one:
                upd = _apply_headline_position_best_effort(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), item_one, position)
                if upd.get("ok"):
                    verify = _verify_headline_pin(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), owner_id, data.get("headline"), position)
                    if verify.get("ok"):
                        item = verify.get("item") or upd.get("item") or item
                        detail_msg = f"등록 완료 (위치 {position} 반영 · {verify.get('detail')})"
                    else:
                        item = upd.get("item") or item
                        warnings.append(verify.get('detail'))
                        detail_msg = f"등록 완료 (위치 {position} 반영 요청)"
                else:
                    warnings.append(f"추가제목 위치 {position}번 반영 미확인: {upd.get('detail')}")
            results.append({"ok": True, "owner_id": owner_id, "detail": detail_msg, "item": item})
        else:
            if res.status_code == 400 and "A record with the same name already exists." in (res.text or ""):
                existing = _find_existing_extension(d.get("api_key"), d.get("secret_key"), d.get("customer_id"), owner_id, ext_type, data)
                if existing:
                    success += 1
                    results.append({"ok": True, "owner_id": owner_id, "detail": "이미 동일한 확장소재가 등록되어 있습니다.", "item": existing, "exists": True})
                    continue
            fail += 1
            extra_detail = res.text
            if ext_type == "SHOPPING_EXTRA" and attempts:
                extra_detail = f"{res.text}\n\n시도한 payload 후보 {len(attempts)}건:\n" + "\n".join(
                    f"{a['try']}. status={a['status']} detail={a['detail']}" for a in attempts[:8]
                )
            results.append({"ok": False, "owner_id": owner_id, "detail": extra_detail, "payload": payload})
    if len(owner_ids) == 1:
        r = results[0]
        if r.get("ok"):
            out = {"ok": True, "item": r.get("item"), "message": "확장소재 등록 완료"}
            if warnings:
                out["warnings"] = warnings
            return jsonify(out)
        return jsonify({"error": "확장소재 등록 실패", "details": r.get("detail"), "payload": r.get("payload"), "warnings": warnings}), 400
    out = {
        "ok": fail == 0,
        "total": len(owner_ids),
        "success": success,
        "fail": fail,
        "results": results,
        "message": f"확장소재 등록 완료 (성공 {success}건 / 실패 {fail}건)"
    }
    if warnings:
        out["warnings"] = warnings
    return jsonify(out)
@app.route("/bulk_upload_headlines", methods=["POST"])
def bulk_upload_headlines():
    api_key = str(request.form.get("api_key") or "").strip()
    secret_key = str(request.form.get("secret_key") or "").strip()
    cid = str(request.form.get("customer_id") or "").strip()
    upload = request.files.get("file")
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API Key / Secret Key / Customer ID가 필요합니다."}), 400
    if upload is None or not getattr(upload, "filename", ""):
        return jsonify({"error": "업로드할 파일을 선택해주세요."}), 400
    try:
        rows = _read_uploaded_table(upload)
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except Exception as e:
        return jsonify({"error": f"업로드 파일 처리 실패: {e}"}), 400
    if not rows:
        return jsonify({"error": "파일에서 등록할 데이터가 없습니다."}), 400
    owner_keys = ["ownerId", "owner_id", "ownerid", "nccAdgroupId", "adgroupId", "광고그룹id", "광고그룹아이디", "그룹id", "그룹아이디"]
    headline_keys = ["headline", "title", "name", "추가제목", "제목"]
    pin_keys = ["pin", "position", "노출위치", "노출가능위치", "노출가능위치지정", "위치", "위치지정"]
    results: List[Dict[str, Any]] = []
    success = fail = 0
    for idx, row in enumerate(rows, start=1):
        owner_id = str(_row_pick_value(row, owner_keys) or "").strip()
        headline = str(_row_pick_value(row, headline_keys) or "").strip()
        raw_pin = _row_pick_value(row, pin_keys)
        has_pin_field = _row_has_any_key(row, pin_keys)
        if not owner_id or not headline:
            fail += 1
            results.append(_result_item(idx, False, headline or f"{idx}행", "그룹ID/광고그룹ID와 추가제목은 필수입니다."))
            continue
        if len(headline) > 15:
            fail += 1
            results.append(_result_item(idx, False, headline, "추가제목은 15자 이내로 입력해야 합니다."))
            continue
        try:
            position = _normalize_headline_pin_input(raw_pin)
        except ValueError as e:
            fail += 1
            results.append(_result_item(idx, False, headline, str(e)))
            continue
        existing = _find_existing_extension(api_key, secret_key, cid, owner_id, "HEADLINE", {"headline": headline})
        if existing:
            if has_pin_field:
                upd = _apply_headline_pin_best_effort(api_key, secret_key, cid, existing, position)
                if not upd.get("ok"):
                    fail += 1
                    results.append(_result_item(idx, False, headline, upd.get("detail") or "노출위치 반영 실패"))
                    continue
            verify = _verify_headline_pin(api_key, secret_key, cid, owner_id, headline, position)
            if verify.get("ok"):
                success += 1
                results.append(_result_item(idx, True, headline, f"이미 등록됨 · {verify.get('detail')}"))
            else:
                fail += 1
                results.append(_result_item(idx, False, headline, f"이미 등록됨 · {verify.get('detail')}"))
            continue
        payload = _build_extension_payload(owner_id, "HEADLINE", {"headline": headline}, int(cid), position=position)
        res = _do_req("POST", api_key, secret_key, cid, "/ncc/ad-extensions", params={"ownerId": owner_id}, json_body=payload)
        if res.status_code not in [200, 201]:
            fail += 1
            results.append(_result_item(idx, False, headline, res.text))
            continue
        created_item = None
        try:
            created_item = _extract_extension_item(res.json())
        except Exception:
            created_item = None
        if has_pin_field:
            target_item = created_item or _find_existing_extension(api_key, secret_key, cid, owner_id, "HEADLINE", {"headline": headline})
            if isinstance(target_item, dict):
                upd = _apply_headline_pin_best_effort(api_key, secret_key, cid, target_item, position)
                if not upd.get("ok"):
                    fail += 1
                    results.append(_result_item(idx, False, headline, f"등록은 되었으나 노출위치 반영 실패: {upd.get('detail')}"))
                    continue
        verify = _verify_headline_pin(api_key, secret_key, cid, owner_id, headline, position)
        if verify.get("ok"):
            success += 1
            results.append(_result_item(idx, True, headline, f"등록 완료 · {verify.get('detail')}"))
        else:
            fail += 1
            results.append(_result_item(idx, False, headline, f"등록은 되었으나 {verify.get('detail')}"))
    return jsonify({
        "ok": fail == 0,
        "total": len(rows),
        "success": success,
        "fail": fail,
        "results": results,
        "message": f"추가제목 대량등록 완료 (성공 {success}건 / 실패 {fail}건)",
    })
@app.route("/create_shopping_ad_simple", methods=["POST"])
def create_shopping_ad_simple():
    d = request.json or {}
    adgroup_ids = _parse_target_ids(d, "adgroup_id", "adgroup_ids")
    reference_key = str(d.get("reference_key") or "").strip()
    ad_type = str(d.get("ad_type") or "SHOPPING_PRODUCT_AD").strip() or "SHOPPING_PRODUCT_AD"
    if not adgroup_ids or not reference_key:
        return jsonify({"error": "광고그룹과 상품번호(referenceKey)는 필수입니다."}), 400
    results = [_create_shopping_ad_for_adgroup(d, adg_id, reference_key, ad_type) for adg_id in adgroup_ids]
    success = sum(1 for r in results if r.get("ok"))
    fail = len(results) - success
    if len(adgroup_ids) == 1:
        r = results[0]
        if r.get("ok"):
            return jsonify({"ok": True, "item": r.get("item"), "message": "쇼핑 소재 등록 완료"})
        payload = {
            "customerId": int(d.get("customer_id")),
            "nccAdgroupId": adgroup_ids[0],
            "type": ad_type,
            "referenceKey": reference_key,
            "useGroupBidAmt": bool(d.get("use_group_bid_amt", True)),
            "bidAmt": int(d.get("bid_amt") or 0),
            "userLock": bool(d.get("user_lock", False)),
            "ad": {},
        }
        return jsonify({"error": "쇼핑 소재 등록 실패", "details": r.get("detail"), "payload": payload}), 400
    return jsonify({
        "ok": True,
        "total": len(results),
        "success": success,
        "fail": fail,
        "results": results,
        "message": f"쇼핑 소재 등록 완료 · 성공 {success}건 / 실패 {fail}건",
    })
@app.route("/create_extension_raw", methods=["POST"])
def create_extension_raw():
    d = request.json or {}
    owner_id = str(d.get("owner_id") or "").strip()
    raw_json = str(d.get("raw_json") or "").strip()
    if not owner_id or not raw_json:
        return jsonify({"error": "광고그룹과 JSON 입력이 필요합니다."}), 400
    try:
        payload = json.loads(raw_json)
    except Exception as e:
        return jsonify({"error": f"JSON 파싱 실패: {e}"}), 400
    payload.setdefault("ownerId", owner_id)
    payload.setdefault("customerId", int(d.get("customer_id")))
    ext_type = _normalize_extension_type(payload.get("type"))
    payload["type"] = ext_type
    res = _do_req("POST", d.get("api_key"), d.get("secret_key"), d.get("customer_id"), "/ncc/ad-extensions", params={"ownerId": owner_id}, json_body=payload)
    if res.status_code in [200, 201]:
        return jsonify({"ok": True, "item": res.json(), "message": "확장소재 생성 완료"})
    return jsonify({"error": "확장소재 생성 실패", "details": res.text, "payload": payload}), 400
@app.route("/create_restricted_keywords_simple", methods=["POST"])
def create_restricted_keywords_simple():
    d = request.json or {}
    raw_adgroup_ids = d.get("adgroup_ids") or []
    adgroup_ids: List[str] = []
    if isinstance(raw_adgroup_ids, list):
        for x in raw_adgroup_ids:
            sx = str(x or "").strip()
            if sx and sx not in adgroup_ids:
                adgroup_ids.append(sx)
    adgroup_id = str(d.get("adgroup_id") or "").strip()
    if adgroup_id and adgroup_id not in adgroup_ids:
        adgroup_ids.append(adgroup_id)
    keywords = [x.strip() for x in str(d.get("keywords") or "").replace(",", "\n").splitlines() if x.strip()]
    if not adgroup_ids or not keywords:
        return jsonify({"error": "광고그룹과 제외키워드는 필수입니다."}), 400
    api_key = d.get("api_key")
    secret_key = d.get("secret_key")
    customer_id = d.get("customer_id")
    requested_type = str(d.get("match_type") or "EXACT").strip()
    new_rows = []
    skipped = []
    lookup_fail = []
    for adgroup_id in adgroup_ids:
        res_existing, existing_rows = _fetch_restricted_keywords(api_key, secret_key, customer_id, adgroup_id)
        existing_set = set()
        if res_existing.status_code == 200:
            for item in existing_rows:
                if isinstance(item, dict):
                    kw = str(item.get("keyword") or item.get("restrictedKeyword") or "").strip().lower()
                    if kw:
                        existing_set.add(kw)
                elif isinstance(item, str):
                    kw = item.strip().lower()
                    if kw:
                        existing_set.add(kw)
        else:
            lookup_fail.append(adgroup_id)
        for kw in keywords:
            if kw.lower() in existing_set:
                skipped.append({"adgroup_id": adgroup_id, "keyword": kw})
                continue
            new_rows.append({"nccAdgroupId": adgroup_id, "keyword": kw, "type": requested_type})
    success, fail, results = _bulk_create_restricted_keywords(api_key, secret_key, customer_id, new_rows)
    return jsonify({
        "ok": True,
        "total_adgroups": len(adgroup_ids),
        "total_keywords": len(keywords),
        "submitted": len(new_rows),
        "skipped_duplicates": skipped,
        "lookup_fail_adgroups": lookup_fail,
        "success": success,
        "fail": fail,
        "results": results,
    })
@app.route("/copy_entities_to_adgroups", methods=["POST"])
def copy_entities_to_adgroups():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    source_ids = _unique_keep_order(d.get("source_ids") or [])
    target_adgroup_ids = _unique_keep_order(d.get("target_adgroup_ids") or [])
    include_keywords = _boolish(d.get("include_keywords"), True)
    include_ads = _boolish(d.get("include_ads"), True)
    include_extensions = _boolish(d.get("include_extensions"), True)
    include_negatives = _boolish(d.get("include_negatives"), True)
    if not source_ids:
        return jsonify({"error": "복사 원본 광고그룹을 선택해주세요."}), 400
    if not target_adgroup_ids:
        return jsonify({"error": "대상 광고그룹을 선택해주세요."}), 400
    success = 0
    fail = 0
    skipped_same = 0
    all_errors: List[str] = []
    for src_id in source_ids:
        for target_id in target_adgroup_ids:
            if str(src_id) == str(target_id):
                skipped_same += 1
                continue
            errs, summary = _copy_adgroup_children(
                api_key, secret_key, cid,
                str(src_id), str(target_id), None,
                include_keywords=include_keywords,
                include_ads=include_ads,
                include_extensions=include_extensions,
                include_negatives=include_negatives,
                return_summary=True,
            )
            summary_line = _format_copy_summary(summary)
            if errs:
                fail += 1
                all_errors.append(f"[원본 {src_id} → 대상 {target_id}] {summary_line}")
                all_errors.extend([f"[원본 {src_id} → 대상 {target_id}] {e}" for e in errs])
            else:
                success += 1
                all_errors.append(f"[원본 {src_id} → 대상 {target_id}] {summary_line}")
    msg = f"항목 복사 완료! (성공: {success}, 실패: {fail}, 동일 그룹 건너뜀: {skipped_same})"
    if all_errors:
        msg += "\n" + "\n".join(all_errors[:10])
    _cache_invalidate(d.get("api_key"), d.get("secret_key"), d.get("customer_id"))
    return jsonify({"ok": True, "message": msg})
@app.route("/copy_campaigns", methods=["POST"])
def copy_campaigns():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    src_ids, suffix = d.get("source_ids", []), d.get("suffix", "_복사본")
    try:
        copy_count = int(d.get("copy_count") or 1)
    except (TypeError, ValueError):
        return jsonify({"error": "복사 개수는 1 이상의 숫자여야 합니다."}), 400
    if copy_count <= 0:
        return jsonify({"error": "복사 개수는 1 이상의 숫자여야 합니다."}), 400
    custom_names_raw = d.get("custom_names") if d.get("custom_names") is not None else d.get("custom_names_text")
    results, all_errors = {"success": 0, "fail": 0}, []
    source_name_map: Dict[str, str] = {}
    for src_id in src_ids:
        r_get = _do_req("GET", api_key, secret_key, cid, f"/ncc/campaigns/{src_id}")
        if r_get.status_code != 200:
            results["fail"] += 1
            all_errors.append(f"[원본 캠페인 {src_id}] 조회 실패: {r_get.text}")
            continue
        src = r_get.json() or {}
        source_name_map[str(src_id)] = str(src.get("name") or "").strip()
    valid_src_ids = [str(src_id) for src_id in src_ids if str(src_id) in source_name_map]
    if valid_src_ids:
        try:
            planned_names = _resolve_copy_names([source_name_map[x] for x in valid_src_ids], copy_count, suffix, custom_names_raw)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        camp_res, camp_rows = _fetch_campaigns(api_key, secret_key, cid)
        if camp_res.status_code != 200:
            return jsonify({"error": f"기존 캠페인 이름 조회 실패: {camp_res.text}"}), 400
        name_err = _validate_name_candidates(planned_names, [row.get("name") for row in (camp_rows or [])], "캠페인")
        if name_err:
            return jsonify({"error": name_err}), 400
        planned_iter = iter(planned_names)
    else:
        planned_iter = iter([])
    for src_id in src_ids:
        if str(src_id) not in source_name_map:
            continue
        r_get = _do_req("GET", api_key, secret_key, cid, f"/ncc/campaigns/{src_id}")
        if r_get.status_code != 200:
            continue
        src = r_get.json() or {}
        source_name = str(src.get("name") or "").strip()
        r_adgs, src_adgroups = _fetch_adgroups(api_key, secret_key, cid, str(src_id))
        if r_adgs.status_code != 200:
            results["fail"] += copy_count
            all_errors.append(f"[{source_name or src_id}] 원본 광고그룹 조회 실패: {r_adgs.text}")
            continue
        for idx in range(1, copy_count + 1):
            new_camp = {
                "customerId": int(cid),
                "name": next(planned_iter, _build_copy_name(source_name, idx, copy_count, suffix)),
                "campaignTp": src.get("campaignTp"),
                "useDailyBudget": src.get("useDailyBudget", False),
                "dailyBudget": src.get("dailyBudget", 0),
            }
            r_post = _do_req("POST", api_key, secret_key, cid, "/ncc/campaigns", json_body=new_camp)
            if r_post.status_code not in [200, 201]:
                results["fail"] += 1
                all_errors.append(f"[{new_camp['name']}] 생성 실패: {r_post.text}")
                continue
            results["success"] += 1
            new_campaign_id = str((r_post.json() or {}).get("nccCampaignId") or "").strip()
            if not new_campaign_id:
                all_errors.append(f"[{new_camp['name']}] 신규 캠페인 ID 확인 실패")
                continue
            for row in (src_adgroups or []):
                src_adg_id = str(row.get("id") or row.get("nccAdgroupId") or "").strip()
                if not src_adg_id:
                    continue
                r_adg = _do_req("GET", api_key, secret_key, cid, f"/ncc/adgroups/{src_adg_id}")
                if r_adg.status_code != 200:
                    all_errors.append(f"[{new_camp['name']}] 원본 광고그룹 {src_adg_id} 조회 실패: {r_adg.text}")
                    continue
                src_adg = r_adg.json() or {}
                new_adg = _extract_adgroup(src_adg, new_campaign_id, cid, None)
                r_new_adg = _do_req("POST", api_key, secret_key, cid, "/ncc/adgroups", json_body=new_adg)
                if r_new_adg.status_code not in [200, 201]:
                    all_errors.append(f"[{new_camp['name']} > {src_adg.get('name')}] 광고그룹 생성 실패: {r_new_adg.text}")
                    continue
                new_adg_id = str((r_new_adg.json() or {}).get("nccAdgroupId") or "").strip()
                errs, summary = _copy_adgroup_children(
                    api_key, secret_key, cid,
                    src_adg_id, new_adg_id, None,
                    include_keywords=True,
                    include_ads=True,
                    include_extensions=True,
                    include_negatives=True,
                    return_summary=True,
                )
                all_errors.append(f"[{new_camp['name']} > {src_adg.get('name')}] {_format_copy_summary(summary)}")
                all_errors.extend([f"[{new_camp['name']} > {src_adg.get('name')}] {e}" for e in errs])
                _, media_msgs = _copy_adgroup_media_settings(api_key, secret_key, cid, src_adg_id, new_adg_id)
                if media_msgs:
                    for msg in media_msgs:
                        if msg and (("실패" in msg) or ("없음" in msg) or ("비어" in msg) or ("조회" in msg)):
                            all_errors.append(f"[{new_camp['name']} > {src_adg.get('name')}] 매체 설정: {msg}")
                _, extra_target_msgs = _copy_adgroup_extra_target_settings(api_key, secret_key, cid, src_adg_id, new_adg_id)
                if extra_target_msgs:
                    for msg in extra_target_msgs:
                        if msg and (("실패" in msg) or ("없음" in msg) or ("비어" in msg) or ("조회" in msg)):
                            all_errors.append(f"[{new_camp['name']} > {src_adg.get('name')}] 타겟 설정: {msg}")
                ok_search_opts, search_opts_msg = _copy_adgroup_search_option_settings(api_key, secret_key, cid, src_adg_id, new_adg_id)
                if (not ok_search_opts) and search_opts_msg:
                    all_errors.append(f"[{new_camp['name']} > {src_adg.get('name')}] 검색옵션: {search_opts_msg}")
    msg = f"캠페인 복사 완료!\n(성공: {results['success']}개, 실패: {results['fail']}개)"
    if all_errors:
        msg += "\n" + "\n".join(all_errors[:10])
    _cache_invalidate(d.get("api_key"), d.get("secret_key"), d.get("customer_id"))
    return jsonify({"ok": True, "message": msg})
@app.route("/copy_adgroups_to_target", methods=["POST"])
def copy_adgroups_to_target():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    source_ids = _unique_keep_order(d.get("source_ids") or [])
    source_campaign_ids = _unique_keep_order(d.get("source_campaign_ids") or [])
    src_ids, resolve_warnings = _resolve_copy_source_adgroup_ids(api_key, secret_key, cid, source_ids, source_campaign_ids)
    target_camp_id = d.get("target_campaign_id")
    suffix = d.get("suffix", "_복사본")
    try:
        copy_count = int(d.get("copy_count") or 1)
    except (TypeError, ValueError):
        return jsonify({"error": "복사 개수는 1 이상의 숫자여야 합니다."}), 400
    if copy_count <= 0:
        return jsonify({"error": "복사 개수는 1 이상의 숫자여야 합니다."}), 400
    custom_names_raw = d.get("custom_names") if d.get("custom_names") is not None else d.get("custom_names_text")
    biz_channel_id = d.get("biz_channel_id")
    include_keywords = _boolish(d.get("include_keywords"), True)
    include_ads = _boolish(d.get("include_ads"), True)
    include_extensions = _boolish(d.get("include_extensions"), True)
    include_negatives = _boolish(d.get("include_negatives"), True)
    copy_media = _boolish(d.get("copy_media"), True)
    copy_as_off = _boolish(d.get("copy_as_off"), False)
    if not src_ids:
        return jsonify({"error": "복사할 광고그룹 또는 캠페인을 선택해주세요."}), 400
    if not str(target_camp_id or "").strip():
        return jsonify({"error": "대상 캠페인을 선택해주세요."}), 400
    source_name_map: Dict[str, str] = {}
    for src_id in src_ids:
        r_get = _do_req("GET", api_key, secret_key, cid, f"/ncc/adgroups/{src_id}")
        if r_get.status_code != 200:
            continue
        src_obj = r_get.json() or {}
        source_name_map[str(src_id)] = str(src_obj.get("name") or "").strip()
    valid_src_ids = [str(src_id) for src_id in src_ids if str(src_id) in source_name_map]
    if valid_src_ids:
        try:
            planned_names = _resolve_copy_names([source_name_map[x] for x in valid_src_ids], copy_count, suffix, custom_names_raw)
        except ValueError as exc:
            return jsonify({"error": str(exc)}), 400
        adg_res, existing_target_rows = _fetch_adgroups(api_key, secret_key, cid, str(target_camp_id), enrich_media=False)
        if adg_res.status_code != 200:
            return jsonify({"error": f"대상 캠페인 광고그룹 이름 조회 실패: {adg_res.text}"}), 400
        name_err = _validate_name_candidates(planned_names, [row.get("name") for row in (existing_target_rows or [])], "광고그룹")
        if name_err:
            return jsonify({"error": name_err}), 400
        planned_iter = iter(planned_names)
    else:
        planned_iter = iter([])
    results, all_errors = {"success": 0, "fail": 0}, list(resolve_warnings)
    for src_id in src_ids:
        r_get = _do_req("GET", api_key, secret_key, cid, f"/ncc/adgroups/{src_id}")
        if r_get.status_code != 200:
            results["fail"] += 1
            all_errors.append(f"[원본 {src_id}] 조회 실패: {r_get.text}")
            continue
        src_obj = r_get.json() or {}
        source_name = str(src_obj.get("name") or "").strip()
        for idx in range(1, copy_count + 1):
            new_adg = _extract_adgroup(src_obj, target_camp_id, cid, biz_channel_id)
            new_adg["name"] = next(planned_iter, _build_copy_name(source_name, idx, copy_count, suffix))
            r_post = _do_req("POST", api_key, secret_key, cid, "/ncc/adgroups", json_body=new_adg)
            if r_post.status_code in [200, 201]:
                results["success"] += 1
                new_adg_id = str((r_post.json() or {}).get("nccAdgroupId") or "").strip()
                errs, summary = _copy_adgroup_children(
                    api_key, secret_key, cid, src_id, new_adg_id, biz_channel_id,
                    include_keywords=include_keywords, include_ads=include_ads,
                    include_extensions=include_extensions, include_negatives=include_negatives,
                    return_summary=True,
                )
                all_errors.append(f"[{new_adg['name']}] {_format_copy_summary(summary)}")
                all_errors.extend([f"[{new_adg['name']}] {e}" for e in errs])
                if copy_media and new_adg_id:
                    ok_media, media_msgs = _copy_adgroup_media_settings(api_key, secret_key, cid, str(src_id), new_adg_id)
                    if media_msgs:
                        for msg in media_msgs:
                            if msg and (("실패" in msg) or ("없음" in msg) or ("비어" in msg) or ("조회" in msg)):
                                all_errors.append(f"[{new_adg['name']}] 매체 설정: {msg}")
                if new_adg_id:
                    _, extra_target_msgs = _copy_adgroup_extra_target_settings(api_key, secret_key, cid, str(src_id), new_adg_id)
                    if extra_target_msgs:
                        for msg in extra_target_msgs:
                            if msg and (("실패" in msg) or ("없음" in msg) or ("비어" in msg) or ("조회" in msg)):
                                all_errors.append(f"[{new_adg['name']}] 타겟 설정: {msg}")
                if new_adg_id:
                    ok_search_opts, search_opts_msg = _copy_adgroup_search_option_settings(api_key, secret_key, cid, str(src_id), new_adg_id)
                    if (not ok_search_opts) and search_opts_msg:
                        all_errors.append(f"[{new_adg['name']}] 검색옵션: {search_opts_msg}")
                if copy_as_off and new_adg_id:
                    ok_off, off_msg = _set_user_lock_for_entity(api_key, secret_key, cid, "adgroup", new_adg_id, False)
                    if not ok_off and off_msg:
                        all_errors.append(f"[{new_adg['name']}] OFF 설정 실패: {off_msg}")
            else:
                results["fail"] += 1
                all_errors.append(f"[{new_adg['name']}] 생성 실패: {r_post.text}")
    _cache_invalidate(api_key, secret_key, cid)
    return jsonify({"ok": True, "message": f"복사 완료! (성공: {results['success']}, 실패: {results['fail']})\n" + "\n".join(all_errors[:10])})
@app.route("/rename_adgroups_bulk", methods=["POST"])
def rename_adgroups_bulk():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    entity_ids = [str(x).strip() for x in (d.get("entity_ids") or []) if str(x).strip()]
    target_names = _parse_multiline_names(d.get("target_names") if d.get("target_names") is not None else d.get("target_names_text"))
    if not entity_ids:
        return jsonify({"error": "이름을 변경할 광고그룹을 선택해주세요."}), 400
    if not target_names:
        return jsonify({"error": "변경할 광고그룹명을 1개 이상 입력해주세요."}), 400
    if len(entity_ids) != len(target_names):
        return jsonify({"error": f"선택한 광고그룹 수({len(entity_ids)})와 입력한 이름 수({len(target_names)})가 다릅니다."}), 400
    detail_map: Dict[str, Dict[str, Any]] = {}
    campaign_to_selected: Dict[str, List[str]] = defaultdict(list)
    for adg_id in entity_ids:
        res_get, obj = _fetch_adgroup_detail(api_key, secret_key, cid, adg_id)
        if res_get.status_code != 200 or not obj:
            return jsonify({"error": f"광고그룹 조회 실패 ({adg_id}): {res_get.text}"}), 400
        detail_map[adg_id] = obj
        campaign_id = str(obj.get("nccCampaignId") or "").strip()
        campaign_to_selected[campaign_id].append(adg_id)
    rename_map = {entity_ids[idx]: target_names[idx] for idx in range(len(entity_ids))}
    for campaign_id, selected_ids in campaign_to_selected.items():
        adg_res, rows = _fetch_adgroups(api_key, secret_key, cid, campaign_id, enrich_media=False)
        if adg_res.status_code != 200:
            return jsonify({"error": f"캠페인 내 광고그룹 이름 조회 실패 ({campaign_id}): {adg_res.text}"}), 400
        existing_names = []
        selected_id_set = set(selected_ids)
        for row in (rows or []):
            row_id = str(row.get("id") or row.get("nccAdgroupId") or "").strip()
            if row_id in selected_id_set:
                continue
            existing_names.append(row.get("name"))
        planned_names = [rename_map[x] for x in selected_ids]
        name_err = _validate_name_candidates(planned_names, existing_names, "광고그룹")
        if name_err:
            return jsonify({"error": name_err}), 400
    success = fail = 0
    messages: List[str] = []
    for adg_id in entity_ids:
        obj = detail_map.get(adg_id) or {}
        new_name = rename_map.get(adg_id, "")
        old_name = str(obj.get("name") or adg_id)
        obj["name"] = new_name
        res_put = _do_req("PUT", api_key, secret_key, cid, f"/ncc/adgroups/{adg_id}", params={"fields": "name"}, json_body=obj)
        if res_put.status_code in [200, 201]:
            success += 1
            messages.append(f"[{old_name}] → [{new_name}] 변경 완료")
        else:
            fail += 1
            messages.append(f"[{old_name}] 이름 변경 실패: {res_put.text}")
    _cache_invalidate(api_key, secret_key, cid)
    return jsonify({
        "ok": success > 0,
        "message": f"광고그룹명 일괄 변경 완료! (성공: {success}, 실패: {fail})" + ("\n" + "\n".join(messages[:10]) if messages else ""),
        "success": success,
        "fail": fail,
    }), (200 if success > 0 else 400)
@app.route("/update_media", methods=["POST"])
def update_media():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    entity_ids = [str(x).strip() for x in (d.get("entity_ids") or []) if str(x).strip()]
    media_type = str(d.get("media_type") or "ALL").strip().upper()
    media_detail = _normalize_media_detail(d.get("media_detail"))
    if not entity_ids:
        return jsonify({"error": "선택된 광고그룹이 없습니다."}), 400
    results = []
    success = fail = 0
    for eid in entity_ids:
        row_ok = True
        row_msgs: List[str] = []
        ok_pm, detail_pm = _update_pc_mobile_target(api_key, secret_key, cid, eid, media_type)
        row_ok = row_ok and ok_pm
        row_msgs.append(f"PC/모바일: {detail_pm}")
        ok_media, detail_media = _update_media_target(api_key, secret_key, cid, eid, media_detail)
        row_ok = row_ok and ok_media
        row_msgs.append(f"세부매체: {detail_media}")
        results.append({"nccAdgroupId": eid, "ok": row_ok, "detail": " | ".join(row_msgs)})
        if row_ok:
            success += 1
        else:
            fail += 1
    status_code = 200 if success > 0 else 400
    return jsonify({"ok": success > 0, "message": f"총 {len(entity_ids)}개 매체 변경 성공: {success}개 / 실패: {fail}개", "success": success, "fail": fail, "results": results}), status_code
@app.route("/update_adgroup_options", methods=["POST"])
def update_adgroup_options():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    entity_ids = [str(x).strip() for x in (d.get("entity_ids") or []) if str(x).strip()]
    media_type = d.get("media_type")
    media_detail = _normalize_media_detail(d.get("media_detail"))
    use_keyword_plus = d.get("use_keyword_plus")
    use_close_variant = d.get("use_close_variant")
    keyword_plus_weight = d.get("keyword_plus_weight")
    if not entity_ids:
        return jsonify({"error": "선택된 광고그룹이 없습니다."}), 400
    results = []
    success = fail = skipped = 0
    for adg_id in entity_ids:
        detail_res, adgroup_obj = _fetch_adgroup_detail(api_key, secret_key, cid, adg_id)
        if detail_res.status_code != 200 or not adgroup_obj:
            fail += 1
            results.append({"nccAdgroupId": adg_id, "ok": False, "detail": f"광고그룹 조회 실패: {detail_res.text}"})
            continue
        row_msgs: List[str] = []
        row_ok = True
        if str(media_type or "").strip() != "" or d.get("media_detail") is not None:
            ok_media_pm, detail_media_pm = _update_pc_mobile_target(api_key, secret_key, cid, adg_id, media_type or "ALL")
            row_msgs.append(f"PC/모바일 매체: {detail_media_pm}")
            row_ok = row_ok and ok_media_pm
            ok_media_network, detail_media_network = _update_media_target(api_key, secret_key, cid, adg_id, media_detail)
            row_msgs.append(f"세부 매체: {detail_media_network}")
            row_ok = row_ok and ok_media_network
        is_web_site = str(adgroup_obj.get("adgroupType") or "").upper() == "WEB_SITE"
        if use_keyword_plus is not None or use_close_variant is not None or str(keyword_plus_weight or "").strip() != "":
            if not is_web_site:
                skipped += 1
                row_msgs.append("파워링크 그룹이 아니어서 확장검색/일치검색은 건너뜀")
            else:
                kwp = None
                if str(keyword_plus_weight or "").strip() != "":
                    try:
                        kwp = int(keyword_plus_weight)
                    except Exception:
                        kwp = None
                ok_opts, detail_opts = _update_adgroup_search_options(api_key, secret_key, cid, adg_id, use_keyword_plus=None if use_keyword_plus is None else bool(use_keyword_plus), keyword_plus_weight=kwp, use_close_variant=None if use_close_variant is None else bool(use_close_variant))
                row_msgs.append(f"검색옵션: {detail_opts}")
                row_ok = row_ok and ok_opts
        if row_ok:
            success += 1
        else:
            fail += 1
        results.append({"nccAdgroupId": adg_id, "ok": row_ok, "detail": " | ".join(row_msgs) if row_msgs else "변경 완료"})
    status_code = 200 if success > 0 else 400
    return jsonify({
        "ok": success > 0,
        "message": f"총 {len(entity_ids)}개 광고그룹 설정 변경 완료 · 성공 {success}개 / 실패 {fail}개 / 건너뜀 {skipped}개",
        "success": success,
        "fail": fail,
        "skipped": skipped,
        "results": results,
    }), status_code
@app.route("/apply_target_settings_bulk", methods=["POST"])
def apply_target_settings_bulk():
    d = request.get_json(silent=True) or {}
    api_key = d.get("api_key")
    secret_key = d.get("secret_key")
    cid = d.get("customer_id")
    source_adgroup_id = str(d.get("source_adgroup_id") or "").strip()
    target_scope = str(d.get("target_scope") or "adgroup").strip().lower()
    include_extra_targets = _boolish(d.get("include_extra_targets"), True)
    include_schedule = _boolish(d.get("include_schedule"), True)
    campaign_ids = [str(x).strip() for x in (d.get("campaign_ids") or []) if str(x).strip()]
    adgroup_ids = [str(x).strip() for x in (d.get("adgroup_ids") or []) if str(x).strip()]
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API Key / Secret Key / Customer ID가 필요합니다."}), 400
    if not source_adgroup_id:
        return jsonify({"error": "원본 광고그룹을 선택해주세요."}), 400
    if not include_extra_targets and not include_schedule:
        return jsonify({"error": "적용할 타겟 항목을 1개 이상 선택해주세요."}), 400
    target_adgroup_ids, warnings = _resolve_bulk_target_adgroup_ids(api_key, secret_key, cid, target_scope, campaign_ids, adgroup_ids)
    target_adgroup_ids = [x for x in target_adgroup_ids if x and x != source_adgroup_id]
    if not target_adgroup_ids:
        msg = "적용할 대상 광고그룹이 없습니다."
        if warnings:
            msg += "\n" + "\n".join(warnings[:10])
        return jsonify({"error": msg}), 400
    source_copyable_target_types: List[str] = []
    source_extra_error = ""
    source_schedule_ready = False
    source_schedule_msg = ""
    if include_extra_targets:
        res_src_targets, src_targets = _fetch_all_target_objects(api_key, secret_key, cid, source_adgroup_id)
        if res_src_targets.status_code == 200:
            source_copyable_target_types = _unique_keep_order([
                _normalize_target_type_name((row or {}).get("targetTp"))
                for row in (src_targets or [])
                if _is_extra_copy_target_type((row or {}).get("targetTp"))
            ])
        else:
            source_extra_error = f"추가 타겟 원본 조회 실패: {res_src_targets.text}"
    if include_schedule:
        res_schedule_src, schedule_rows = _fetch_schedule_entries(api_key, secret_key, cid, source_adgroup_id)
        if res_schedule_src.status_code == 200 and schedule_rows:
            source_schedule_ready = True
        elif res_schedule_src.status_code == 404:
            source_schedule_msg = "SCHEDULE 원본 설정 없음"
        elif res_schedule_src.status_code == 405:
            source_schedule_msg = f"SCHEDULE 원본 조회 미지원(405): {res_schedule_src.text}"
        else:
            source_schedule_msg = f"SCHEDULE 원본 조회 실패: {res_schedule_src.text}"
    success = 0
    fail = 0
    details: List[str] = list(warnings)
    for target_adgroup_id in target_adgroup_ids:
        row_msgs: List[str] = []
        hard_fail = False
        applied_any = False
        if include_extra_targets:
            if source_copyable_target_types:
                ok_extra, extra_msgs = _copy_adgroup_extra_targets_only(api_key, secret_key, cid, source_adgroup_id, target_adgroup_id)
                row_msgs.extend([msg for msg in (extra_msgs or []) if msg])
                if ok_extra:
                    applied_any = True
                else:
                    hard_fail = True
            else:
                row_msgs.append(source_extra_error or "추가 타겟 원본 설정 없음")
        if include_schedule:
            if source_schedule_ready:
                ok_schedule, msg_schedule = _copy_schedule_criterion_exact(api_key, secret_key, cid, source_adgroup_id, target_adgroup_id)
                if ok_schedule:
                    applied_any = True
                elif msg_schedule:
                    row_msgs.append(msg_schedule)
                    hard_fail = True
            else:
                row_msgs.append(source_schedule_msg or "SCHEDULE 원본 설정 없음")
        if not applied_any:
            row_msgs.insert(0, "실제 반영된 항목 없음")
        row_ok = applied_any and not hard_fail
        if row_ok:
            success += 1
        else:
            fail += 1
        if row_msgs:
            details.append(f"[{target_adgroup_id}] " + " | ".join(_unique_keep_order(row_msgs)[:5]))
    status_code = 200 if success > 0 else 400
    return jsonify({
        "ok": success > 0,
        "message": f"타겟 설정 일괄 적용 완료 · 성공 {success}개 / 실패 {fail}개" + (("\n" + "\n".join(details[:12])) if details else ""),
        "success": success,
        "fail": fail,
        "details": details[:50],
    }), status_code
@app.route("/update_budget", methods=["POST"])
def update_budget():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    entity_type, entity_ids, budget = d.get("entity_type"), d.get("entity_ids", []), int(d.get("budget", 0))
    use_daily_budget = bool(d.get("use_daily_budget", budget > 0))
    results = {"success": 0, "fail": 0}
    for eid in entity_ids:
        uri = f"/ncc/campaigns/{str(eid).strip()}" if entity_type == "campaign" else f"/ncc/adgroups/{str(eid).strip()}"
        r_get = _do_req("GET", api_key, secret_key, cid, uri)
        if r_get.status_code != 200:
            results["fail"] += 1
            continue
        obj = r_get.json()
        obj["useDailyBudget"] = use_daily_budget
        obj["dailyBudget"] = budget if use_daily_budget else 0
        if "budget" in obj:
            del obj["budget"]
        r_put = _do_req("PUT", api_key, secret_key, cid, uri, params={"fields": "budget"}, json_body=obj)
        if r_put.status_code == 200:
            results["success"] += 1
        else:
            results["fail"] += 1
    return jsonify({"ok": True, "message": f"총 {len(entity_ids)}개 예산 업데이트 성공: {results['success']}개 / 실패: {results['fail']}개"})
def _normalize_schedule_days(values: Any) -> List[int]:
    out: List[int] = []
    if not isinstance(values, list):
        return out
    for v in values:
        try:
            n = int(v)
        except Exception:
            continue
        if n in DAY_NUM_TO_CODE and n not in out:
            out.append(n)
    return out
def _normalize_schedule_hours(values: Any) -> List[int]:
    raw: List[int] = []
    if not isinstance(values, list):
        return raw
    uniq: List[int] = []
    seen = set()
    for v in values:
        try:
            n = int(v)
        except Exception:
            continue
        if 0 <= n <= 23 and n not in seen:
            seen.add(n)
            uniq.append(n)
    uniq.sort()
    if not uniq:
        return []
    is_contiguous = len(uniq) >= 2 and all(uniq[i] == uniq[i - 1] + 1 for i in range(1, len(uniq)))
    if is_contiguous:
        return uniq[:-1]
    return uniq
def _build_schedule_codes(days: List[int], hours: List[int]) -> List[str]:
    codes: List[str] = []
    for d_num in days:
        day_code = DAY_NUM_TO_CODE.get(int(d_num))
        if not day_code:
            continue
        for h in hours:
            start_h = int(h)
            end_h = start_h + 1
            if 0 <= start_h <= 23 and 1 <= end_h <= 24:
                codes.append(f"SD{day_code}{start_h:02d}{end_h:02d}")
    return codes

def _normalize_schedule_blocks(values: Any) -> List[Dict[str, Any]]:
    blocks: List[Dict[str, Any]] = []
    if not isinstance(values, list):
        return blocks
    for item in values:
        if not isinstance(item, dict):
            continue
        try:
            start_h = int(item.get("startHour"))
            end_h = int(item.get("endHour"))
            bid_weight = int(item.get("bidWeight", 100))
        except Exception:
            continue
        if not (0 <= start_h <= 23):
            continue
        if not (1 <= end_h <= 24):
            continue
        if end_h <= start_h:
            continue
        if not (10 <= bid_weight <= 500):
            continue
        block_days = _normalize_schedule_days(item.get("days") or [])
        blocks.append({"startHour": start_h, "endHour": end_h, "bidWeight": bid_weight, "days": block_days})
    return blocks

def _build_schedule_weighted_codes(days: List[int], hours: List[int], bid_weight: int = 100, schedule_blocks: Optional[List[Dict[str, Any]]] = None) -> List[Tuple[str, int]]:
    dedup: Dict[str, int] = {}
    if schedule_blocks:
        for block in schedule_blocks:
            block_days = _normalize_schedule_days(block.get("days") or []) or list(days or [])
            start_h = int(block.get("startHour", 0))
            end_h = int(block.get("endHour", 0))
            bw = int(block.get("bidWeight", 100))
            for d_num in block_days:
                day_code = DAY_NUM_TO_CODE.get(int(d_num))
                if not day_code:
                    continue
                for h in range(start_h, end_h):
                    if 0 <= h <= 23:
                        dedup[f"SD{day_code}{h:02d}{h+1:02d}"] = bw
    else:
        for code in _build_schedule_codes(days, hours):
            dedup[code] = int(bid_weight)
    return [(code, weight) for code, weight in dedup.items()]

def _extract_schedule_weight_map(rows: List[Dict[str, Any]]) -> Dict[str, int]:
    out: Dict[str, int] = {}
    for row in rows or []:
        code = str((row or {}).get("dictionaryCode") or (row or {}).get("code") or "").strip()
        if not code:
            continue
        try:
            out[code] = int((row or {}).get("bidWeight") or 100)
        except Exception:
            out[code] = 100
    return out

def _apply_schedule_action(existing_map: Dict[str, int], incoming_weighted_codes: List[Tuple[str, int]], action_mode: str) -> Dict[str, int]:
    mode = str(action_mode or "overwrite").strip().lower()
    incoming_map: Dict[str, int] = {}
    for code, weight in incoming_weighted_codes or []:
        s = str(code or "").strip()
        if not s:
            continue
        try:
            incoming_map[s] = int(weight)
        except Exception:
            incoming_map[s] = 100
    base = dict(existing_map or {})
    if mode == "overwrite":
        return incoming_map
    if mode == "add":
        merged = dict(base)
        merged.update(incoming_map)
        return merged
    if mode == "delete":
        for code in incoming_map.keys():
            base.pop(code, None)
        return base
    return incoming_map

def _put_schedule_weight_map(api_key: str, secret_key: str, cid: str, owner_id: str, final_map: Dict[str, int]):
    owner_id = str(owner_id or "").strip()
    final_codes = _unique_keep_order(list((final_map or {}).keys()))
    uri = f"/ncc/criterion/{owner_id}/SD"
    target_body = [{"customerId": int(cid), "ownerId": owner_id, "dictionaryCode": c, "type": "SD"} for c in final_codes]
    put_res = _do_req("PUT", api_key, secret_key, cid, uri, json_body=target_body)
    if put_res.status_code != 200:
        return False, "criterion_put", put_res.text
    weight_map: Dict[int, List[str]] = {}
    for code in final_codes:
        try:
            weight = int((final_map or {}).get(code, 100) or 100)
        except Exception:
            weight = 100
        weight_map.setdefault(weight, []).append(code)
    for weight, weight_codes in weight_map.items():
        if not weight_codes or int(weight) == 100:
            continue
        for i in range(0, len(weight_codes), 50):
            chunk = weight_codes[i:i + 50]
            bw_res = _do_req(
                "PUT",
                api_key,
                secret_key,
                cid,
                f"/ncc/criterion/{owner_id}/bidWeight",
                params={"codes": ",".join(chunk), "bidWeight": int(weight)},
            )
            if bw_res.status_code != 200:
                return False, "bid_weight_put", bw_res.text
    return True, "", ""

@app.route("/update_schedule", methods=["POST"])
def update_schedule():
    d = request.get_json(silent=True) or {}
    api_key = d.get("api_key")
    secret_key = d.get("secret_key")
    cid = d.get("customer_id")
    action_mode = str(d.get("actionMode") or "overwrite").strip().lower()
    if action_mode not in {"overwrite", "add", "delete"}:
        action_mode = "overwrite"
    adgroup_ids = _unique_keep_order([str(x).strip() for x in (d.get("adgroup_ids") or []) if str(x).strip()])
    days = _normalize_schedule_days(d.get("days") or [])
    raw_hours = []
    for x in (d.get("hours") or []):
        try:
            n = int(x)
        except Exception:
            continue
        if 0 <= n <= 23:
            raw_hours.append(n)
    raw_hours = sorted(set(raw_hours))
    hours = _normalize_schedule_hours(d.get("hours") or [])
    schedule_blocks = _normalize_schedule_blocks(d.get("scheduleBlocks") or [])
    try:
        bid_weight = int(d.get("bidWeight", 100))
    except Exception:
        bid_weight = 100
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API Key / Secret Key / Customer ID가 필요합니다."}), 400
    if not adgroup_ids:
        return jsonify({"error": "광고그룹이 선택되지 않았습니다."}), 400
    if not days and not any(block.get("days") for block in schedule_blocks):
        return jsonify({"error": "요일을 1개 이상 선택해주세요."}), 400
    if not schedule_blocks and not raw_hours:
        return jsonify({"error": "시간을 1개 이상 선택해주세요."}), 400
    if not schedule_blocks and not hours:
        return jsonify({"error": "연속 시간 선택 시 마지막 시간은 종료시각으로 처리됩니다. 예: 8~20 선택 → 실제 적용 8~19"}), 400
    weighted_codes = _build_schedule_weighted_codes(days, hours, bid_weight, schedule_blocks)
    codes = [code for code, _ in weighted_codes]
    if not codes and action_mode != "delete":
        return jsonify({"error": "적용할 시간대 코드가 없습니다."}), 400
    results = []
    success = 0
    fail = 0
    fetch_errors = []
    for owner_id in adgroup_ids:
        owner_id = str(owner_id).strip()
        existing_map: Dict[str, int] = {}
        if action_mode in {"add", "delete"}:
            res_existing, existing_rows = _fetch_schedule_entries(api_key, secret_key, cid, owner_id)
            if res_existing.status_code not in {200, 404}:
                fail += 1
                fetch_errors.append(f"[{owner_id}] 기존 스케줄 조회 실패: {res_existing.text}")
                results.append({"ownerId": owner_id, "ok": False, "step": "criterion_fetch", "details": res_existing.text})
                continue
            existing_map = _extract_schedule_weight_map(existing_rows or [])
        final_map = _apply_schedule_action(existing_map, weighted_codes, action_mode)
        ok_put, step, detail = _put_schedule_weight_map(api_key, secret_key, cid, owner_id, final_map)
        if not ok_put:
            fail += 1
            results.append({
                "ownerId": owner_id,
                "ok": False,
                "step": step,
                "details": detail,
                "codes": list(final_map.keys()),
            })
            continue
        success += 1
        results.append({"ownerId": owner_id, "ok": True, "codes": list(final_map.keys()), "weightMap": final_map, "actionMode": action_mode})
    status_code = 200 if success > 0 else 400
    action_label = {"overwrite": "덮어쓰기", "add": "추가", "delete": "삭제"}.get(action_mode, "덮어쓰기")
    return jsonify({
        "ok": success > 0,
        "message": f"총 {len(adgroup_ids)}개 스케줄 {action_label} 성공: {success}개 / 실패: {fail}개",
        "success": success,
        "fail": fail,
        "action_mode": action_mode,
        "raw_hours": raw_hours,
        "applied_hours": hours,
        "schedule_blocks": schedule_blocks,
        "fetch_errors": fetch_errors[:10],
        "results": results[:20],
    }), status_code

@app.route("/update_schedule_campaign_bulk", methods=["POST"])
def update_schedule_campaign_bulk():
    d = request.get_json(silent=True) or {}
    api_key = d.get("api_key")
    secret_key = d.get("secret_key")
    cid = d.get("customer_id")
    action_mode = str(d.get("actionMode") or "overwrite").strip().lower()
    if action_mode not in {"overwrite", "add", "delete"}:
        action_mode = "overwrite"
    campaign_ids = _unique_keep_order([str(x).strip() for x in (d.get("campaign_ids") or []) if str(x).strip()])
    days = _normalize_schedule_days(d.get("days") or [])
    raw_hours = []
    for x in (d.get("hours") or []):
        try:
            n = int(x)
        except Exception:
            continue
        if 0 <= n <= 23:
            raw_hours.append(n)
    raw_hours = sorted(set(raw_hours))
    hours = _normalize_schedule_hours(d.get("hours") or [])
    schedule_blocks = _normalize_schedule_blocks(d.get("scheduleBlocks") or [])
    try:
        bid_weight = int(d.get("bidWeight", 100))
    except Exception:
        bid_weight = 100
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API Key / Secret Key / Customer ID가 필요합니다."}), 400
    if not campaign_ids:
        return jsonify({"error": "캠페인이 선택되지 않았습니다."}), 400
    if not days and not any(block.get("days") for block in schedule_blocks):
        return jsonify({"error": "요일을 1개 이상 선택해주세요."}), 400
    if not schedule_blocks and not raw_hours:
        return jsonify({"error": "시간을 1개 이상 선택해주세요."}), 400
    if not schedule_blocks and not hours:
        return jsonify({"error": "연속 시간 선택 시 마지막 시간은 종료시각으로 처리됩니다. 예: 8~20 선택 → 실제 적용 8~19"}), 400
    adgroup_ids = []
    fetch_errors = []
    for camp_id in campaign_ids:
        r_adgs = _do_req("GET", api_key, secret_key, cid, "/ncc/adgroups", params={"nccCampaignId": camp_id})
        if r_adgs.status_code == 200:
            adgroup_ids.extend([adg.get("nccAdgroupId") for adg in r_adgs.json() or [] if adg.get("nccAdgroupId")])
        else:
            fetch_errors.append(f"[{camp_id}] {r_adgs.text}")
    adgroup_ids = _unique_keep_order(adgroup_ids)
    if not adgroup_ids:
        return jsonify({"error": "하위 광고그룹을 불러오지 못했습니다.", "details": fetch_errors[:10]}), 400
    weighted_codes = _build_schedule_weighted_codes(days, hours, bid_weight, schedule_blocks)
    codes = [code for code, _ in weighted_codes]
    if not codes and action_mode != "delete":
        return jsonify({"error": "적용할 시간대 코드가 없습니다."}), 400
    results = []
    success = 0
    fail = 0
    for owner_id in adgroup_ids:
        owner_id = str(owner_id).strip()
        existing_map: Dict[str, int] = {}
        if action_mode in {"add", "delete"}:
            res_existing, existing_rows = _fetch_schedule_entries(api_key, secret_key, cid, owner_id)
            if res_existing.status_code not in {200, 404}:
                fail += 1
                fetch_errors.append(f"[{owner_id}] 기존 스케줄 조회 실패: {res_existing.text}")
                results.append({"ownerId": owner_id, "ok": False, "step": "criterion_fetch", "details": res_existing.text})
                continue
            existing_map = _extract_schedule_weight_map(existing_rows or [])
        final_map = _apply_schedule_action(existing_map, weighted_codes, action_mode)
        ok_put, step, detail = _put_schedule_weight_map(api_key, secret_key, cid, owner_id, final_map)
        if not ok_put:
            fail += 1
            results.append({"ownerId": owner_id, "ok": False, "step": step, "details": detail, "codes": list(final_map.keys())})
            continue
        success += 1
        results.append({"ownerId": owner_id, "ok": True, "codes": list(final_map.keys()), "weightMap": final_map, "actionMode": action_mode})
    status_code = 200 if success > 0 else 400
    action_label = {"overwrite": "덮어쓰기", "add": "추가", "delete": "삭제"}.get(action_mode, "덮어쓰기")
    return jsonify({
        "ok": success > 0,
        "message": f"하위 광고그룹 총 {len(adgroup_ids)}개 스케줄 {action_label} 성공: {success}개 / 실패: {fail}개",
        "success": success,
        "fail": fail,
        "action_mode": action_mode,
        "raw_hours": raw_hours,
        "applied_hours": hours,
        "schedule_blocks": schedule_blocks,
        "fetch_errors": fetch_errors[:10],
        "results": results[:20],
    }), status_code

def _unique_keep_order(values: List[str]) -> List[str]:
    seen = set()
    out: List[str] = []
    for v in values:
        s = str(v or "").strip()
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return out
def _resolve_adgroup_ids(api_key: str, secret_key: str, cid: str, entity_type: str, entity_ids: List[str]):
    adgroup_ids: List[str] = []
    warnings: List[str] = []
    entity_type = str(entity_type or "adgroup").strip().lower()
    entity_ids = _unique_keep_order(entity_ids or [])
    if entity_type == "campaign":
        for camp_id in entity_ids:
            res, rows = _fetch_adgroups(api_key, secret_key, cid, str(camp_id).strip())
            if res.status_code != 200:
                warnings.append(f"캠페인 {camp_id} 하위 광고그룹 조회 실패: {res.text}")
                continue
            for row in rows:
                adg_id = str(row.get("id") or row.get("nccAdgroupId") or "").strip()
                if adg_id:
                    adgroup_ids.append(adg_id)
    else:
        adgroup_ids.extend([str(x).strip() for x in (entity_ids or []) if str(x).strip()])
    return _unique_keep_order(adgroup_ids), warnings
def _resolve_copy_source_adgroup_ids(api_key: str, secret_key: str, cid: str, source_ids: List[str], source_campaign_ids: List[str]):
    direct_ids = _unique_keep_order(source_ids or [])
    campaign_adgroup_ids, warnings = _resolve_adgroup_ids(api_key, secret_key, cid, "campaign", source_campaign_ids or [])
    resolved = _unique_keep_order(direct_ids + campaign_adgroup_ids)
    return resolved, warnings
def _resolve_web_site_adgroup_ids(api_key: str, secret_key: str, cid: str, entity_type: str, entity_ids: List[str]):
    adgroup_ids: List[str] = []
    skipped_non_web = 0
    warnings: List[str] = []
    if entity_type == "campaign":
        for camp_id in entity_ids:
            res, rows = _fetch_adgroups(api_key, secret_key, cid, str(camp_id).strip())
            if res.status_code != 200:
                warnings.append(f"캠페인 {camp_id} 하위 광고그룹 조회 실패: {res.text}")
                continue
            for row in rows:
                if str(row.get("adgroupType") or "").upper() == "WEB_SITE":
                    adgroup_ids.append(str(row.get("id") or "").strip())
                else:
                    skipped_non_web += 1
    else:
        for adg_id in entity_ids:
            res, obj = _fetch_adgroup_detail(api_key, secret_key, cid, str(adg_id).strip())
            if res.status_code != 200 or not obj:
                warnings.append(f"광고그룹 {adg_id} 조회 실패: {res.text}")
                continue
            if str(obj.get("adgroupType") or "").upper() == "WEB_SITE":
                adgroup_ids.append(str(obj.get("nccAdgroupId") or adg_id).strip())
            else:
                skipped_non_web += 1
    return _unique_keep_order(adgroup_ids), skipped_non_web, warnings
def _resolve_shopping_adgroup_ids(api_key: str, secret_key: str, cid: str, entity_type: str, entity_ids: List[str]):
    adgroup_ids: List[str] = []
    skipped_non_shopping = 0
    warnings: List[str] = []
    entity_type = str(entity_type or "adgroup").strip().lower()
    entity_ids = _unique_keep_order(entity_ids or [])
    if entity_type == "campaign":
        for camp_id in entity_ids:
            res, rows = _fetch_adgroups(api_key, secret_key, cid, str(camp_id).strip())
            if res.status_code != 200:
                warnings.append(f"캠페인 {camp_id} 하위 광고그룹 조회 실패: {res.text}")
                continue
            for row in rows:
                adgroup_type = str(row.get("adgroupType") or "").upper()
                if adgroup_type in SHOPPING_TARGETABLE_ADGROUP_TYPES:
                    adgroup_ids.append(str(row.get("id") or row.get("nccAdgroupId") or "").strip())
                else:
                    skipped_non_shopping += 1
    else:
        for adg_id in entity_ids:
            res, obj = _fetch_adgroup_detail(api_key, secret_key, cid, str(adg_id).strip())
            if res.status_code != 200 or not obj:
                warnings.append(f"광고그룹 {adg_id} 조회 실패: {res.text}")
                continue
            adgroup_type = str(obj.get("adgroupType") or "").upper()
            if adgroup_type in SHOPPING_TARGETABLE_ADGROUP_TYPES:
                adgroup_ids.append(str(obj.get("nccAdgroupId") or adg_id).strip())
            else:
                skipped_non_shopping += 1
    return _unique_keep_order(adgroup_ids), skipped_non_shopping, warnings
def _normalize_bid_amt(value: Any, max_bid: Optional[int] = None) -> Optional[int]:
    try:
        bid = int(float(value))
    except Exception:
        return None
    bid = int(round(bid / 10.0) * 10)
    if max_bid is not None:
        bid = min(bid, int(max_bid))
    bid = max(70, min(100000, bid))
    return bid
def _normalize_bid_weight(value: Any) -> Optional[int]:
    try:
        weight = int(float(value))
    except Exception:
        return None
    return max(10, min(500, weight))
def _calc_bid_stats(values: List[int]):
    if not values:
        return {"min": None, "max": None, "median": None}
    vals = sorted(values)
    n = len(vals)
    if n % 2:
        median = vals[n // 2]
    else:
        median = int(round((vals[n // 2 - 1] + vals[n // 2]) / 2.0 / 10.0) * 10)
    return {"min": vals[0], "max": vals[-1], "median": median}
def _is_keyword_editable_for_avg_position(row: Dict[str, Any], include_paused: bool = False, include_pending: bool = False):
    if bool(row.get("delFlag")):
        return False, "삭제됨"
    status = str(row.get("status") or "").upper()
    if (not include_paused) and status in {"PAUSE", "PAUSED", "STOP", "STOPPED", "LIMITEDBYBUDGET"}:
        return False, f"상태:{status or 'PAUSE'}"
    inspect = str(row.get("inspectStatus") or "").upper()
    if (not include_pending) and inspect and inspect not in {"APPROVED", "NONE", "NORMAL"}:
        return False, f"검수:{inspect}"
    return True, ""
def _is_ad_editable_for_avg_position(row: Dict[str, Any], include_paused: bool = False, include_pending: bool = False):
    if bool(row.get("delFlag")):
        return False, "삭제됨"
    status = str(row.get("status") or "").upper()
    if (not include_paused) and status in {"PAUSE", "PAUSED", "STOP", "STOPPED", "LIMITEDBYBUDGET"}:
        return False, f"상태:{status or 'PAUSE'}"
    inspect = str(row.get("inspectStatus") or "").upper()
    if (not include_pending) and inspect and inspect not in {"APPROVED", "NONE", "NORMAL"}:
        return False, f"검수:{inspect}"
    return True, ""
def _estimate_keyword_bids_by_avg_position(api_key: str, secret_key: str, cid: str, keyword_ids: List[str], device: str, position: int, max_bid: Optional[int] = None):
    estimated: Dict[str, int] = {}
    warnings: List[str] = []
    device = str(device or "PC").upper()
    for i in range(0, len(keyword_ids), 100):
        chunk = [str(x).strip() for x in keyword_ids[i:i + 100] if str(x).strip()]
        if not chunk:
            continue
        body = {
            "device": device,
            "items": [{"key": kid, "position": int(position)} for kid in chunk],
        }
        res = _do_req("POST", api_key, secret_key, cid, "/npc-estimate/average-position-bid/id", json_body=body)
        if res.status_code != 200:
            warnings.append(f"평균순위 추정 실패({device}, {position}위): {res.text}")
            continue
        payload = res.json() or {}
        items = None
        if isinstance(payload, dict):
            items = payload.get("items")
            if not isinstance(items, list):
                items = payload.get("estimate")
        elif isinstance(payload, list):
            items = payload
        if not isinstance(items, list):
            warnings.append(f"평균순위 추정 응답 형식이 예상과 다릅니다: {payload}")
            continue
        for item in items:
            try:
                key = str(item.get("key") or item.get("nccKeywordId") or "").strip()
                bid = _normalize_bid_amt(item.get("bid"), max_bid=max_bid)
            except Exception:
                bid = None
                key = ""
            if key and bid is not None:
                estimated[key] = bid
    return estimated, warnings
def _estimate_ad_bids_by_avg_position(api_key: str, secret_key: str, cid: str, ad_ids: List[str], device: str, position: int, max_bid: Optional[int] = None):
    estimated: Dict[str, int] = {}
    warnings: List[str] = []
    device = str(device or "PC").upper()
    for i in range(0, len(ad_ids), 100):
        chunk = [str(x).strip() for x in ad_ids[i:i + 100] if str(x).strip()]
        if not chunk:
            continue
        body = {
            "device": device,
            "items": [{"key": ad_id, "position": int(position)} for ad_id in chunk],
        }
        res = _do_req("POST", api_key, secret_key, cid, "/npla-estimate/average-position-bid/id", json_body=body)
        if res.status_code != 200:
            warnings.append(f"쇼핑 평균순위 추정 실패({device}, {position}위): {res.text}")
            continue
        payload = res.json() or {}
        items = None
        if isinstance(payload, dict):
            items = payload.get("items")
            if not isinstance(items, list):
                items = payload.get("estimate")
        elif isinstance(payload, list):
            items = payload
        if not isinstance(items, list):
            warnings.append(f"쇼핑 평균순위 추정 응답 형식이 예상과 다릅니다: {payload}")
            continue
        for item in items:
            try:
                key = str(item.get("key") or item.get("nccAdId") or item.get("id") or "").strip()
                bid = _normalize_bid_amt(item.get("bid") or item.get("bidAmt"), max_bid=max_bid)
            except Exception:
                bid = None
                key = ""
            if key and bid is not None:
                estimated[key] = bid
    return estimated, warnings
def _apply_keyword_bid_map(api_key: str, secret_key: str, cid: str, adgroup_ids: List[str], bid_map: Dict[str, int], keyword_meta: Optional[Dict[str, Dict[str, Any]]] = None):
    success_cnt = fail_cnt = skipped_cnt = 0
    err_details: List[str] = []
    cleanup_keys = ['regTm', 'editTm', 'status', 'statusReason', 'inspectStatus', 'delFlag', 'managedKeyword', 'referenceKey']
    keyword_meta = keyword_meta or {}
    for adg_id in adgroup_ids:
        r_kw = _do_req("GET", api_key, secret_key, cid, "/ncc/keywords", params={"nccAdgroupId": adg_id})
        if r_kw.status_code != 200:
            fail_cnt += 1
            if len(err_details) < 5:
                err_details.append(f"[광고그룹 {adg_id}] 키워드 조회 실패: {r_kw.text}")
            continue
        kws = r_kw.json() or []
        update_payload = []
        for kw in kws:
            kid = str(kw.get("nccKeywordId") or "").strip()
            if not kid:
                continue
            if kid not in bid_map:
                skipped_cnt += 1
                continue
            target_bid = int(bid_map[kid])
            meta = keyword_meta.get(kid, {})
            current_bid = _normalize_bid_amt(meta.get("current_bid", kw.get("bidAmt")))
            current_use_group = bool(meta.get("use_group_bid", kw.get("useGroupBidAmt")))
            if (current_bid == target_bid) and (not current_use_group):
                skipped_cnt += 1
                continue
            item = copy.deepcopy(kw)
            item["useGroupBidAmt"] = False
            item["bidAmt"] = target_bid
            for k in cleanup_keys:
                item.pop(k, None)
            update_payload.append(item)
        for i in range(0, len(update_payload), 100):
            batch = update_payload[i:i + 100]
            if not batch:
                continue
            r_put = _do_req("PUT", api_key, secret_key, cid, "/ncc/keywords", params={"fields": "bidAmt,useGroupBidAmt"}, json_body=batch)
            if r_put.status_code in [200, 201]:
                success_cnt += len(batch)
            else:
                for item in batch:
                    r_single = _do_req("PUT", api_key, secret_key, cid, f"/ncc/keywords/{item['nccKeywordId']}", params={"fields": "bidAmt,useGroupBidAmt"}, json_body=item)
                    if r_single.status_code in [200, 201]:
                        success_cnt += 1
                    else:
                        fail_cnt += 1
                        if len(err_details) < 5:
                            err_details.append(f"[{item.get('keyword', '알수없음')}] 변경 실패: {r_single.text}")
    return success_cnt, fail_cnt, skipped_cnt, err_details
@app.route("/update_non_search_keyword_exclusion", methods=["POST"])
def update_non_search_keyword_exclusion():
    d = request.get_json(force=True) or {}
    api_key = str(d.get("api_key") or "").strip()
    secret_key = str(d.get("secret_key") or "").strip()
    cid = str(d.get("customer_id") or "").strip()
    entity_type = str(d.get("entity_type") or "adgroup").strip().lower()
    entity_ids = _unique_keep_order(d.get("entity_ids") or [])
    if entity_type not in {"campaign", "adgroup"}:
        return jsonify({"error": "entity_type은 campaign 또는 adgroup 이어야 합니다."}), 400
    if not entity_ids:
        return jsonify({"error": "대상이 없습니다."}), 400
    if "excluded" not in d:
        return jsonify({"error": "사용 여부(excluded)는 필수입니다."}), 400
    excluded = _boolish(d.get("excluded"), False)
    adgroup_ids, skipped_non_shopping, warnings = _resolve_shopping_adgroup_ids(api_key, secret_key, cid, entity_type, entity_ids)
    if not adgroup_ids:
        msg = "선택 범위에서 적용 가능한 쇼핑 광고그룹을 찾지 못했습니다."
        if warnings:
            msg += f" / {warnings[0]}"
        return jsonify({"error": msg}), 400
    success = 0
    fail = 0
    details: List[Dict[str, Any]] = []
    for adg_id in adgroup_ids:
        ok, msg = _update_non_search_keyword_target(api_key, secret_key, cid, adg_id, excluded)
        details.append({"adgroup_id": adg_id, "success": bool(ok), "message": msg})
        if ok:
            success += 1
        else:
            fail += 1
    message_bits = [f"검색어 없는 경우 광고 노출 제외 {'사용' if excluded else '사용 안함'} 적용 완료: 성공 {success}건 / 실패 {fail}건"]
    if skipped_non_shopping > 0:
        message_bits.append(f"비쇼핑 광고그룹 {skipped_non_shopping}건은 건너뜀")
    if warnings:
        message_bits.append(f"조회 경고 {len(warnings)}건")
    return jsonify({
        "ok": fail == 0,
        "success": success,
        "fail": fail,
        "skipped_non_shopping": skipped_non_shopping,
        "warnings": warnings,
        "results": details,
        "message": " / ".join(message_bits),
    }), (200 if fail == 0 else 207)
@app.route("/preview_keyword_bids_by_search", methods=["POST"])
def preview_keyword_bids_by_search():
    d = request.json or {}
    api_key = str(d.get("api_key") or "").strip()
    secret_key = str(d.get("secret_key") or "").strip()
    cid = str(d.get("customer_id") or "").strip()
    search_text = str(d.get("search_text") or d.get("keyword_query") or "").strip()
    exact_match = _boolish(d.get("exact_match"), False)
    search_scope = str(d.get("search_scope") or "account").strip().lower()
    campaign_ids = [str(x or "").strip() for x in (d.get("campaign_ids") or []) if str(x or "").strip()]
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API Key / Secret Key / Customer ID가 필요합니다."}), 400
    if not search_text:
        return jsonify({"error": "검색할 단어를 입력해주세요."}), 400
    scan = _scan_powerlink_keywords_by_search(api_key, secret_key, cid, search_text, exact_match=exact_match)
    powerlink_campaigns = scan["powerlink_campaigns"]
    adgroup_contexts = scan["adgroup_contexts"]
    matched_rows = scan["matched_rows"]
    err_details = scan["warnings"]
    scanned_keyword_count = int(scan["scanned_keyword_count"])
    matched_count = int(scan["matched_count"])
    preview_token = hashlib.sha256(f"{cid}|{search_text}|{'1' if exact_match else '0'}|{matched_count}".encode('utf-8')).hexdigest()[:20]
    return jsonify({
        "ok": True,
        "message": _build_powerlink_keyword_search_message(search_text, exact_match, scan, row_preview_limit=30),
        "search_text": search_text,
        "search_groups": scan.get("search_groups") or _parse_keyword_search_groups(search_text),
        "exact_match": bool(exact_match),
        "matched_count": matched_count,
        "scanned_keyword_count": scanned_keyword_count,
        "total_powerlink_campaign_count": len(powerlink_campaigns),
        "total_powerlink_adgroup_count": len(adgroup_contexts),
        "rows": matched_rows[:100],
        "warnings": err_details[:10],
        "preview_token": preview_token,
    })
@app.route("/update_keyword_bids_by_search", methods=["POST"])
def update_keyword_bids_by_search():
    d = request.json or {}
    api_key = str(d.get("api_key") or "").strip()
    secret_key = str(d.get("secret_key") or "").strip()
    cid = str(d.get("customer_id") or "").strip()
    search_text = str(d.get("search_text") or d.get("keyword_query") or "").strip()
    exact_match = _boolish(d.get("exact_match"), False)
    search_scope = str(d.get("search_scope") or "account").strip().lower()
    campaign_ids = [str(x or "").strip() for x in (d.get("campaign_ids") or []) if str(x or "").strip()]
    target_bid = _normalize_bid_amt(d.get("bid_amt"))
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API Key / Secret Key / Customer ID가 필요합니다."}), 400
    if not search_text:
        return jsonify({"error": "검색할 단어를 입력해주세요."}), 400
    if target_bid is None:
        return jsonify({"error": "적용할 입찰가를 올바르게 입력해주세요."}), 400
    scan = _scan_powerlink_keywords_by_search(api_key, secret_key, cid, search_text, exact_match=exact_match)
    powerlink_campaigns = scan["powerlink_campaigns"]
    adgroup_contexts = scan["adgroup_contexts"]
    matched_rows = scan["matched_rows"]
    base_payload = scan["update_payload"]
    updated_adgroup_ids = scan["updated_adgroup_ids"]
    err_details = list(scan["warnings"])
    scanned_keyword_count = int(scan["scanned_keyword_count"])
    matched_count = int(scan["matched_count"])
    if not powerlink_campaigns:
        return jsonify({"ok": True, "message": "현재 계정에 파워링크 캠페인이 없습니다.", "matched_count": 0, "updated_count": 0, "skipped_count": 0, "fail_count": 0, "rows": [], "updated_adgroup_ids": []})
    if not adgroup_contexts:
        return jsonify({"ok": True, "message": "현재 계정에서 조회 가능한 파워링크 광고그룹이 없습니다.", "matched_count": 0, "updated_count": 0, "skipped_count": 0, "fail_count": 0, "rows": [], "warnings": err_details[:10], "updated_adgroup_ids": []})
    update_payload: List[Dict[str, Any]] = []
    skipped_count = 0
    bid_by_id = {str(row.get("ncc_keyword_id") or "").strip(): row for row in matched_rows}
    for item in (base_payload or []):
        item2 = copy.deepcopy(item)
        item2["useGroupBidAmt"] = False
        item2["bidAmt"] = int(target_bid)
        item_id = str(item2.get("nccKeywordId") or "").strip()
        row = bid_by_id.get(item_id) or {}
        current_bid = _normalize_bid_amt(row.get("current_bid")) or _normalize_bid_amt(item2.get("bidAmt")) or 70
        current_use_group = bool(row.get("current_use_group"))
        if current_bid == target_bid and not current_use_group:
            skipped_count += 1
            continue
        update_payload.append(item2)
    updated_count = 0
    fail_count = 0
    single_fallback_jobs: List[Dict[str, Any]] = []
    for i in range(0, len(update_payload), 100):
        batch = update_payload[i:i + 100]
        if not batch:
            continue
        r_put = _do_req("PUT", api_key, secret_key, cid, "/ncc/keywords", params={"fields": "bidAmt,useGroupBidAmt"}, json_body=batch)
        if r_put.status_code in [200, 201]:
            updated_count += len(batch)
        else:
            single_fallback_jobs.extend(batch)
    if single_fallback_jobs:
        fb_workers = min(max(BID_IO_WORKERS, 10), max(1, len(single_fallback_jobs)))
        with ThreadPoolExecutor(max_workers=fb_workers) as fb_ex:
            fb_future_map = {
                fb_ex.submit(_do_req, "PUT", api_key, secret_key, cid, f"/ncc/keywords/{item['nccKeywordId']}", {"fields": "bidAmt,useGroupBidAmt"}, item): item
                for item in single_fallback_jobs
            }
            for fb_fut in as_completed(fb_future_map):
                item = fb_future_map[fb_fut]
                try:
                    r_single = fb_fut.result()
                except Exception as exc:
                    fail_count += 1
                    if len(err_details) < 10:
                        err_details.append(f"[{item.get('keyword', '알수없음')}] 키워드 입찰가 변경 실패: {exc}")
                    continue
                if r_single.status_code in [200, 201]:
                    updated_count += 1
                else:
                    fail_count += 1
                    if len(err_details) < 10:
                        err_details.append(f"[{item.get('keyword', '알수없음')}] 키워드 입찰가 변경 실패: {r_single.text}")
    mode_label = "완전일치" if exact_match else "부분일치"
    lines = [
        f"검색어 기준 파워링크 키워드 입찰가 변경 완료! ({mode_label})",
        f"검색어: {search_text}",
        f"대상 입찰가: {int(target_bid):,}원",
        f"조회한 파워링크 캠페인: {len(powerlink_campaigns)}개 | 광고그룹: {len(adgroup_contexts)}개 | 키워드 스캔: {scanned_keyword_count}개",
        f"검색 일치 키워드: {matched_count}개 | 변경 성공: {updated_count}개 | 유지: {skipped_count}개 | 실패: {fail_count}개",
    ]
    if matched_rows:
        lines.append("\n[일치 키워드 예시]")
        for row in matched_rows[:20]:
            current_label = f"{int(row.get('current_bid') or 0):,}원"
            if row.get("current_use_group"):
                current_label += " (그룹입찰가 사용)"
            matched_term_label = str(row.get('matched_terms_text') or '')
            if matched_term_label:
                matched_term_label = f" | 매칭: {matched_term_label}"
            lines.append(f"- {row.get('campaign_name')} > {row.get('adgroup_name')} > {row.get('keyword')}{matched_term_label} | 현재 {current_label}")
    if err_details:
        lines.append("\n[상세 내역]")
        lines.extend(err_details[:10])
    _cache_invalidate(api_key, secret_key, cid)
    return jsonify({
        "ok": True,
        "message": "\n".join(lines),
        "search_text": search_text,
        "exact_match": bool(exact_match),
        "target_bid": int(target_bid),
        "matched_count": int(matched_count),
        "updated_count": int(updated_count),
        "skipped_count": int(skipped_count),
        "fail_count": int(fail_count),
        "scanned_keyword_count": int(scanned_keyword_count),
        "total_powerlink_campaign_count": len(powerlink_campaigns),
        "total_powerlink_adgroup_count": len(adgroup_contexts),
        "rows": matched_rows[:100],
        "warnings": err_details[:10],
        "updated_adgroup_ids": _unique_keep_order(updated_adgroup_ids),
    })
@app.route("/preview_keyword_bid_weights_by_search", methods=["POST"])
def preview_keyword_bid_weights_by_search():
    d = request.json or {}
    api_key = str(d.get("api_key") or "").strip()
    secret_key = str(d.get("secret_key") or "").strip()
    cid = str(d.get("customer_id") or "").strip()
    search_text = str(d.get("search_text") or d.get("keyword_query") or "").strip()
    exact_match = _boolish(d.get("exact_match"), False)
    search_scope = str(d.get("search_scope") or "account").strip().lower()
    campaign_ids = [str(x or "").strip() for x in (d.get("campaign_ids") or []) if str(x or "").strip()]
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API Key / Secret Key / Customer ID가 필요합니다."}), 400
    if not search_text:
        return jsonify({"error": "검색할 단어를 입력해주세요."}), 400
    scan = _scan_powerlink_keywords_by_search(api_key, secret_key, cid, search_text, exact_match=exact_match)
    powerlink_campaigns = scan["powerlink_campaigns"]
    adgroup_contexts = scan["adgroup_contexts"]
    matched_rows = scan["matched_rows"]
    err_details = scan["warnings"]
    scanned_keyword_count = int(scan["scanned_keyword_count"])
    matched_count = int(scan["matched_count"])
    mode_label = "완전일치" if exact_match else "부분일치"
    lines = [
        f"검색어 기준 파워링크 키워드 입찰가중치 조회 완료! ({mode_label})",
        f"검색어: {search_text}",
        f"조회한 파워링크 캠페인: {len(powerlink_campaigns)}개 | 광고그룹: {len(adgroup_contexts)}개 | 키워드 스캔: {scanned_keyword_count}개",
        f"검색 일치 키워드: {matched_count}개",
    ]
    if matched_rows:
        lines.append("\n[일치 키워드 예시]")
        for row in matched_rows[:30]:
            current_label = f"{int(row.get('current_bid_weight') or 100):,}%"
            matched_term_label = str(row.get('matched_terms_text') or '')
            if matched_term_label:
                matched_term_label = f" | 매칭: {matched_term_label}"
            lines.append(f"- {row.get('campaign_name')} > {row.get('adgroup_name')} > {row.get('keyword')}{matched_term_label} | 현재 입찰가중치 {current_label}")
    if err_details:
        lines.append("\n[상세 내역]")
        lines.extend(err_details[:10])
    preview_token = hashlib.sha256(f"BW|{cid}|{search_text}|{'1' if exact_match else '0'}|{matched_count}".encode('utf-8')).hexdigest()[:20]
    return jsonify({
        "ok": True,
        "message": "\n".join(lines),
        "search_text": search_text,
        "exact_match": bool(exact_match),
        "matched_count": matched_count,
        "scanned_keyword_count": scanned_keyword_count,
        "total_powerlink_campaign_count": len(powerlink_campaigns),
        "total_powerlink_adgroup_count": len(adgroup_contexts),
        "rows": matched_rows[:100],
        "warnings": err_details[:10],
        "preview_token": preview_token,
    })
@app.route("/update_keyword_bid_weights_by_search", methods=["POST"])
def update_keyword_bid_weights_by_search():
    d = request.json or {}
    api_key = str(d.get("api_key") or "").strip()
    secret_key = str(d.get("secret_key") or "").strip()
    cid = str(d.get("customer_id") or "").strip()
    search_text = str(d.get("search_text") or d.get("keyword_query") or "").strip()
    exact_match = _boolish(d.get("exact_match"), False)
    search_scope = str(d.get("search_scope") or "account").strip().lower()
    campaign_ids = [str(x or "").strip() for x in (d.get("campaign_ids") or []) if str(x or "").strip()]
    target_weight = _normalize_bid_weight(d.get("bid_weight"))
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API Key / Secret Key / Customer ID가 필요합니다."}), 400
    if not search_text:
        return jsonify({"error": "검색할 단어를 입력해주세요."}), 400
    if target_weight is None:
        return jsonify({"error": "적용할 입찰가중치(%)를 올바르게 입력해주세요. (10~500)"}), 400
    scan = _scan_powerlink_keywords_by_search(api_key, secret_key, cid, search_text, exact_match=exact_match)
    powerlink_campaigns = scan["powerlink_campaigns"]
    adgroup_contexts = scan["adgroup_contexts"]
    matched_rows = scan["matched_rows"]
    base_payload = scan["update_payload"]
    updated_adgroup_ids = scan["updated_adgroup_ids"]
    err_details = list(scan["warnings"])
    scanned_keyword_count = int(scan["scanned_keyword_count"])
    matched_count = int(scan["matched_count"])
    if not powerlink_campaigns:
        return jsonify({"ok": True, "message": "현재 계정에 파워링크 캠페인이 없습니다.", "matched_count": 0, "updated_count": 0, "skipped_count": 0, "fail_count": 0, "rows": [], "updated_adgroup_ids": []})
    if not adgroup_contexts:
        return jsonify({"ok": True, "message": "현재 계정에서 조회 가능한 파워링크 광고그룹이 없습니다.", "matched_count": 0, "updated_count": 0, "skipped_count": 0, "fail_count": 0, "rows": [], "warnings": err_details[:10], "updated_adgroup_ids": []})
    update_payload: List[Dict[str, Any]] = []
    skipped_count = 0
    row_by_id = {str(row.get("ncc_keyword_id") or "").strip(): row for row in matched_rows}
    for item in (base_payload or []):
        item2 = copy.deepcopy(item)
        item2["bidWeight"] = int(target_weight)
        item_id = str(item2.get("nccKeywordId") or "").strip()
        row = row_by_id.get(item_id) or {}
        current_weight = _normalize_bid_weight(row.get("current_bid_weight")) or 100
        if current_weight == target_weight:
            skipped_count += 1
            continue
        update_payload.append(item2)
    updated_count = 0
    fail_count = 0
    single_fallback_jobs: List[Dict[str, Any]] = []
    for i in range(0, len(update_payload), 100):
        batch = update_payload[i:i + 100]
        if not batch:
            continue
        r_put = _do_req("PUT", api_key, secret_key, cid, "/ncc/keywords", params={"fields": "bidWeight"}, json_body=batch)
        if r_put.status_code in [200, 201]:
            updated_count += len(batch)
        else:
            single_fallback_jobs.extend(batch)
    if single_fallback_jobs:
        fb_workers = min(max(BID_IO_WORKERS, 10), max(1, len(single_fallback_jobs)))
        with ThreadPoolExecutor(max_workers=fb_workers) as fb_ex:
            fb_future_map = {
                fb_ex.submit(_do_req, "PUT", api_key, secret_key, cid, f"/ncc/keywords/{item['nccKeywordId']}", {"fields": "bidWeight"}, item): item
                for item in single_fallback_jobs
            }
            for fb_fut in as_completed(fb_future_map):
                item = fb_future_map[fb_fut]
                try:
                    r_single = fb_fut.result()
                except Exception as exc:
                    fail_count += 1
                    if len(err_details) < 10:
                        err_details.append(f"[{item.get('keyword', '알수없음')}] 키워드 입찰가중치 변경 실패: {exc}")
                    continue
                if r_single.status_code in [200, 201]:
                    updated_count += 1
                else:
                    fail_count += 1
                    if len(err_details) < 10:
                        err_details.append(f"[{item.get('keyword', '알수없음')}] 키워드 입찰가중치 변경 실패: {r_single.text}")
    mode_label = "완전일치" if exact_match else "부분일치"
    lines = [
        f"검색어 기준 파워링크 키워드 입찰가중치 변경 완료! ({mode_label})",
        f"검색어: {search_text}",
        f"대상 입찰가중치: {int(target_weight):,}%",
        f"조회한 파워링크 캠페인: {len(powerlink_campaigns)}개 | 광고그룹: {len(adgroup_contexts)}개 | 키워드 스캔: {scanned_keyword_count}개",
        f"검색 일치 키워드: {matched_count}개 | 변경 성공: {updated_count}개 | 유지: {skipped_count}개 | 실패: {fail_count}개",
    ]
    if matched_rows:
        lines.append("\n[일치 키워드 예시]")
        for row in matched_rows[:20]:
            current_label = f"{int(row.get('current_bid_weight') or 100):,}%"
            matched_term_label = str(row.get('matched_terms_text') or '')
            if matched_term_label:
                matched_term_label = f" | 매칭: {matched_term_label}"
            lines.append(f"- {row.get('campaign_name')} > {row.get('adgroup_name')} > {row.get('keyword')}{matched_term_label} | 현재 입찰가중치 {current_label}")
    if err_details:
        lines.append("\n[상세 내역]")
        lines.extend(err_details[:10])
    _cache_invalidate(api_key, secret_key, cid)
    return jsonify({
        "ok": True,
        "message": "\n".join(lines),
        "search_text": search_text,
        "exact_match": bool(exact_match),
        "target_weight": int(target_weight),
        "matched_count": int(matched_count),
        "updated_count": int(updated_count),
        "skipped_count": int(skipped_count),
        "fail_count": int(fail_count),
        "scanned_keyword_count": int(scanned_keyword_count),
        "total_powerlink_campaign_count": len(powerlink_campaigns),
        "total_powerlink_adgroup_count": len(adgroup_contexts),
        "rows": matched_rows[:100],
        "warnings": err_details[:10],
        "updated_adgroup_ids": _unique_keep_order(updated_adgroup_ids),
    })
@app.route("/update_keyword_bids", methods=["POST"])
def update_keyword_bids():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    entity_type, entity_ids = d.get("entity_type"), d.get("entity_ids", [])
    bid_amt = int(d.get("bid_amt", 70))
    if not entity_ids:
        return jsonify({"error": "대상이 없습니다."}), 400
    adgroup_contexts, resolve_warnings = _resolve_adgroup_contexts(api_key, secret_key, cid, entity_type, entity_ids)
    if not adgroup_contexts:
        return jsonify({"error": "대상 광고그룹을 찾지 못했습니다."}), 400
    use_group_bid = bid_amt == 0
    target_bid = _normalize_bid_amt(bid_amt if bid_amt else 70) or 70
    kw_success = kw_fail = kw_skipped = 0
    ad_success = ad_fail = ad_skipped = 0
    err_details: List[str] = list(resolve_warnings)
    keyword_groups = 0
    shopping_groups = 0
    cleanup_keys = ['regTm', 'editTm', 'status', 'statusReason', 'inspectStatus', 'delFlag', 'managedKeyword', 'referenceKey']
    keyword_contexts = [ctx for ctx in adgroup_contexts if not _adgroup_uses_ad_level_bid(ctx.get("adgroup_type"))]
    shopping_contexts = [ctx for ctx in adgroup_contexts if _adgroup_uses_ad_level_bid(ctx.get("adgroup_type"))]
    keyword_groups = len(keyword_contexts)
    shopping_groups = len(shopping_contexts)
    if keyword_contexts:
        max_workers = min(BID_IO_WORKERS, max(1, len(keyword_contexts)))
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            future_map = {
                ex.submit(_do_req, "GET", api_key, secret_key, cid, "/ncc/keywords", {"nccAdgroupId": str(ctx.get("adgroup_id") or "").strip()}, None): ctx
                for ctx in keyword_contexts if str(ctx.get("adgroup_id") or "").strip()
            }
            single_fallback_jobs: List[Dict[str, Any]] = []
            for fut in as_completed(future_map):
                ctx = future_map[fut]
                adg_id = str(ctx.get("adgroup_id") or "").strip()
                adgroup_bid = ctx.get("adgroup_bid")
                try:
                    r_kw = fut.result()
                except Exception as exc:
                    kw_fail += 1
                    if len(err_details) < 10:
                        err_details.append(f"[광고그룹 {adg_id}] 키워드 조회 실패: {exc}")
                    continue
                if r_kw.status_code != 200:
                    kw_fail += 1
                    if len(err_details) < 10:
                        err_details.append(f"[광고그룹 {adg_id}] 키워드 조회 실패: {r_kw.text}")
                    continue
                kws = r_kw.json() or []
                update_payload = []
                for kw in kws:
                    item = copy.deepcopy(kw)
                    current_use_group = bool(kw.get("useGroupBidAmt"))
                    effective_current_bid = _resolve_effective_bid(kw.get("bidAmt"), current_use_group, adgroup_bid)
                    raw_keyword_bid = _normalize_bid_amt(kw.get("bidAmt")) or 70
                    should_skip = current_use_group if use_group_bid else ((effective_current_bid == target_bid) and (not current_use_group))
                    if should_skip:
                        kw_skipped += 1
                        continue
                    item["useGroupBidAmt"] = bool(use_group_bid)
                    item["bidAmt"] = raw_keyword_bid if use_group_bid else target_bid
                    for k in cleanup_keys:
                        item.pop(k, None)
                    update_payload.append(item)
                for i in range(0, len(update_payload), 100):
                    batch = update_payload[i:i + 100]
                    if not batch:
                        continue
                    r_put = _do_req("PUT", api_key, secret_key, cid, "/ncc/keywords", params={"fields": "bidAmt,useGroupBidAmt"}, json_body=batch)
                    if r_put.status_code in [200, 201]:
                        kw_success += len(batch)
                    else:
                        for item in batch:
                            single_fallback_jobs.append(item)
            if single_fallback_jobs:
                fallback_workers = min(BID_IO_WORKERS, max(1, len(single_fallback_jobs)))
                with ThreadPoolExecutor(max_workers=fallback_workers) as fb_ex:
                    fb_future_map = {
                        fb_ex.submit(_do_req, "PUT", api_key, secret_key, cid, f"/ncc/keywords/{item['nccKeywordId']}", {"fields": "bidAmt,useGroupBidAmt"}, item): item
                        for item in single_fallback_jobs
                    }
                    for fb_fut in as_completed(fb_future_map):
                        item = fb_future_map[fb_fut]
                        try:
                            r_single = fb_fut.result()
                        except Exception as exc:
                            kw_fail += 1
                            if len(err_details) < 10:
                                err_details.append(f"[{item.get('keyword', '알수없음')}] 키워드 입찰가 변경 실패: {exc}")
                            continue
                        if r_single.status_code in [200, 201]:
                            kw_success += 1
                        else:
                            kw_fail += 1
                            if len(err_details) < 10:
                                err_details.append(f"[{item.get('keyword', '알수없음')}] 키워드 입찰가 변경 실패: {r_single.text}")
    if shopping_contexts:
        max_workers = min(BID_IO_WORKERS, max(1, len(shopping_contexts)))
        put_jobs: List[Dict[str, Any]] = []
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            future_map = {
                ex.submit(_fetch_ads, api_key, secret_key, cid, str(ctx.get("adgroup_id") or "").strip()): ctx
                for ctx in shopping_contexts if str(ctx.get("adgroup_id") or "").strip()
            }
            for fut in as_completed(future_map):
                ctx = future_map[fut]
                adg_id = str(ctx.get("adgroup_id") or "").strip()
                adgroup_bid = ctx.get("adgroup_bid")
                try:
                    res_ads, ads = fut.result()
                except Exception as exc:
                    ad_fail += 1
                    if len(err_details) < 10:
                        err_details.append(f"[광고그룹 {adg_id}] 소재 조회 실패: {exc}")
                    continue
                if res_ads.status_code != 200:
                    ad_fail += 1
                    if len(err_details) < 10:
                        err_details.append(f"[광고그룹 {adg_id}] 소재 조회 실패: {res_ads.text}")
                    continue
                for ad in (ads or []):
                    if not _ad_item_has_bid_attr(ad):
                        continue
                    ad_attr = _extract_ad_attr(ad)
                    current_use_group = bool(ad_attr.get("useGroupBidAmt"))
                    effective_current_bid = _resolve_effective_bid(ad_attr.get("bidAmt"), current_use_group, adgroup_bid)
                    raw_item_bid = _normalize_bid_amt(ad_attr.get("bidAmt")) or 70
                    should_skip = current_use_group if use_group_bid else ((effective_current_bid == target_bid) and (not current_use_group))
                    if should_skip:
                        ad_skipped += 1
                        continue
                    item = copy.deepcopy(ad)
                    new_attr = _extract_ad_attr(item)
                    new_attr["useGroupBidAmt"] = bool(use_group_bid)
                    new_attr["bidAmt"] = raw_item_bid if use_group_bid else target_bid
                    item["adAttr"] = new_attr
                    put_jobs.append(item)
        if put_jobs:
            put_workers = min(BID_IO_WORKERS, max(1, len(put_jobs)))
            with ThreadPoolExecutor(max_workers=put_workers) as put_ex:
                put_future_map = {put_ex.submit(_put_single_ad_with_ad_attr, api_key, secret_key, cid, item): item for item in put_jobs}
                for put_fut in as_completed(put_future_map):
                    item = put_future_map[put_fut]
                    try:
                        r_put = put_fut.result()
                    except Exception as exc:
                        ad_fail += 1
                        if len(err_details) < 10:
                            name = str((((item.get("ad") or {}).get("productName")) if isinstance(item.get("ad"), dict) else "") or _extract_ad_id(item) or "알수없음")
                            err_details.append(f"[{name}] 소재 입찰가 변경 실패: {exc}")
                        continue
                    if r_put.status_code in [200, 201]:
                        ad_success += 1
                    else:
                        ad_fail += 1
                        if len(err_details) < 10:
                            name = str((((item.get("ad") or {}).get("productName")) if isinstance(item.get("ad"), dict) else "") or _extract_ad_id(item) or "알수없음")
                            err_details.append(f"[{name}] 소재 입찰가 변경 실패: {r_put.text}")
    label = "그룹 입찰가 사용 전환" if use_group_bid else f"고정 입찰가 {target_bid:,}원 적용"
    lines = [
        f"입찰가 변경 완료! ({label})",
        f"파워링크/키워드 처리 광고그룹: {keyword_groups}개 | 쇼핑/소재 처리 광고그룹: {shopping_groups}개",
        f"키워드 변경 성공: {kw_success}개 / 실패: {kw_fail}개 / 유지: {kw_skipped}개",
        f"쇼핑 소재 변경 성공: {ad_success}개 / 실패: {ad_fail}개 / 유지: {ad_skipped}개",
    ]
    if err_details:
        lines.append("\n[상세 내역]")
        lines.extend(err_details[:10])
    return jsonify({"ok": True, "message": "\n".join(lines)})
@app.route("/update_bid_mode_by_scope", methods=["POST"])
def update_bid_mode_by_scope():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    entity_type = str(d.get("entity_type") or "adgroup").strip().lower()
    entity_ids = d.get("entity_ids") or []
    bid_mode = str(d.get("bid_mode") or "").strip().lower()
    if bid_mode not in {"group", "individual"}:
        return jsonify({"error": "입찰가 설정 방식을 확인해주세요."}), 400
    if not entity_ids:
        return jsonify({"error": "대상이 없습니다."}), 400
    adgroup_contexts, resolve_warnings = _resolve_adgroup_contexts(api_key, secret_key, cid, entity_type, entity_ids)
    if not adgroup_contexts:
        return jsonify({"error": "대상 광고그룹을 찾지 못했습니다."}), 400

    target_use_group = bid_mode == "group"
    keyword_contexts = [ctx for ctx in adgroup_contexts if not _adgroup_uses_ad_level_bid(ctx.get("adgroup_type"))]
    shopping_contexts = [ctx for ctx in adgroup_contexts if _adgroup_uses_ad_level_bid(ctx.get("adgroup_type"))]
    keyword_groups = len(keyword_contexts)
    shopping_groups = len(shopping_contexts)
    kw_success = kw_fail = kw_skipped = 0
    ad_success = ad_fail = ad_skipped = 0
    err_details: List[str] = list(resolve_warnings)
    cleanup_keys = ['regTm', 'editTm', 'status', 'statusReason', 'inspectStatus', 'delFlag', 'managedKeyword', 'referenceKey']

    if keyword_contexts:
        max_workers = min(BID_IO_WORKERS, max(1, len(keyword_contexts)))
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            future_map = {
                ex.submit(_do_req, "GET", api_key, secret_key, cid, "/ncc/keywords", {"nccAdgroupId": str(ctx.get("adgroup_id") or "").strip()}, None): ctx
                for ctx in keyword_contexts if str(ctx.get("adgroup_id") or "").strip()
            }
            single_fallback_jobs: List[Dict[str, Any]] = []
            for fut in as_completed(future_map):
                ctx = future_map[fut]
                adg_id = str(ctx.get("adgroup_id") or "").strip()
                adgroup_bid = ctx.get("adgroup_bid")
                try:
                    r_kw = fut.result()
                except Exception as exc:
                    kw_fail += 1
                    if len(err_details) < 10:
                        err_details.append(f"[광고그룹 {adg_id}] 키워드 조회 실패: {exc}")
                    continue
                if r_kw.status_code != 200:
                    kw_fail += 1
                    if len(err_details) < 10:
                        err_details.append(f"[광고그룹 {adg_id}] 키워드 조회 실패: {r_kw.text}")
                    continue
                kws = r_kw.json() or []
                update_payload = []
                for kw in kws:
                    item = copy.deepcopy(kw)
                    current_use_group = bool(kw.get("useGroupBidAmt"))
                    raw_keyword_bid = _normalize_bid_amt(kw.get("bidAmt")) or 70
                    effective_current_bid = _resolve_effective_bid(kw.get("bidAmt"), current_use_group, adgroup_bid) or raw_keyword_bid
                    target_bid = raw_keyword_bid if target_use_group else effective_current_bid
                    should_skip = current_use_group == target_use_group
                    if should_skip:
                        kw_skipped += 1
                        continue
                    item["useGroupBidAmt"] = bool(target_use_group)
                    item["bidAmt"] = target_bid
                    for k in cleanup_keys:
                        item.pop(k, None)
                    update_payload.append(item)
                for i in range(0, len(update_payload), 100):
                    batch = update_payload[i:i + 100]
                    if not batch:
                        continue
                    r_put = _do_req("PUT", api_key, secret_key, cid, "/ncc/keywords", params={"fields": "bidAmt,useGroupBidAmt"}, json_body=batch)
                    if r_put.status_code in [200, 201]:
                        kw_success += len(batch)
                    else:
                        single_fallback_jobs.extend(batch)
            if single_fallback_jobs:
                fallback_workers = min(BID_IO_WORKERS, max(1, len(single_fallback_jobs)))
                with ThreadPoolExecutor(max_workers=fallback_workers) as fb_ex:
                    fb_future_map = {
                        fb_ex.submit(_do_req, "PUT", api_key, secret_key, cid, f"/ncc/keywords/{item['nccKeywordId']}", {"fields": "bidAmt,useGroupBidAmt"}, item): item
                        for item in single_fallback_jobs
                    }
                    for fb_fut in as_completed(fb_future_map):
                        item = fb_future_map[fb_fut]
                        try:
                            r_single = fb_fut.result()
                        except Exception as exc:
                            kw_fail += 1
                            if len(err_details) < 10:
                                err_details.append(f"[{item.get('keyword', '알수없음')}] 입찰가 설정 방식 변경 실패: {exc}")
                            continue
                        if r_single.status_code in [200, 201]:
                            kw_success += 1
                        else:
                            kw_fail += 1
                            if len(err_details) < 10:
                                err_details.append(f"[{item.get('keyword', '알수없음')}] 입찰가 설정 방식 변경 실패: {r_single.text}")

    if shopping_contexts:
        max_workers = min(BID_IO_WORKERS, max(1, len(shopping_contexts)))
        put_jobs: List[Dict[str, Any]] = []
        with ThreadPoolExecutor(max_workers=max_workers) as ex:
            future_map = {
                ex.submit(_fetch_ads, api_key, secret_key, cid, str(ctx.get("adgroup_id") or "").strip()): ctx
                for ctx in shopping_contexts if str(ctx.get("adgroup_id") or "").strip()
            }
            for fut in as_completed(future_map):
                ctx = future_map[fut]
                adg_id = str(ctx.get("adgroup_id") or "").strip()
                adgroup_bid = ctx.get("adgroup_bid")
                try:
                    res_ads, ads = fut.result()
                except Exception as exc:
                    ad_fail += 1
                    if len(err_details) < 10:
                        err_details.append(f"[광고그룹 {adg_id}] 소재 조회 실패: {exc}")
                    continue
                if res_ads.status_code != 200:
                    ad_fail += 1
                    if len(err_details) < 10:
                        err_details.append(f"[광고그룹 {adg_id}] 소재 조회 실패: {res_ads.text}")
                    continue
                for ad in (ads or []):
                    if not _ad_item_has_bid_attr(ad):
                        continue
                    ad_attr = _extract_ad_attr(ad)
                    current_use_group = bool(ad_attr.get("useGroupBidAmt"))
                    raw_item_bid = _normalize_bid_amt(ad_attr.get("bidAmt")) or 70
                    effective_current_bid = _resolve_effective_bid(ad_attr.get("bidAmt"), current_use_group, adgroup_bid) or raw_item_bid
                    target_bid = raw_item_bid if target_use_group else effective_current_bid
                    if current_use_group == target_use_group:
                        ad_skipped += 1
                        continue
                    item = copy.deepcopy(ad)
                    new_attr = _extract_ad_attr(item)
                    new_attr["useGroupBidAmt"] = bool(target_use_group)
                    new_attr["bidAmt"] = target_bid
                    item["adAttr"] = new_attr
                    put_jobs.append(item)
        if put_jobs:
            put_workers = min(BID_IO_WORKERS, max(1, len(put_jobs)))
            with ThreadPoolExecutor(max_workers=put_workers) as put_ex:
                put_future_map = {put_ex.submit(_put_single_ad_with_ad_attr, api_key, secret_key, cid, item): item for item in put_jobs}
                for put_fut in as_completed(put_future_map):
                    item = put_future_map[put_fut]
                    try:
                        r_put = put_fut.result()
                    except Exception as exc:
                        ad_fail += 1
                        if len(err_details) < 10:
                            name = str((((item.get("ad") or {}).get("productName")) if isinstance(item.get("ad"), dict) else "") or _extract_ad_id(item) or "알수없음")
                            err_details.append(f"[{name}] 입찰가 설정 방식 변경 실패: {exc}")
                        continue
                    if r_put.status_code in [200, 201]:
                        ad_success += 1
                    else:
                        ad_fail += 1
                        if len(err_details) < 10:
                            name = str((((item.get("ad") or {}).get("productName")) if isinstance(item.get("ad"), dict) else "") or _extract_ad_id(item) or "알수없음")
                            err_details.append(f"[{name}] 입찰가 설정 방식 변경 실패: {r_put.text}")

    label = "그룹 입찰가 사용" if target_use_group else "개별 입찰가 사용"
    lines = [
        f"입찰가 설정 방식 변경 완료! ({label})",
        f"파워링크/키워드 처리 광고그룹: {keyword_groups}개 | 쇼핑/소재 처리 광고그룹: {shopping_groups}개",
        f"키워드 변경 성공: {kw_success}개 / 실패: {kw_fail}개 / 유지: {kw_skipped}개",
        f"쇼핑 소재 변경 성공: {ad_success}개 / 실패: {ad_fail}개 / 유지: {ad_skipped}개",
    ]
    if not target_use_group:
        lines.append("개별 입찰가 전환 시 현재 적용 중인 유효 입찰가를 각 키워드/소재의 개별 입찰가로 저장합니다.")
    if err_details:
        lines.append("\n[상세 내역]")
        lines.extend(err_details[:10])
    return jsonify({"ok": True, "message": "\n".join(lines)})

@app.route("/adjust_keyword_bids_by_threshold", methods=["POST"])
def adjust_keyword_bids_by_threshold():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    entity_type = str(d.get("entity_type") or "adgroup").strip().lower()
    entity_ids = _unique_keep_order(d.get("entity_ids") or [])
    if not entity_ids:
        return jsonify({"error": "대상이 없습니다."}), 400
    def _parse_optional_bound(value: Any, label: str) -> Optional[int]:
        if str(value or "").strip() == "":
            return None
        bid = _normalize_bid_amt(value)
        if bid is None:
            raise ValueError(f"{label} 값을 확인해주세요.")
        return bid
    def _parse_signed_step(value: Any, label: str) -> int:
        s = str(value or "").strip().replace(",", "")
        if not s:
            raise ValueError(f"{label} 값을 입력해주세요.")
        try:
            return int(float(s))
        except Exception as exc:
            raise ValueError(f"{label} 값을 확인해주세요.") from exc
    try:
        upper_bid = _parse_optional_bound(d.get("upper_bid"), "상한가")
        lower_bid = _parse_optional_bound(d.get("lower_bid"), "하한가")
        upper_delta = _parse_signed_step(d.get("upper_delta"), "상한가 조정값") if upper_bid is not None else 0
        lower_delta = _parse_signed_step(d.get("lower_delta"), "하한가 조정값") if lower_bid is not None else 0
    except ValueError as exc:
        return jsonify({"error": str(exc)}), 400
    if upper_bid is None and lower_bid is None:
        return jsonify({"error": "상한가 또는 하한가 중 하나 이상 입력해주세요."}), 400
    if upper_bid is not None and lower_bid is not None and lower_bid >= upper_bid:
        return jsonify({"error": "하한가는 상한가보다 작아야 합니다."}), 400
    if upper_bid is not None and upper_delta == 0:
        return jsonify({"error": "상한가 조정값은 0이 될 수 없습니다."}), 400
    if lower_bid is not None and lower_delta == 0:
        return jsonify({"error": "하한가 조정값은 0이 될 수 없습니다."}), 400
    adgroup_contexts, resolve_warnings = _resolve_adgroup_contexts(api_key, secret_key, cid, entity_type, entity_ids)
    if not adgroup_contexts:
        return jsonify({"error": "대상 광고그룹을 찾지 못했습니다."}), 400
    keyword_meta: Dict[str, Dict[str, Any]] = {}
    keyword_bid_map: Dict[str, int] = {}
    ad_meta: Dict[str, Dict[str, Any]] = {}
    ad_bid_map: Dict[str, int] = {}
    warnings: List[str] = list(resolve_warnings)
    keyword_count = 0
    shopping_ad_count = 0
    upper_hit = 0
    lower_hit = 0
    unchanged_preview = 0
    keyword_groups = 0
    shopping_groups = 0
    for ctx in adgroup_contexts:
        adg_id = str(ctx.get("adgroup_id") or "").strip()
        adgroup_type = str(ctx.get("adgroup_type") or "").upper()
        adgroup_bid = ctx.get("adgroup_bid")
        if _adgroup_uses_ad_level_bid(adgroup_type):
            shopping_groups += 1
            res_ads, ads = _fetch_ads(api_key, secret_key, cid, adg_id)
            if res_ads.status_code != 200:
                warnings.append(f"[광고그룹 {adg_id}] 소재 조회 실패: {res_ads.text}")
                continue
            for ad in (ads or []):
                if not _ad_item_has_bid_attr(ad):
                    continue
                shopping_ad_count += 1
                ad_id = _extract_ad_id(ad)
                if not ad_id:
                    continue
                ad_attr = _extract_ad_attr(ad)
                current_use_group = bool(ad_attr.get("useGroupBidAmt"))
                effective_current_bid = _resolve_effective_bid(ad_attr.get("bidAmt"), current_use_group, adgroup_bid)
                if effective_current_bid is None:
                    continue
                ad_name = str(((ad.get("ad") or {}).get("productName") if isinstance(ad.get("ad"), dict) else "") or ad_id)
                ad_meta[ad_id] = {
                    "current_bid": effective_current_bid,
                    "use_group_bid": current_use_group,
                    "name": ad_name,
                    "adgroup_id": adg_id,
                }
                target_bid = None
                if upper_bid is not None and effective_current_bid >= upper_bid:
                    upper_hit += 1
                    target_bid = _normalize_bid_amt(upper_bid + upper_delta)
                elif lower_bid is not None and effective_current_bid <= lower_bid:
                    lower_hit += 1
                    target_bid = _normalize_bid_amt(lower_bid + lower_delta)
                if target_bid is None:
                    continue
                if target_bid == effective_current_bid and (not current_use_group):
                    unchanged_preview += 1
                    continue
                ad_bid_map[ad_id] = target_bid
        else:
            keyword_groups += 1
            r_kw = _do_req("GET", api_key, secret_key, cid, "/ncc/keywords", params={"nccAdgroupId": adg_id})
            if r_kw.status_code != 200:
                warnings.append(f"[광고그룹 {adg_id}] 키워드 조회 실패: {r_kw.text}")
                continue
            kws = r_kw.json() or []
            keyword_count += len(kws)
            for kw in kws:
                kid = str(kw.get("nccKeywordId") or "").strip()
                if not kid:
                    continue
                current_use_group = bool(kw.get("useGroupBidAmt"))
                effective_current_bid = _resolve_effective_bid(kw.get("bidAmt"), current_use_group, adgroup_bid)
                if effective_current_bid is None:
                    continue
                keyword_meta[kid] = {
                    "current_bid": effective_current_bid,
                    "use_group_bid": current_use_group,
                    "keyword": kw.get("keyword"),
                    "adgroup_id": adg_id,
                }
                target_bid = None
                if upper_bid is not None and effective_current_bid >= upper_bid:
                    upper_hit += 1
                    target_bid = _normalize_bid_amt(upper_bid + upper_delta)
                elif lower_bid is not None and effective_current_bid <= lower_bid:
                    lower_hit += 1
                    target_bid = _normalize_bid_amt(lower_bid + lower_delta)
                if target_bid is None:
                    continue
                if target_bid == effective_current_bid and (not current_use_group):
                    unchanged_preview += 1
                    continue
                keyword_bid_map[kid] = target_bid
    if not keyword_bid_map and not ad_bid_map:
        lines = [
            "조건에 맞는 키워드/소재가 없습니다.",
            f"파워링크 키워드 조회: {keyword_count}개 / 쇼핑 소재 조회: {shopping_ad_count}개 / 상한 조건 매칭: {upper_hit}개 / 하한 조건 매칭: {lower_hit}개",
        ]
        if warnings:
            lines.append("\n[참고]")
            lines.extend(warnings[:10])
        return jsonify({"ok": True, "message": "\n".join(lines)})
    kw_success = kw_fail = kw_skipped = 0
    ad_success = ad_fail = ad_skipped = 0
    err_details: List[str] = []
    if keyword_bid_map:
        kw_success, kw_fail, kw_skipped, kw_err_details = _apply_keyword_bid_map(
            api_key, secret_key, cid, [str(x.get("adgroup_id") or "") for x in adgroup_contexts if not _adgroup_uses_ad_level_bid(x.get("adgroup_type"))], keyword_bid_map, keyword_meta=keyword_meta
        )
        err_details.extend(kw_err_details)
    if ad_bid_map:
        ad_success, ad_fail, ad_skipped, ad_err_details = _apply_ad_bid_map(
            api_key, secret_key, cid, adgroup_contexts, ad_bid_map, ad_meta=ad_meta
        )
        err_details.extend(ad_err_details)
    stats = _calc_bid_stats(list(keyword_bid_map.values()) + list(ad_bid_map.values()))
    lines = [
        "상/하한 기준 입찰가 조정 완료!",
        f"파워링크/키워드 처리 광고그룹: {keyword_groups}개 | 쇼핑/소재 처리 광고그룹: {shopping_groups}개",
        f"파워링크 키워드 조회: {keyword_count}개 / 쇼핑 소재 조회: {shopping_ad_count}개 / 실제 변경 대상: {len(keyword_bid_map) + len(ad_bid_map)}개 / 동일해서 유지 예상: {unchanged_preview}개",
        f"상한 조건 매칭: {upper_hit}개 / 하한 조건 매칭: {lower_hit}개",
        f"키워드 변경 성공: {kw_success}개 / 실패: {kw_fail}개 / 유지/생략: {kw_skipped}개",
        f"쇼핑 소재 변경 성공: {ad_success}개 / 실패: {ad_fail}개 / 유지/생략: {ad_skipped}개",
    ]
    if stats.get("min") is not None:
        lines.append(f"변경 후 입찰가 범위: 최소 {stats['min']:,}원 · 중앙 {stats['median']:,}원 · 최대 {stats['max']:,}원")
    rule_bits = []
    if upper_bid is not None:
        upper_target = _normalize_bid_amt(upper_bid + upper_delta)
        rule_bits.append(f"상한 {upper_bid:,}원 이상 → {upper_bid:,}원 기준 {upper_delta:+,}원 = {upper_target:,}원")
    if lower_bid is not None:
        lower_target = _normalize_bid_amt(lower_bid + lower_delta)
        rule_bits.append(f"하한 {lower_bid:,}원 이하 → {lower_bid:,}원 기준 {lower_delta:+,}원 = {lower_target:,}원")
    if rule_bits:
        lines.append("적용 규칙: " + " / ".join(rule_bits))
    detail_lines = warnings + err_details
    if detail_lines:
        lines.append("\n[상세 내역]")
        lines.extend(detail_lines[:10])
    return jsonify({
        "ok": True,
        "message": "\n".join(lines),
        "stats": {
            "keyword_count": keyword_count,
            "shopping_ad_count": shopping_ad_count,
            "upper_hit": upper_hit,
            "lower_hit": lower_hit,
            "targets": len(keyword_bid_map) + len(ad_bid_map),
            "success": kw_success + ad_success,
            "fail": kw_fail + ad_fail,
            "unchanged_preview": unchanged_preview,
            "skipped": kw_skipped + ad_skipped,
        },
    })
@app.route("/update_keyword_bids_avg_position", methods=["POST"])
def update_keyword_bids_avg_position():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    entity_type = str(d.get("entity_type") or "adgroup").strip()
    entity_ids = d.get("entity_ids", []) or []
    device = str(d.get("device") or "PC").upper()
    position = int(d.get("position") or 1)
    preview_only = bool(d.get("preview_only"))
    include_paused = _boolish(d.get("include_paused"), False)
    include_pending = _boolish(d.get("include_pending"), False)
    if "exclude_paused" in d:
        include_paused = not _boolish(d.get("exclude_paused"), True)
    if "exclude_pending" in d:
        include_pending = not _boolish(d.get("exclude_pending"), True)
    max_bid_raw = d.get("max_bid")
    max_bid = None
    if str(max_bid_raw or "").strip() != "":
        try:
            max_bid = _normalize_bid_amt(max_bid_raw)
        except Exception:
            max_bid = None
    if device not in {"PC", "MOBILE"}:
        return jsonify({"error": "device는 PC 또는 MOBILE 이어야 합니다."}), 400
    max_position = 10 if device == "PC" else 5
    if position < 1 or position > max_position:
        return jsonify({"error": f"목표 평균순위는 {device} 기준 1~{max_position} 사이로 입력해주세요."}), 400
    if not entity_ids:
        return jsonify({"error": "대상이 없습니다."}), 400
    adgroup_contexts, resolve_warnings = _resolve_adgroup_contexts(api_key, secret_key, cid, entity_type, entity_ids)
    if not adgroup_contexts:
        msg = "적용 가능한 광고그룹을 찾지 못했습니다."
        if resolve_warnings:
            msg += "\n" + "\n".join(resolve_warnings[:5])
        return jsonify({"error": msg}), 400
    warnings: List[str] = list(resolve_warnings)
    keyword_adgroup_ids: List[str] = []
    shopping_contexts: List[Dict[str, Any]] = []
    skipped_other_groups = 0
    for ctx in adgroup_contexts:
        if _adgroup_uses_ad_level_bid(ctx.get("adgroup_type")):
            shopping_contexts.append(ctx)
        else:
            keyword_adgroup_ids.append(str(ctx.get("adgroup_id") or "").strip())
            if not str(ctx.get("adgroup_type") or "").upper() in {"WEB_SITE", ""}:
                skipped_other_groups += 1
    keyword_ids: List[str] = []
    keyword_meta: Dict[str, Dict[str, Any]] = {}
    keyword_count = 0
    keyword_fetch_fail = 0
    skipped_keyword_paused = 0
    skipped_keyword_pending = 0
    for adg_id in keyword_adgroup_ids:
        if not adg_id:
            continue
        r_kw = _do_req("GET", api_key, secret_key, cid, "/ncc/keywords", params={"nccAdgroupId": adg_id})
        if r_kw.status_code != 200:
            keyword_fetch_fail += 1
            if len(warnings) < 10:
                warnings.append(f"광고그룹 {adg_id} 키워드 조회 실패: {r_kw.text}")
            continue
        rows = r_kw.json() or []
        keyword_count += len(rows)
        for kw in rows:
            kid = str(kw.get("nccKeywordId") or "").strip()
            if not kid:
                continue
            ok, reason = _is_keyword_editable_for_avg_position(kw, include_paused=include_paused, include_pending=include_pending)
            if not ok:
                if str(reason).startswith("상태:"):
                    skipped_keyword_paused += 1
                else:
                    skipped_keyword_pending += 1
                continue
            keyword_ids.append(kid)
            keyword_meta[kid] = {
                "keyword": str(kw.get("keyword") or ""),
                "current_bid": kw.get("bidAmt"),
                "use_group_bid": bool(kw.get("useGroupBidAmt")),
                "adgroup_id": adg_id,
            }
    keyword_ids = _unique_keep_order(keyword_ids)
    ad_ids: List[str] = []
    ad_meta: Dict[str, Dict[str, Any]] = {}
    shopping_ad_count = 0
    shopping_fetch_fail = 0
    skipped_ad_paused = 0
    skipped_ad_pending = 0
    for ctx in shopping_contexts:
        adg_id = str(ctx.get("adgroup_id") or "").strip()
        adgroup_bid = ctx.get("adgroup_bid")
        res_ads, ads = _fetch_ads(api_key, secret_key, cid, adg_id)
        if res_ads.status_code != 200:
            shopping_fetch_fail += 1
            if len(warnings) < 10:
                warnings.append(f"광고그룹 {adg_id} 소재 조회 실패: {res_ads.text}")
            continue
        for ad in (ads or []):
            if not _ad_item_has_bid_attr(ad):
                continue
            shopping_ad_count += 1
            ad_id = _extract_ad_id(ad)
            if not ad_id:
                continue
            ok, reason = _is_ad_editable_for_avg_position(ad, include_paused=include_paused, include_pending=include_pending)
            if not ok:
                if str(reason).startswith("상태:"):
                    skipped_ad_paused += 1
                else:
                    skipped_ad_pending += 1
                continue
            ad_attr = _extract_ad_attr(ad)
            effective_current_bid = _resolve_effective_bid(ad_attr.get("bidAmt"), bool(ad_attr.get("useGroupBidAmt")), adgroup_bid)
            ad_name = str(((ad.get("ad") or {}).get("productName") if isinstance(ad.get("ad"), dict) else "") or ad_id)
            ad_ids.append(ad_id)
            ad_meta[ad_id] = {
                "name": ad_name,
                "current_bid": effective_current_bid,
                "use_group_bid": bool(ad_attr.get("useGroupBidAmt")),
                "adgroup_id": adg_id,
            }
    ad_ids = _unique_keep_order(ad_ids)
    if not keyword_ids and not ad_ids:
        return jsonify({"error": "평균순위 추정에 사용할 활성 키워드/소재가 없습니다."}), 400
    estimated_keyword_bid_map: Dict[str, int] = {}
    estimated_ad_bid_map: Dict[str, int] = {}
    if keyword_ids:
        estimated_keyword_bid_map, estimate_warnings = _estimate_keyword_bids_by_avg_position(api_key, secret_key, cid, keyword_ids, device, position, max_bid=max_bid)
        warnings.extend(estimate_warnings)
    if ad_ids:
        estimated_ad_bid_map, ad_estimate_warnings = _estimate_ad_bids_by_avg_position(api_key, secret_key, cid, ad_ids, device, position, max_bid=max_bid)
        warnings.extend(ad_estimate_warnings)
    if not estimated_keyword_bid_map and not estimated_ad_bid_map:
        msg = f"{device} {position}위 평균순위 추정값을 받지 못했습니다."
        if warnings:
            msg += "\n" + "\n".join(warnings[:5])
        return jsonify({"error": msg}), 400
    preview_rows = []
    changed_bids: List[int] = []
    unchanged_cnt = 0
    for kid in keyword_ids:
        if kid not in estimated_keyword_bid_map:
            continue
        meta = keyword_meta.get(kid, {})
        current_bid = _normalize_bid_amt(meta.get("current_bid"))
        new_bid = estimated_keyword_bid_map[kid]
        will_change = (current_bid != new_bid) or bool(meta.get("use_group_bid"))
        if will_change:
            changed_bids.append(new_bid)
        else:
            unchanged_cnt += 1
        if len(preview_rows) < 30:
            preview_rows.append({
                "target_type": "keyword",
                "name": meta.get("keyword") or kid,
                "keyword": meta.get("keyword") or kid,
                "keyword_id": kid,
                "current_bid": current_bid,
                "new_bid": new_bid,
                "use_group_bid": bool(meta.get("use_group_bid")),
                "will_change": will_change,
            })
    for ad_id in ad_ids:
        if ad_id not in estimated_ad_bid_map:
            continue
        meta = ad_meta.get(ad_id, {})
        current_bid = _normalize_bid_amt(meta.get("current_bid"))
        new_bid = estimated_ad_bid_map[ad_id]
        will_change = (current_bid != new_bid) or bool(meta.get("use_group_bid"))
        if will_change:
            changed_bids.append(new_bid)
        else:
            unchanged_cnt += 1
        if len(preview_rows) < 30:
            preview_rows.append({
                "target_type": "ad",
                "name": meta.get("name") or ad_id,
                "ad_id": ad_id,
                "current_bid": current_bid,
                "new_bid": new_bid,
                "use_group_bid": bool(meta.get("use_group_bid")),
                "will_change": will_change,
            })
    def _calc_stats(values: List[int]):
        if not values:
            return {"min": None, "max": None, "median": None}
        vals = sorted(values)
        n = len(vals)
        if n % 2:
            median = vals[n // 2]
        else:
            median = int(round((vals[n // 2 - 1] + vals[n // 2]) / 2.0 / 10.0) * 10)
        return {"min": vals[0], "max": vals[-1], "median": median}
    stats = _calc_stats(list(estimated_keyword_bid_map.values()) + list(estimated_ad_bid_map.values()))
    if preview_only:
        lines = [
            f"{device} 평균순위 {position}위 기준 변경사항 확인 완료",
            f"파워링크 추정 성공: {len(estimated_keyword_bid_map)}개 / 적용 대상 키워드: {len(keyword_ids)}개",
            f"쇼핑 추정 성공: {len(estimated_ad_bid_map)}개 / 적용 대상 소재: {len(ad_ids)}개",
            f"변경 예정: {len(changed_bids)}개 / 동일해서 유지: {unchanged_cnt}개",
        ]
        if stats["min"] is not None:
            lines.append(f"예상 입찰가 범위: 최소 {stats['min']:,}원 · 중앙 {stats['median']:,}원 · 최대 {stats['max']:,}원")
        if max_bid is not None:
            lines.append(f"최대 입찰가 상한 적용: {max_bid:,}원")
        if skipped_keyword_paused or skipped_ad_paused:
            lines.append(f"중지 상태 제외: {skipped_keyword_paused + skipped_ad_paused}개")
        if skipped_keyword_pending or skipped_ad_pending:
            lines.append(f"검수보류/비승인 제외: {skipped_keyword_pending + skipped_ad_pending}개")
        if skipped_other_groups:
            lines.append(f"기타 광고그룹 제외: {skipped_other_groups}개")
        for msg in warnings[:5]:
            lines.append(msg)
        return jsonify({
            "ok": True,
            "preview": True,
            "message": "\n".join(lines),
            "stats": stats,
            "rows": preview_rows,
            "estimated": len(estimated_keyword_bid_map) + len(estimated_ad_bid_map),
            "estimated_keywords": len(estimated_keyword_bid_map),
            "estimated_ads": len(estimated_ad_bid_map),
            "changed": len(changed_bids),
            "unchanged": unchanged_cnt,
            "skipped_paused": skipped_keyword_paused + skipped_ad_paused,
            "skipped_pending": skipped_keyword_pending + skipped_ad_pending,
        })
    kw_success = kw_fail = kw_skipped = 0
    ad_success = ad_fail = ad_skipped = 0
    err_details: List[str] = []
    if estimated_keyword_bid_map:
        kw_success, kw_fail, kw_skipped, kw_err_details = _apply_keyword_bid_map(api_key, secret_key, cid, keyword_adgroup_ids, estimated_keyword_bid_map, keyword_meta=keyword_meta)
        err_details.extend(kw_err_details)
    if estimated_ad_bid_map:
        ad_success, ad_fail, ad_skipped, ad_err_details = _apply_ad_bid_map(api_key, secret_key, cid, shopping_contexts, estimated_ad_bid_map, ad_meta=ad_meta)
        err_details.extend(ad_err_details)
    lines = [
        f"{device} 평균순위 {position}위 기준 입찰가 적용 완료!",
        f"파워링크 광고그룹: {len(keyword_adgroup_ids)}개 / 쇼핑 광고그룹: {len(shopping_contexts)}개",
        f"파워링크 추정 성공: {len(estimated_keyword_bid_map)}개 / 전체 조회 키워드: {keyword_count}개",
        f"쇼핑 추정 성공: {len(estimated_ad_bid_map)}개 / 전체 조회 소재: {shopping_ad_count}개",
        f"키워드 변경 성공: {kw_success}개 / 실패: {kw_fail}개 / 유지/생략: {kw_skipped}개",
        f"쇼핑 소재 변경 성공: {ad_success}개 / 실패: {ad_fail}개 / 유지/생략: {ad_skipped}개",
        f"동일해서 유지: {unchanged_cnt}개",
    ]
    if stats["min"] is not None:
        lines.append(f"예상 입찰가 범위: 최소 {stats['min']:,}원 · 중앙 {stats['median']:,}원 · 최대 {stats['max']:,}원")
    if max_bid is not None:
        lines.append(f"최대 입찰가 상한 적용: {max_bid:,}원")
    if skipped_other_groups:
        lines.append(f"기타 광고그룹 건너뜀: {skipped_other_groups}개")
    if keyword_fetch_fail:
        lines.append(f"키워드 조회 실패 광고그룹: {keyword_fetch_fail}개")
    if shopping_fetch_fail:
        lines.append(f"소재 조회 실패 광고그룹: {shopping_fetch_fail}개")
    if skipped_keyword_paused or skipped_ad_paused:
        lines.append(f"중지 상태 제외: {skipped_keyword_paused + skipped_ad_paused}개")
    if skipped_keyword_pending or skipped_ad_pending:
        lines.append(f"검수보류/비승인 제외: {skipped_keyword_pending + skipped_ad_pending}개")
    if kw_skipped + ad_skipped:
        lines.append(f"추정값 없음/변경 생략: {kw_skipped + ad_skipped}개")
    for msg in (warnings[:5] + err_details[:5]):
        lines.append(msg)
    return jsonify({
        "ok": True,
        "message": "\n".join(lines),
        "estimated": len(estimated_keyword_bid_map) + len(estimated_ad_bid_map),
        "estimated_keywords": len(estimated_keyword_bid_map),
        "estimated_ads": len(estimated_ad_bid_map),
        "success": kw_success + ad_success,
        "fail": kw_fail + ad_fail,
        "skipped": kw_skipped + ad_skipped,
        "unchanged": unchanged_cnt,
        "stats": stats,
    })
def _delete_payload_rows(api_key: str, secret_key: str, cid: str, entity_type: str, rows: List[Dict[str, Any]]):
    def _delete_one(idx: int, row: Dict[str, Any]):
        name = f"{idx}행"
        if entity_type in {"campaign", "adgroup", "keyword", "ad", "ad_extension"}:
            key_map = {
                "campaign": ["nccCampaignId", "캠페인ID", "campaign_id"],
                "adgroup": ["nccAdgroupId", "광고그룹ID", "adgroup_id"],
                "keyword": ["nccKeywordId", "키워드ID", "keyword_id"],
                "ad": ["nccAdId", "소재ID", "ad_id"],
                "ad_extension": ["adExtensionId", "확장소재ID", "ad_extension_id"],
            }
            entity_id = ""
            for k in key_map[entity_type]:
                entity_id = str(row.get(k) or "").strip()
                if entity_id:
                    break
            res = _delete_entity_by_id(api_key, secret_key, cid, entity_type, entity_id)
            name = entity_id or name
            ok = res.status_code in [200, 201, 204]
            return idx, ok, _result_item(idx, ok, name, "삭제 완료" if ok else getattr(res, 'text', '삭제 실패'))
        elif entity_type == "restricted_keyword":
            adgroup_id = str(row.get("nccAdgroupId") or row.get("광고그룹ID") or row.get("adgroup_id") or "").strip()
            keyword = str(row.get("keyword") or row.get("제외키워드") or "").strip()
            name = keyword or name
            if not adgroup_id or not keyword:
                return idx, False, _result_item(idx, False, name, "광고그룹ID / 제외키워드는 필수입니다.")
            res = _do_req("DELETE", api_key, secret_key, cid, "/ncc/restricted-keywords", params={"nccAdgroupId": adgroup_id, "keyword": keyword})
            ok = res.status_code in [200, 201, 204]
            return idx, ok, _result_item(idx, ok, name, "삭제 완료" if ok else getattr(res, 'text', '삭제 실패'))
        return idx, False, _result_item(idx, False, name, "지원하지 않는 삭제 유형")
    indexed_rows = list(enumerate(rows, start=1))
    if not indexed_rows:
        return 0, 0, []
    results_map: Dict[int, Dict[str, Any]] = {}
    success = fail = 0
    max_workers = min(DELETE_IO_WORKERS, max(1, len(indexed_rows)))
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        future_map = {ex.submit(_delete_one, idx, row): idx for idx, row in indexed_rows}
        for fut in as_completed(future_map):
            idx = future_map[fut]
            try:
                row_idx, ok, result_item = fut.result()
            except Exception as exc:
                row_idx = idx
                ok = False
                result_item = _result_item(idx, False, f"{idx}행", str(exc))
            results_map[row_idx] = result_item
            if ok:
                success += 1
            else:
                fail += 1
    results = [results_map[i] for i, _ in indexed_rows]
    return success, fail, results
def _normalize_bulk_extension_delete_type(value: Any) -> str:
    s = str(value or "").strip().upper()
    if not s or s == "ALL":
        return "ALL"
    if s == "SHOPPING_PROMO_TEXT":
        return "PROMOTION"
    if s == "쇼핑상품부가정보".upper():
        return "SHOPPING_EXTRA"
    return _normalize_extension_type(s) or s
def _bulk_extension_type_label(value: Any) -> str:
    normalized = _normalize_bulk_extension_delete_type(value)
    if normalized == "ALL":
        return "모든 확장소재"
    if str(value or "").strip().upper() == "SHOPPING_PROMO_TEXT":
        return "쇼핑 추가홍보문구"
    return AD_EXTENSION_TYPE_LABELS.get(normalized, normalized or "확장소재")
def _looks_like_shopping_ad(ad_item: Dict[str, Any] | None) -> bool:
    ad_type = str((ad_item or {}).get("type") or (ad_item or {}).get("adType") or "").upper()
    return ("SHOPPING" in ad_type) or ("CATALOG" in ad_type) or ("PRODUCT" in ad_type)
def _extract_ad_id(ad_item: Dict[str, Any] | None) -> str:
    return str((ad_item or {}).get("nccAdId") or (ad_item or {}).get("id") or (ad_item or {}).get("ownerId") or "").strip()
def _collect_target_adgroup_ids(api_key: str, secret_key: str, cid: str, parent_type: str, parent_ids: List[str]) -> Tuple[List[str], List[str]]:
    adgroup_ids: List[str] = []
    errors: List[str] = []
    seen = set()
    if parent_type == "campaign":
        normalized_campaign_ids = [str(x or "").strip() for x in (parent_ids or []) if str(x or "").strip()]
        if normalized_campaign_ids:
            max_workers = min(FAST_IO_WORKERS, max(1, len(normalized_campaign_ids)))
            with ThreadPoolExecutor(max_workers=max_workers) as ex:
                future_map = {ex.submit(_fetch_adgroups, api_key, secret_key, cid, campaign_id): campaign_id for campaign_id in normalized_campaign_ids}
                ordered_rows: Dict[str, List[Dict[str, Any]]] = {}
                for fut in as_completed(future_map):
                    campaign_id = future_map[fut]
                    try:
                        res, rows = fut.result()
                    except Exception as exc:
                        errors.append(f"[캠페인 {campaign_id}] 광고그룹 조회 실패: {exc}")
                        continue
                    if res.status_code != 200:
                        errors.append(f"[캠페인 {campaign_id}] 광고그룹 조회 실패: {res.text}")
                        continue
                    ordered_rows[campaign_id] = rows or []
            for campaign_id in normalized_campaign_ids:
                for row in ordered_rows.get(campaign_id, []):
                    adgroup_id = str(row.get("id") or row.get("nccAdgroupId") or "").strip()
                    if adgroup_id and adgroup_id not in seen:
                        seen.add(adgroup_id)
                        adgroup_ids.append(adgroup_id)
    else:
        for adgroup_id in parent_ids:
            adgroup_id = str(adgroup_id or "").strip()
            if adgroup_id and adgroup_id not in seen:
                seen.add(adgroup_id)
                adgroup_ids.append(adgroup_id)
    return adgroup_ids, errors
def _collect_extension_delete_rows(api_key: str, secret_key: str, cid: str, adgroup_ids: List[str], ext_type: str, campaign_ids: List[str] | None = None) -> Tuple[List[Dict[str, Any]], List[str]]:
    rows: List[Dict[str, Any]] = []
    errors: List[str] = []
    seen_ext_ids = set()
    requested_type = str(ext_type or "ALL").strip().upper()
    normalized_type = _normalize_bulk_extension_delete_type(requested_type)
    shopping_promo_only = requested_type == "SHOPPING_PROMO_TEXT"
    # 확장소재는 계정/생성 경로/API 응답 형태에 따라 campaign/adgroup/ad owner 기준으로 섞여 내려올 수 있다.
    # 특히 추가제목/홍보문구/추가설명이 ad owner 기준으로 존재하는 계정이 있어, 타입별로 owner 범위를 좁게 잡으면
    # "삭제할 추가제목이 없습니다"처럼 잘못 보일 수 있다. 삭제 전 조회 단계에서는 가능한 owner scope를 모두 확인하고,
    # adExtensionId 기준으로 dedupe 한다.
    include_campaign_owner = bool(campaign_ids)
    include_adgroup_owner = True
    include_ad_owner = True
    def _push_ext_items(items: List[Dict[str, Any]], owner_label: str):
        for item in (items or []):
            if not isinstance(item, dict):
                continue
            ext_id = str(item.get("adExtensionId") or item.get("id") or "").strip()
            if not ext_id or ext_id in seen_ext_ids:
                continue
            matched, resolved_type = _extension_item_matches_delete_type(item, requested_type)
            if shopping_promo_only:
                if resolved_type != "PROMOTION":
                    continue
            elif not matched:
                continue
            seen_ext_ids.add(ext_id)
            rows.append({
                "adExtensionId": ext_id,
                "_owner": owner_label,
                "_type": resolved_type or _normalize_bulk_extension_delete_type(item.get("type")) or requested_type,
                "_rawType": str(item.get("type") or ""),
            })
    if include_campaign_owner:
        seen_campaign_ids = set()
        for campaign_id in (campaign_ids or []):
            campaign_id = str(campaign_id or "").strip()
            if not campaign_id or campaign_id in seen_campaign_ids:
                continue
            seen_campaign_ids.add(campaign_id)
            res_ext, ext_items = _fetch_extensions(api_key, secret_key, cid, campaign_id)
            if res_ext.status_code == 200:
                _push_ext_items(ext_items, f"campaign:{campaign_id}")
            else:
                errors.append(f"[캠페인 {campaign_id}] 확장소재 조회 실패: {res_ext.text}")
    for adgroup_id in adgroup_ids:
        adgroup_id = str(adgroup_id or "").strip()
        if not adgroup_id:
            continue
        adgroup_obj = None
        res_detail, adgroup_obj = _fetch_adgroup_detail(api_key, secret_key, cid, adgroup_id)
        if res_detail.status_code != 200:
            errors.append(f"[광고그룹 {adgroup_id}] 상세 조회 실패: {res_detail.text}")
        if include_adgroup_owner:
            res_ext, ext_items = _fetch_extensions(api_key, secret_key, cid, adgroup_id)
            if res_ext.status_code == 200:
                _push_ext_items(ext_items, f"adgroup:{adgroup_id}")
            else:
                errors.append(f"[광고그룹 {adgroup_id}] 확장소재 조회 실패: {res_ext.text}")
        if include_ad_owner:
            res_ads, ads = _fetch_ads(api_key, secret_key, cid, adgroup_id)
            if res_ads.status_code != 200:
                errors.append(f"[광고그룹 {adgroup_id}] 소재 조회 실패: {res_ads.text}")
                continue
            adgroup_is_shopping = _is_shopping_adgroup(adgroup_obj)
            for ad_item in (ads or []):
                ad_id = _extract_ad_id(ad_item)
                if not ad_id:
                    continue
                # 쇼핑 외 타입도 ad owner 기준으로 확장소재가 매달린 경우가 있어 유형과 무관하게 확인한다.
                # 다만 ALL/쇼핑 계열은 기존처럼 그대로 포함되고, 파워링크 추가제목/홍보문구/추가설명도 여기서 잡힌다.
                if shopping_promo_only and not (adgroup_is_shopping or _looks_like_shopping_ad(ad_item)):
                    continue
                res_ad_ext, ad_ext_items = _fetch_extensions(api_key, secret_key, cid, ad_id)
                if res_ad_ext.status_code == 200:
                    _push_ext_items(ad_ext_items, f"ad:{ad_id}")
                else:
                    errors.append(f"[소재 {ad_id}] 확장소재 조회 실패: {res_ad_ext.text}")
    return rows, errors
@app.route("/bulk_delete_by_parent", methods=["POST"])
def bulk_delete_by_parent():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    parent_type = str(d.get("parent_type") or "campaign").strip()
    parent_ids = [str(x).strip() for x in (d.get("parent_ids") or []) if str(x).strip()]
    target_entity = str(d.get("target_entity") or "").strip()
    ext_type = str(d.get("ext_type") or "ALL").strip() or "ALL"
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API Key / Secret Key / Customer ID가 필요합니다."}), 400
    if parent_type not in {"campaign", "adgroup"}:
        return jsonify({"error": "지원하지 않는 부모 범위입니다."}), 400
    if target_entity not in {"keyword", "ad", "extension"}:
        return jsonify({"error": "지원하지 않는 삭제 대상입니다."}), 400
    if not parent_ids:
        return jsonify({"error": "선택된 대상이 없습니다."}), 400
    adgroup_ids, collect_errors = _collect_target_adgroup_ids(api_key, secret_key, cid, parent_type, parent_ids)
    if not adgroup_ids and collect_errors:
        return jsonify({"error": "하위 광고그룹 조회에 실패했습니다.", "details": collect_errors[:10]}), 400
    rows: List[Dict[str, Any]] = []
    entity_type = {"keyword": "keyword", "ad": "ad", "extension": "ad_extension"}[target_entity]
    seen = set()
    if target_entity == "keyword":
        if adgroup_ids:
            max_workers = min(FAST_IO_WORKERS, max(1, len(adgroup_ids)))
            with ThreadPoolExecutor(max_workers=max_workers) as ex:
                future_map = {ex.submit(_fetch_keywords, api_key, secret_key, cid, adgroup_id): adgroup_id for adgroup_id in adgroup_ids}
                ordered_items: Dict[str, List[Dict[str, Any]]] = {}
                for fut in as_completed(future_map):
                    adgroup_id = future_map[fut]
                    try:
                        res, items = fut.result()
                    except Exception as exc:
                        collect_errors.append(f"[광고그룹 {adgroup_id}] 키워드 조회 실패: {exc}")
                        continue
                    if res.status_code != 200:
                        collect_errors.append(f"[광고그룹 {adgroup_id}] 키워드 조회 실패: {res.text}")
                        continue
                    ordered_items[adgroup_id] = items or []
            for adgroup_id in adgroup_ids:
                for item in ordered_items.get(adgroup_id, []):
                    keyword_id = str((item or {}).get("nccKeywordId") or "").strip()
                    if keyword_id and keyword_id not in seen:
                        seen.add(keyword_id)
                        rows.append({"nccKeywordId": keyword_id})
    elif target_entity == "ad":
        for adgroup_id in adgroup_ids:
            res, items = _fetch_ads(api_key, secret_key, cid, adgroup_id)
            if res.status_code != 200:
                collect_errors.append(f"[광고그룹 {adgroup_id}] 소재 조회 실패: {res.text}")
                continue
            for item in (items or []):
                ad_id = _extract_ad_id(item)
                if ad_id and ad_id not in seen:
                    seen.add(ad_id)
                    rows.append({"nccAdId": ad_id})
    else:
        campaign_ids_for_ext = parent_ids if parent_type == "campaign" else []
        rows, ext_errors = _collect_extension_delete_rows(api_key, secret_key, cid, adgroup_ids, ext_type, campaign_ids=campaign_ids_for_ext)
        collect_errors.extend(ext_errors)
    if not rows:
        scope_label = "캠페인" if parent_type == "campaign" else "광고그룹"
        target_label = {
            "keyword": "키워드",
            "ad": "소재",
            "extension": _bulk_extension_type_label(ext_type),
        }[target_entity]
        msg = f"선택한 {scope_label} 범위에서 삭제할 {target_label}가 없습니다."
        if target_entity == "extension" and str(ext_type or "").strip().upper() == "HEADLINE":
            msg += "\n(캠페인/광고그룹/소재 owner 기준으로 확장소재를 모두 조회했고, type 누락 응답은 headline 필드 기준으로도 다시 판별했습니다.)"
        if str(ext_type or "").strip().upper() == "SHOPPING_PROMO_TEXT" and parent_type == "campaign":
            msg += "\n(캠페인/광고그룹/쇼핑소재 owner 기준으로 모두 조회했지만 대상이 발견되지 않았습니다.)"
        if collect_errors:
            msg += "\n" + "\n".join(collect_errors[:10])
        return jsonify({"ok": True, "total": 0, "success": 0, "fail": 0, "results": [], "message": msg})
    success, fail, results = _delete_payload_rows(api_key, secret_key, cid, entity_type, rows)
    msg_target = {
        "keyword": "키워드",
        "ad": "소재",
        "extension": _bulk_extension_type_label(ext_type),
    }[target_entity]
    msg = f"{msg_target} 일괄 삭제 완료 (대상 {len(rows)}건 / 성공 {success} / 실패 {fail})"
    if collect_errors:
        msg += "\n" + "\n".join(collect_errors[:10])
    return jsonify({
        "ok": True,
        "entity_type": entity_type,
        "total": len(rows),
        "success": success,
        "fail": fail,
        "results": results,
        "message": msg,
    })
@app.route("/bulk_register", methods=["POST"])
def bulk_register():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    entity_type = d.get("entity_type")
    rows = d.get("rows") or []
    raw_text = d.get("raw_text") or ""
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API Key / Secret Key / Customer ID가 필요합니다."}), 400
    if not rows and raw_text:
        rows = _read_table_text(raw_text)
    if not rows:
        return jsonify({"error": "등록할 데이터가 없습니다."}), 400
    handlers = {
        "campaign": _bulk_create_campaigns,
        "adgroup": _bulk_create_adgroups,
        "keyword": _bulk_create_keywords,
        "ad": _bulk_create_ads,
        "ad_extension": _bulk_create_extensions,
        "restricted_keyword": _bulk_create_restricted_keywords,
    }
    handler = handlers.get(entity_type)
    if not handler:
        return jsonify({"error": f"지원하지 않는 entity_type: {entity_type}"}), 400
    success, fail, results = handler(api_key, secret_key, cid, rows)
    return jsonify({"ok": True, "entity_type": entity_type, "total": len(rows), "success": success, "fail": fail, "results": results})
@app.route("/bulk_delete", methods=["POST"])
def bulk_delete():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    entity_type = d.get("entity_type")
    rows = d.get("rows") or []
    raw_text = d.get("raw_text") or ""
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API Key / Secret Key / Customer ID가 필요합니다."}), 400
    if not rows and raw_text:
        rows = _read_table_text(raw_text)
    if not rows:
        return jsonify({"error": "삭제할 데이터가 없습니다."}), 400
    success, fail, results = _delete_payload_rows(api_key, secret_key, cid, entity_type, rows)
    return jsonify({"ok": True, "entity_type": entity_type, "total": len(rows), "success": success, "fail": fail, "results": results})
def _set_user_lock_for_entity(api_key: str, secret_key: str, cid: str, entity_type: str, entity_id: str, enabled: bool) -> Tuple[bool, str]:
    entity_id = str(entity_id or "").strip()
    if not entity_id:
        return False, "ID가 비어 있습니다."
    uri = f"/ncc/campaigns/{entity_id}" if entity_type == "campaign" else f"/ncc/adgroups/{entity_id}"
    r_get = _do_req("GET", api_key, secret_key, cid, uri)
    if r_get.status_code != 200:
        return False, f"조회 실패: {r_get.text}"
    obj = r_get.json() or {}
    obj["userLock"] = not enabled
    r_put = _do_req("PUT", api_key, secret_key, cid, uri, params={"fields": "userLock"}, json_body=obj)
    if r_put.status_code in [200, 201]:
        return True, ""
    return False, f"변경 실패: {r_put.text}"
@app.route("/set_searched_powerlink_keyword_state", methods=["POST"])
def set_searched_powerlink_keyword_state():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    ids = d.get("ids") or []
    enabled = bool(d.get("enabled", True))
    if not api_key or not secret_key or not cid:
        return jsonify({"error": "API Key / Secret Key / Customer ID가 필요합니다."}), 400
    if not isinstance(ids, list) or not ids:
        return jsonify({"error": "상태를 변경할 키워드를 선택해주세요."}), 400
    normalized_ids: List[str] = []
    seen: set[str] = set()
    for item in ids:
        kid = str(item or "").strip()
        if not kid or kid in seen:
            continue
        seen.add(kid)
        normalized_ids.append(kid)
    if not normalized_ids:
        return jsonify({"error": "유효한 키워드 ID가 없습니다."}), 400
    success = 0
    fail = 0
    details: List[str] = []
    max_workers = min(max(1, BID_IO_WORKERS), max(1, len(normalized_ids)))
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        future_map = {ex.submit(_set_keyword_state, api_key, secret_key, cid, kid, enabled): kid for kid in normalized_ids}
        for fut in as_completed(future_map):
            kid = future_map[fut]
            try:
                ok, detail = fut.result()
            except Exception as exc:
                ok, detail = False, str(exc)
            if ok:
                success += 1
            else:
                fail += 1
                if detail and len(details) < 12:
                    details.append(f"[{kid}] {detail}")
    msg = f"조회 키워드 {'ON' if enabled else 'OFF'} 완료! (성공: {success} / 실패: {fail})"
    if details:
        msg += "\n" + "\n".join(details)
    _cache_invalidate(api_key, secret_key, cid)
    return jsonify({"ok": True, "message": msg, "success": success, "fail": fail, "enabled": enabled})
@app.route("/set_campaign_state", methods=["POST"])
def set_campaign_state():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    ids = d.get("ids") or []
    enabled = bool(d.get("enabled", True))
    if not ids:
        return jsonify({"error": "선택된 캠페인이 없습니다."}), 400
    success = fail = 0
    details: List[str] = []
    child_success = child_fail = 0
    for camp_id in ids:
        camp_id = str(camp_id or "").strip()
        ok, detail = _set_user_lock_for_entity(api_key, secret_key, cid, "campaign", camp_id, enabled)
        if ok:
            success += 1
        else:
            fail += 1
            if detail and len(details) < 8:
                details.append(f"[캠페인 {camp_id}] {detail}")
            continue
        if enabled:
            r_adg, rows = _fetch_adgroups(api_key, secret_key, cid, camp_id)
            if r_adg.status_code != 200:
                if len(details) < 8:
                    details.append(f"[캠페인 {camp_id}] 하위 광고그룹 조회 실패: {r_adg.text}")
                continue
            for row in rows:
                adg_id = str(row.get("id") or row.get("nccAdgroupId") or "").strip()
                if not adg_id:
                    continue
                ok_child, child_detail = _set_user_lock_for_entity(api_key, secret_key, cid, "adgroup", adg_id, enabled)
                if ok_child:
                    child_success += 1
                else:
                    child_fail += 1
                    if child_detail and len(details) < 8:
                        details.append(f"[광고그룹 {adg_id}] {child_detail}")
    msg = f"캠페인 {'ON' if enabled else 'OFF'} 완료! (캠페인 성공: {success} / 실패: {fail})"
    if enabled:
        msg += f" · 하위 광고그룹 ON 반영: 성공 {child_success} / 실패 {child_fail}"
    if details:
        msg += "\n" + "\n".join(details)
    _cache_invalidate(d.get("api_key"), d.get("secret_key"), d.get("customer_id"))
    return jsonify({"ok": True, "message": msg})
@app.route("/delete_selected", methods=["POST"])
def delete_selected():
    d = request.json or {}
    api_key, secret_key, cid = d.get("api_key"), d.get("secret_key"), d.get("customer_id")
    entity_type = d.get("entity_type")
    ids = d.get("ids") or []
    rows = []
    key_name = {"campaign": "nccCampaignId", "adgroup": "nccAdgroupId", "keyword": "nccKeywordId", "ad": "nccAdId", "ad_extension": "adExtensionId"}.get(entity_type)
    if not key_name:
        return jsonify({"error": "지원하지 않는 선택삭제 유형입니다."}), 400
    for x in ids:
        rows.append({key_name: x})
    success, fail, results = _delete_payload_rows(api_key, secret_key, cid, entity_type, rows)
    return jsonify({"ok": True, "total": len(rows), "success": success, "fail": fail, "results": results})
@app.route("/sample_headers", methods=["GET"])
def sample_headers():
    entity_type = request.args.get("entity_type", "campaign")
    headers = ENTITY_SAMPLE_HEADERS.get(entity_type, [])
    sample_row = {h: "" for h in headers}
    if entity_type == "campaign":
        sample_row.update({"캠페인명": "브랜드_검색", "캠페인유형": "파워링크", "일예산사용": "예", "일예산": "50000"})
    elif entity_type == "adgroup":
        sample_row.update({"캠페인ID": "cmp-a001", "광고그룹명": "브랜드_기본", "광고그룹유형": "파워링크", "일예산사용": "예", "일예산": "10000", "입찰가": "100", "비즈채널ID": "bsn-a001-..."})
    elif entity_type == "keyword":
        sample_row.update({"광고그룹ID": "grp-a001", "키워드": "브랜드키워드", "그룹입찰가사용": "아니오", "입찰가": "120", "사용자잠금": "아니오"})
    elif entity_type == "ad":
        sample_row.update({"광고그룹ID": "grp-a001", "소재유형": "기본소재", "제목": "브랜드 공식 상담", "설명": "빠른 견적 상담 안내", "PC랜딩URL": "https://example.com", "모바일랜딩URL": "https://m.example.com", "사용자잠금": "아니오"})
    elif entity_type == "ad_extension":
        sample_row.update({"소유ID": "grp-a001", "확장소재유형": "서브링크", "원본JSON": '{"type":"SUB_LINKS","ownerId":"grp-a001","links":[{"title":"상담문의","pc":{"final":"https://example.com"},"mobile":{"final":"https://m.example.com"}}]}'})
    elif entity_type == "restricted_keyword":
        sample_row.update({"광고그룹ID": "grp-a001", "제외키워드": "무료"})
    return jsonify({"headers": headers, "sample_row": sample_row})
@app.route("/delete_sample_headers", methods=["GET"])
def delete_sample_headers():
    entity_type = request.args.get("entity_type", "campaign")
    headers = DELETE_SAMPLE_HEADERS.get(entity_type, [])
    sample_row = {h: "" for h in headers}
    if headers:
        sample_row[headers[0]] = "example-id"
    return jsonify({"headers": headers, "sample_row": sample_row})
if __name__ == "__main__":
    os.makedirs(SAMPLES_DIR, exist_ok=True)
    app.run(debug=True, port=5000)
