# -*- coding: utf-8 -*-
"""Excel export helpers.

This module keeps workbook styling/serialization in one place so each export
route can focus on collecting rows and defining columns.
"""
from __future__ import annotations

import io
from typing import Any, Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

DEFAULT_TITLE_FILL = "1D4ED8"
DEFAULT_HEADER_FILL = "111827"


def safe_sheet_title(title: str, fallback: str = "Sheet1") -> str:
    value = str(title or fallback).strip() or fallback
    for ch in ["\\", "/", "?", "*", "[", "]", ":"]:
        value = value.replace(ch, " ")
    value = " ".join(value.split())
    return value[:31] or fallback


def new_workbook(sheet_title: str = "Sheet1") -> tuple[Workbook, Any]:
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title(sheet_title)
    return wb, ws


def workbook_to_bytesio(wb: Workbook) -> io.BytesIO:
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def write_title_block(
    ws: Any,
    title: str,
    *,
    last_col: int = 1,
    metadata: Sequence[str] | None = None,
    title_fill: str = DEFAULT_TITLE_FILL,
) -> None:
    last_col = max(1, int(last_col or 1))
    ws.cell(row=1, column=1, value=title)
    title_cell = ws.cell(row=1, column=1)
    title_cell.font = Font(size=14, bold=True, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=title_fill)
    title_cell.alignment = Alignment(vertical="center")
    if last_col > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)

    for offset, line in enumerate(metadata or [], start=2):
        ws.cell(row=offset, column=1, value=str(line or ""))


def _width_for(widths: Mapping[Any, Any] | None, col_idx: int, header: str) -> float:
    if not widths:
        return max(12, min(40, len(str(header or "")) + 4))
    letter = get_column_letter(col_idx)
    value = widths.get(col_idx)
    if value is None:
        value = widths.get(letter)
    if value is None:
        value = widths.get(str(col_idx))
    if value is None:
        return max(12, min(40, len(str(header or "")) + 4))
    try:
        return float(value)
    except Exception:
        return max(12, min(40, len(str(header or "")) + 4))


def write_table(
    ws: Any,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    *,
    start_row: int = 1,
    widths: Mapping[Any, Any] | None = None,
    header_fill: str = DEFAULT_HEADER_FILL,
    header_font_color: str | None = "FFFFFF",
    header_font_bold: bool = True,
    freeze_panes: str | None = None,
    auto_filter: bool = False,
    wrap_text: bool = True,
) -> None:
    headers = list(headers or [])
    start_row = int(start_row or 1)
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.font = Font(bold=header_font_bold, color=header_font_color) if header_font_color else Font(bold=header_font_bold)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    written = 0
    for row_idx, values in enumerate(rows or [], start=start_row + 1):
        written += 1
        for col_idx, value in enumerate(list(values), start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    for col_idx, header in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = _width_for(widths, col_idx, str(header))

    end_row = start_row + max(1, written)
    max_col = max(1, len(headers))
    if wrap_text:
        for row_cells in ws.iter_rows(min_row=start_row + 1, max_row=end_row, min_col=1, max_col=max_col):
            for cell in row_cells:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    if freeze_panes:
        ws.freeze_panes = freeze_panes
    if auto_filter and headers:
        ws.auto_filter.ref = ws.dimensions


def build_report_workbook(
    *,
    title: str,
    sheet_title: str | None = None,
    metadata: Sequence[str] | None = None,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    start_row: int,
    widths: Mapping[Any, Any] | None = None,
    freeze_panes: str | None = None,
    auto_filter: bool = False,
    title_fill: str = DEFAULT_TITLE_FILL,
    header_fill: str = DEFAULT_HEADER_FILL,
) -> Workbook:
    wb, ws = new_workbook(sheet_title or title)
    write_title_block(ws, title, last_col=len(headers), metadata=metadata, title_fill=title_fill)
    write_table(
        ws,
        headers,
        rows,
        start_row=start_row,
        widths=widths,
        header_fill=header_fill,
        header_font_color="FFFFFF",
        freeze_panes=freeze_panes,
        auto_filter=auto_filter,
    )
    return wb


def build_table_workbook(
    *,
    sheet_title: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    widths: Mapping[Any, Any] | None = None,
    freeze_panes: str | None = None,
    auto_filter: bool = True,
    header_fill: str = "E8F0FE",
    header_font_color: str | None = None,
) -> Workbook:
    wb, ws = new_workbook(sheet_title)
    write_table(
        ws,
        headers,
        rows,
        start_row=1,
        widths=widths,
        header_fill=header_fill,
        header_font_color=header_font_color,
        freeze_panes=freeze_panes,
        auto_filter=auto_filter,
    )
    return wb
