# -*- coding: utf-8 -*-
from __future__ import annotations

import json
import time
from collections import defaultdict
from typing import Any, Callable, Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.api_response import api_error, file_payload


PURCHASE_STAT_FIELDS = [
    "impCnt",
    "clkCnt",
    "salesAmt",
    "ccnt",
    "convAmt",
    "purchaseCcnt",
    "purchaseConvAmt",
    "purchaseRor",
]


def _number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except Exception:
        return 0.0


def _intish(value: Any) -> int:
    return int(round(_number(value)))


def _campaign_id(row: Dict[str, Any]) -> str:
    return str((row or {}).get("nccCampaignId") or (row or {}).get("id") or "").strip()


def _campaign_type_code(row: Dict[str, Any]) -> str:
    return str((row or {}).get("campaignTp") or (row or {}).get("campaignType") or "").strip().upper()


def _campaign_type_label(value: Any) -> str:
    code = str(value or "").strip().upper()
    if code in {"WEB_SITE", "POWERLINK"}:
        return "파워링크"
    if code in {"SHOPPING", "SHOPPING_PRODUCT"}:
        return "쇼핑검색"
    return "기타"


def _collect_stat_rows_from_payload(payload: Any) -> List[Dict[str, Any]]:
    if isinstance(payload, dict):
        direct = payload.get("data")
        if isinstance(direct, list):
            return [x for x in direct if isinstance(x, dict)]
        summary = payload.get("summaryStatResponse")
        if isinstance(summary, dict) and isinstance(summary.get("data"), list):
            return [x for x in summary.get("data") if isinstance(x, dict)]
    return []


def _format_api_time(value: Any) -> str:
    text = str(value or "").strip()
    if len(text) == 12 and text.isdigit():
        return f"{text[0:4]}-{text[4:6]}-{text[6:8]} {text[8:10]}:{text[10:12]}"
    return text


class PurchaseReportService:
    """Current purchase-complete report based on Naver SearchAd /stats."""

    def __init__(
        self,
        *,
        request_func: Callable[..., Any],
        workbook_to_bytesio_func: Callable[..., Any],
        xlsx_mime: str,
    ) -> None:
        self.request = request_func
        self.workbook_to_bytesio = workbook_to_bytesio_func
        self.xlsx_mime = xlsx_mime

    @staticmethod
    def _credentials(payload: Dict[str, Any]) -> Tuple[str, str, str]:
        d = payload or {}
        return (
            str(d.get("api_key") or "").strip(),
            str(d.get("secret_key") or "").strip(),
            str(d.get("customer_id") or "").strip(),
        )

    @staticmethod
    def _account_name(payload: Dict[str, Any], cid: str) -> str:
        return str((payload or {}).get("account_name") or (payload or {}).get("accountName") or cid or "").strip()

    def _fetch_campaigns(self, api_key: str, secret_key: str, cid: str):
        res = self.request("GET", api_key, secret_key, cid, "/ncc/campaigns")
        if getattr(res, "status_code", 0) != 200:
            return res, []
        try:
            rows = res.json() or []
        except Exception:
            rows = []
        if not isinstance(rows, list):
            rows = []
        return res, [x for x in rows if isinstance(x, dict) and not bool(x.get("delFlag"))]

    def _fetch_today_stats(self, api_key: str, secret_key: str, cid: str, campaign_ids: List[str]):
        rows: List[Dict[str, Any]] = []
        errors: List[str] = []
        cycle_base_tm = ""
        comp_tm = ""
        for start in range(0, len(campaign_ids), 20):
            batch = [x for x in campaign_ids[start:start + 20] if x]
            if not batch:
                continue
            params = {
                "ids": ",".join(batch),
                "fields": json.dumps(PURCHASE_STAT_FIELDS, separators=(",", ":")),
                "datePreset": "today",
                "timeIncrement": "allDays",
            }
            res = self.request("GET", api_key, secret_key, cid, "/stats", params=params)
            if getattr(res, "status_code", 0) != 200:
                errors.append(str(getattr(res, "text", "") or f"/stats {getattr(res, 'status_code', '')}")[:500])
                continue
            try:
                payload = res.json() or {}
            except Exception:
                payload = {}
            if isinstance(payload, dict):
                cycle_base_tm = cycle_base_tm or str(payload.get("cycleBaseTm") or "")
                comp_tm = comp_tm or str(payload.get("compTm") or "")
            rows.extend(_collect_stat_rows_from_payload(payload))
        return rows, errors, cycle_base_tm, comp_tm

    def collect_current(self, payload: Dict[str, Any]):
        api_key, secret_key, cid = self._credentials(payload)
        if not api_key or not secret_key or not cid:
            return api_error("API 정보 및 광고주를 선택해주세요."), 400
        account_name = self._account_name(payload, cid)
        res_camp, campaigns = self._fetch_campaigns(api_key, secret_key, cid)
        if getattr(res_camp, "status_code", 0) != 200:
            return api_error("캠페인 조회 실패", getattr(res_camp, "text", "")), 400
        campaign_map = {_campaign_id(row): row for row in campaigns if _campaign_id(row)}
        stat_rows, errors, cycle_base_tm, comp_tm = self._fetch_today_stats(api_key, secret_key, cid, list(campaign_map.keys()))
        stat_by_campaign = {str(row.get("id") or "").strip(): row for row in stat_rows if str(row.get("id") or "").strip()}

        detail_rows: List[Dict[str, Any]] = []
        summary_map: Dict[str, Dict[str, Any]] = defaultdict(lambda: {
            "campaign_count": 0,
            "impCnt": 0.0,
            "clkCnt": 0.0,
            "salesAmt": 0.0,
            "ccnt": 0.0,
            "convAmt": 0.0,
            "purchaseCcnt": 0.0,
            "purchaseConvAmt": 0.0,
        })
        today = time.strftime("%Y-%m-%d")
        for campaign_id, stat in stat_by_campaign.items():
            campaign = campaign_map.get(campaign_id) or {}
            type_code = _campaign_type_code(campaign)
            type_label = _campaign_type_label(type_code)
            row = {
                "date": today,
                "account_name": account_name,
                "customer_id": cid,
                "campaign_type": type_label,
                "campaign_type_code": type_code,
                "campaign_name": str(campaign.get("name") or campaign_id),
                "campaign_id": campaign_id,
                "impCnt": _intish(stat.get("impCnt")),
                "clkCnt": _intish(stat.get("clkCnt")),
                "salesAmt": _intish(stat.get("salesAmt")),
                "ccnt": _number(stat.get("ccnt")),
                "convAmt": _intish(stat.get("convAmt")),
                "purchaseCcnt": _number(stat.get("purchaseCcnt")),
                "purchaseConvAmt": _intish(stat.get("purchaseConvAmt")),
                "purchaseRor": _number(stat.get("purchaseRor")),
            }
            detail_rows.append(row)
            bucket = summary_map[type_label]
            bucket["campaign_count"] += 1
            for key in ("impCnt", "clkCnt", "salesAmt", "ccnt", "convAmt", "purchaseCcnt", "purchaseConvAmt"):
                bucket[key] += _number(row.get(key))

        summary_rows: List[Dict[str, Any]] = []
        for label in ("파워링크", "쇼핑검색", "기타"):
            if label not in summary_map:
                continue
            item = summary_map[label]
            purchase_ror = (item["purchaseConvAmt"] / item["salesAmt"] * 100) if item["salesAmt"] else 0
            summary_rows.append({
                "campaign_type": label,
                "campaign_count": int(item["campaign_count"]),
                "impCnt": int(item["impCnt"]),
                "clkCnt": int(item["clkCnt"]),
                "salesAmt": int(item["salesAmt"]),
                "ccnt": round(item["ccnt"], 2),
                "convAmt": int(item["convAmt"]),
                "purchaseCcnt": round(item["purchaseCcnt"], 2),
                "purchaseConvAmt": int(item["purchaseConvAmt"]),
                "purchaseRor": round(purchase_ror, 2),
            })

        detail_rows.sort(key=lambda row: (row.get("campaign_type") or "", -_number(row.get("purchaseConvAmt")), row.get("campaign_name") or ""))
        total_purchase_count = round(sum(_number(x.get("purchaseCcnt")) for x in summary_rows), 2)
        total_purchase_amount = int(sum(_number(x.get("purchaseConvAmt")) for x in summary_rows))
        return {
            "ok": True,
            "message": f"오늘 현시점 구매완료 조회 완료: {total_purchase_count:g}건 / {total_purchase_amount:,}원",
            "date": today,
            "account_name": account_name,
            "customer_id": cid,
            "cycleBaseTm": cycle_base_tm,
            "cycleBaseTmLabel": _format_api_time(cycle_base_tm),
            "compTm": comp_tm,
            "compTmLabel": _format_api_time(comp_tm),
            "summary": summary_rows,
            "rows": detail_rows,
            "warnings": [],
            "errors": errors[:20],
        }, 200

    @staticmethod
    def _style_sheet(ws: Any, header_row: int = 1) -> None:
        for cell in ws[header_row]:
            cell.fill = PatternFill("solid", fgColor="111827")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.freeze_panes = f"A{header_row + 1}"
        ws.auto_filter.ref = ws.dimensions
        for idx, cells in enumerate(ws.iter_cols(), start=1):
            width = min(max(max(len(str(cell.value or "")) for cell in cells) + 2, 10), 48)
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _build_workbook(self, result: Dict[str, Any]) -> Workbook:
        wb = Workbook()
        ws = wb.active
        ws.title = "유형별 요약"
        ws.append(["계정", result.get("account_name"), "광고주ID", result.get("customer_id")])
        ws.append(["기준", "오늘 현시점", "cycleBaseTm", result.get("cycleBaseTmLabel") or result.get("cycleBaseTm"), "compTm", result.get("compTmLabel") or result.get("compTm")])
        ws.append([])
        ws.append(["광고유형", "성과발생 캠페인수", "노출수", "클릭수", "광고비(VAT포함)", "전체전환수", "전체전환금액", "구매완료수", "구매완료금액", "구매ROAS(%)"])
        for row in result.get("summary") or []:
            ws.append([
                row.get("campaign_type"),
                row.get("campaign_count"),
                row.get("impCnt"),
                row.get("clkCnt"),
                row.get("salesAmt"),
                row.get("ccnt"),
                row.get("convAmt"),
                row.get("purchaseCcnt"),
                row.get("purchaseConvAmt"),
                row.get("purchaseRor"),
            ])
        self._style_sheet(ws, header_row=4)

        detail = wb.create_sheet("캠페인 상세")
        detail.append(["일자", "계정명", "광고주ID", "광고유형", "캠페인명", "캠페인ID", "캠페인유형코드", "노출수", "클릭수", "광고비(VAT포함)", "전체전환수", "전체전환금액", "구매완료수", "구매완료금액", "구매ROAS(%)"])
        for row in result.get("rows") or []:
            detail.append([
                row.get("date"),
                row.get("account_name"),
                row.get("customer_id"),
                row.get("campaign_type"),
                row.get("campaign_name"),
                row.get("campaign_id"),
                row.get("campaign_type_code"),
                row.get("impCnt"),
                row.get("clkCnt"),
                row.get("salesAmt"),
                row.get("ccnt"),
                row.get("convAmt"),
                row.get("purchaseCcnt"),
                row.get("purchaseConvAmt"),
                row.get("purchaseRor"),
            ])
        self._style_sheet(detail, header_row=1)
        return wb

    def export_current_excel(self, payload: Dict[str, Any]):
        result, status = self.collect_current(payload)
        if status != 200:
            return result, status
        wb = self._build_workbook(result)
        stamp = time.strftime("%Y%m%d_%H%M%S")
        safe_account = "".join(ch if ch.isalnum() or ch in ("_", "-") else "_" for ch in str(result.get("account_name") or "account"))
        output = self.workbook_to_bytesio(wb)
        return file_payload(
            output,
            mimetype=self.xlsx_mime,
            download_name=f"{safe_account}_{result.get('customer_id')}_purchase_current_{stamp}.xlsx",
        ), 200
